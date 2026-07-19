#!/usr/bin/env python3
"""JSON API for evaluation artifacts."""

from __future__ import annotations

import csv
import copy
import io
import json
import mimetypes
import os
import pickle
import pickletools
import re
import shutil
import subprocess
import sys
from ast import literal_eval
from collections import Counter
from functools import lru_cache
from html.parser import HTMLParser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
INDEX_PATH = DATA / ".artifact-viewer-index.json"
WEB_INDEX_PATH = ROOT / "web" / "src" / "qa-index.json"
WEB_SIFLEX_INDEX_PATH = ROOT / "web" / "src" / "siflex-index.json"
WEB_TABULAR_MODELS_INDEX_PATH = ROOT / "web" / "src" / "tabular-models-index.json"
WEB_MISMATCH_INDEX_PATH = ROOT / "web" / "src" / "mismatch-index.json"
SIFLEX_ROOT = DATA / "Log_tabAgent_Siflex"
TABULAR_MODELS_ROOT = DATA / "Log_Tabular_Models"
DATALAKE_CHALLENGE_ROOT = DATA / "Datasets" / "DataLake-challenge"
INDEX_VERSION = 6
QUESTION_KEYS = ("question", "query", "prompt", "qa", "input")
GOLD_KEYS = ("gold", "golden", "answer", "reference", "ground_truth", "target", "label")
PRED_KEYS = ("prediction", "predicted", "pred", "response", "output", "model_answer", "model_output")
csv.field_size_limit(10 * 1024 * 1024)
MAX_PICKLE_BYTES = 32 * 1024 * 1024


class SafeFeatureTree:
    """Inert shell used to read ST-Raptor state without importing its code."""


class SafeIndexTree:
    pass


class SafeIndexNode:
    pass


class SafeBodyNode:
    pass


class SafeBodyTree:
    pass


ST_RAPTOR_PICKLE_CLASSES = {
    ("table2tree.feature_tree", "FeatureTree"): SafeFeatureTree,
    ("table2tree.feature_tree", "IndexTree"): SafeIndexTree,
    ("table2tree.feature_tree", "IndexNode"): SafeIndexNode,
    ("table2tree.feature_tree", "BodyNode"): SafeBodyNode,
    ("table2tree.feature_tree", "BodyTree"): SafeBodyTree,
}


class RestrictedPickleLoader(pickle.Unpickler):
    """Load NumPy caches while rejecting arbitrary pickle globals."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.class_references = []

    def find_class(self, module: str, name: str):
        reference = f"{module}.{name}"
        if reference not in self.class_references:
            self.class_references.append(reference)
        if (module, name) in ST_RAPTOR_PICKLE_CLASSES:
            return ST_RAPTOR_PICKLE_CLASSES[(module, name)]
        if module in {"numpy.core.multiarray", "numpy._core.multiarray"} and name in {"_reconstruct", "scalar"}:
            import importlib
            return getattr(importlib.import_module(module), name)
        if module == "numpy" and name in {"ndarray", "dtype"}:
            import numpy as np
            return getattr(np, name)
        raise pickle.UnpicklingError(f"Unsupported pickle class: {module}.{name}")

    def persistent_load(self, _):
        raise pickle.UnpicklingError("Persistent pickle references are not supported")


def static_pickle_summary(data: bytes) -> dict:
    opcode_counts = Counter()
    strings = []
    class_references = []
    string_stack = []
    try:
        for opcode, argument, _ in pickletools.genops(data):
            opcode_counts[opcode.name] += 1
            if opcode.name == "GLOBAL":
                class_references.append(str(argument).replace(" ", "."))
            elif opcode.name in {"SHORT_BINUNICODE", "BINUNICODE", "UNICODE"}:
                text = str(argument)
                string_stack.append(text)
                if len(string_stack) > 4:
                    string_stack.pop(0)
                if len(strings) < 120 and len(text) <= 240 and text not in strings:
                    strings.append(text)
            elif opcode.name == "STACK_GLOBAL" and len(string_stack) >= 2:
                reference = f"{string_stack[-2]}.{string_stack[-1]}"
                if reference not in class_references:
                    class_references.append(reference)
    except (pickle.UnpicklingError, ValueError, EOFError) as error:
        return {"error": f"Could not inspect pickle opcodes: {error}"}
    technical_strings = {
        "FeatureTree", "IndexTree", "IndexNode", "BodyNode", "BodyTree",
        "index_tree", "body_tree",
        "root", "value", "children", "table", "body", "father",
        "x1", "y1", "x2", "y2", "group_name", "group_id", "group_class",
        "group_type", "group_name_list", "group_id_list", "name2id", "id2name",
        "example_dict", "leaf_nodes", "none", "None",
    }
    content_preview = []
    for text in strings:
        cleaned = text.strip()
        if (
            cleaned in technical_strings
            or "." in cleaned and cleaned.split(".", 1)[0] in {"table2tree", "index_tree", "body_tree"}
            or re.fullmatch(r"[-+]?\d+(?:\.\d+)?", cleaned)
            or cleaned.startswith("[")
            or len(cleaned) < 2
        ):
            continue
        if cleaned not in content_preview:
            content_preview.append(cleaned)
    return {
        "classReferences": class_references[:20],
        "stringPreview": strings[:80],
        "contentPreview": content_preview[:24],
        "opcodeCounts": dict(opcode_counts.most_common(16)),
    }


def array_summary(array) -> dict:
    import numpy as np

    matrix = np.asarray(array)
    flattened = matrix.reshape(matrix.shape[0], -1) if matrix.ndim > 1 else matrix.reshape(1, -1)
    preview = flattened[:8, :12]
    summary = {
        "type": type(array).__name__,
        "shape": list(matrix.shape),
        "dtype": str(matrix.dtype),
        "preview": preview.tolist(),
    }
    if np.issubdtype(matrix.dtype, np.number):
        summary["preview"] = [[float(value) if np.isfinite(value) else None for value in row] for row in preview]
        numeric = matrix.astype(float, copy=False)
        finite = numeric[np.isfinite(numeric)]
        if finite.size:
            summary["stats"] = {
                "min": float(finite.min()),
                "max": float(finite.max()),
                "mean": float(finite.mean()),
                "std": float(finite.std()),
            }
        if flattened.shape[1] and np.issubdtype(flattened.dtype, np.number):
            norms = np.linalg.norm(flattened.astype(float, copy=False), axis=1)
            summary["norms"] = {
                "min": float(norms.min()),
                "max": float(norms.max()),
                "mean": float(norms.mean()),
                "preview": [float(value) for value in norms[:1024]],
            }
    return summary


def safe_value_preview(value, depth: int = 0):
    if depth >= 2:
        return {"type": type(value).__name__}
    try:
        import numpy as np
        if isinstance(value, np.ndarray):
            return array_summary(value)
    except ImportError:
        pass
    if isinstance(value, dict):
        return {
            "type": "dict",
            "length": len(value),
            "items": [{"key": str(key)[:120], "value": safe_value_preview(item, depth + 1)} for key, item in list(value.items())[:12]],
        }
    if isinstance(value, (list, tuple)):
        return {
            "type": type(value).__name__,
            "length": len(value),
            "items": [safe_value_preview(item, depth + 1) for item in list(value)[:12]],
        }
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value if not isinstance(value, str) else value[:240]
    return {"type": type(value).__name__, "preview": repr(value)[:240]}


def compact_table_value(value, limit: int = 180):
    if isinstance(value, SafeFeatureTree):
        return "Nested table"
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    return text if len(text) <= limit else f"{text[:limit - 1]}..."


def feature_tree_leaves(feature_tree) -> list[tuple[object, list[str]]]:
    index_tree = getattr(feature_tree, "index_tree", None)
    root = getattr(index_tree, "root", None)
    leaves = []
    visited = set()

    def walk(node, path, depth=0):
        if node is None or id(node) in visited or depth > 24 or len(visited) >= 400:
            return
        visited.add(id(node))
        value = compact_table_value(getattr(node, "value", None), 80)
        next_path = path + ([str(value)] if value not in {None, ""} else [])
        children = list(getattr(node, "children", None) or [])[:80]
        if children:
            for child in children:
                walk(child, next_path, depth + 1)
        elif node is not root:
            leaves.append((node, next_path or ["Unnamed column"]))

    for child in list(getattr(root, "children", None) or [])[:80]:
        walk(child, [])
    if not leaves:
        for node in list(getattr(index_tree, "leaf_nodes", None) or [])[:80]:
            value = compact_table_value(getattr(node, "value", None), 80)
            leaves.append((node, [str(value) if value not in {None, ""} else "Unnamed column"]))
    return leaves


def index_tree_preview(feature_tree) -> list[dict]:
    root = getattr(getattr(feature_tree, "index_tree", None), "root", None)
    visited = set()

    def convert(node, depth=0):
        if node is None or id(node) in visited or depth > 10 or len(visited) >= 160:
            return None
        visited.add(id(node))
        children = []
        for child in list(getattr(node, "children", None) or [])[:24]:
            converted = convert(child, depth + 1)
            if converted:
                children.append(converted)
        value = compact_table_value(getattr(node, "value", None), 90)
        return {
            "label": str(value) if value not in {None, ""} else "Unnamed header",
            "children": children,
        }

    preview = []
    for child in list(getattr(root, "children", None) or [])[:24]:
        converted = convert(child)
        if converted:
            preview.append(converted)
    return preview


def body_tree_rows(feature_tree, column_count: int, row_limit: int = 40) -> tuple[list[list], int, int]:
    root = getattr(getattr(feature_tree, "body_tree", None), "root", None)
    rows = []
    total_rows = 0
    coordinate_cells = 0
    visited = set()

    def walk(node, path, depth=0):
        nonlocal total_rows, coordinate_cells
        if node is None or id(node) in visited or depth > 80 or len(visited) >= 12000:
            return
        visited.add(id(node))
        next_path = path if node is root else path + [node]
        children = list(getattr(node, "children", None) or [])
        if children:
            for child in children:
                walk(child, next_path, depth + 1)
            return
        if node is root or len(next_path) != column_count:
            return
        total_rows += 1
        if len(rows) >= row_limit:
            return
        row = []
        for body_node in next_path:
            if any(getattr(body_node, name, None) is not None for name in ("x1", "y1", "x2", "y2")):
                coordinate_cells += 1
            row.append(compact_table_value(getattr(body_node, "value", None)))
        rows.append(row)

    walk(root, [])
    return rows, total_rows, coordinate_cells


def body_tree_preview(feature_tree, branch_limit: int = 8, node_limit: int = 120) -> dict:
    root = getattr(getattr(feature_tree, "body_tree", None), "root", None)
    root_children = list(getattr(root, "children", None) or [])
    visited = set()
    truncated = len(root_children) > branch_limit

    def convert(node, depth=0):
        nonlocal truncated
        if node is None or id(node) in visited or depth > 18 or len(visited) >= node_limit:
            truncated = True
            return None
        visited.add(id(node))
        value = getattr(node, "value", None)
        label = "Nested FeatureTree" if isinstance(value, SafeFeatureTree) else compact_table_value(value, 90)
        position = [getattr(node, name, None) for name in ("x1", "y1", "x2", "y2")]
        children = []
        for child in list(getattr(node, "children", None) or []):
            converted = convert(child, depth + 1)
            if converted:
                children.append(converted)
        return {
            "label": str(label) if label not in {None, ""} else "Empty cell",
            "position": position if any(value is not None for value in position) else None,
            "children": children,
        }

    nodes = []
    for child in root_children[:branch_limit]:
        converted = convert(child)
        if converted:
            nodes.append(converted)
    return {
        "nodes": nodes,
        "totalBranches": len(root_children),
        "shownBranches": len(nodes),
        "nodeCount": len(visited),
        "truncated": truncated,
    }


def count_feature_tree_nodes(feature_tree) -> dict:
    seen = {"features": set(), "index": set(), "body": set()}

    def walk_index(node):
        if node is None or id(node) in seen["index"] or len(seen["index"]) >= 4000:
            return
        seen["index"].add(id(node))
        for child in getattr(node, "children", None) or []:
            walk_index(child)
        for body_node in getattr(node, "body", None) or []:
            walk_body(body_node)

    def walk_body(node):
        if node is None or id(node) in seen["body"] or len(seen["body"]) >= 8000:
            return
        seen["body"].add(id(node))
        value = getattr(node, "value", None)
        if isinstance(value, SafeFeatureTree):
            walk_feature(value)
        for child in getattr(node, "children", None) or []:
            walk_body(child)

    def walk_feature(tree):
        if tree is None or id(tree) in seen["features"] or len(seen["features"]) >= 200:
            return
        seen["features"].add(id(tree))
        walk_index(getattr(getattr(tree, "index_tree", None), "root", None))
        walk_body(getattr(getattr(tree, "body_tree", None), "root", None))

    walk_feature(feature_tree)
    return {key: len(value) for key, value in seen.items()}


def feature_tree_summary(feature_tree) -> dict:
    sections = []
    seen_features = set()
    max_feature_depth = 0

    def collect(tree, path, depth=0):
        nonlocal max_feature_depth
        if id(tree) in seen_features or len(sections) >= 24 or depth > 10:
            return
        seen_features.add(id(tree))
        max_feature_depth = max(max_feature_depth, depth)
        leaves = feature_tree_leaves(tree)

        # ST-Raptor commonly wraps the useful table in a single `table` node.
        if len(leaves) == 1:
            body = list(getattr(leaves[0][0], "body", None) or [])
            if len(body) == 1 and isinstance(getattr(body[0], "value", None), SafeFeatureTree):
                label = leaves[0][1][-1] if leaves[0][1] else "table"
                collect(body[0].value, path + [label], depth + 1)
                return

        visible_leaves = leaves[:24]
        has_nested_values = any(
            isinstance(getattr(body_node, "value", None), SafeFeatureTree)
            for node, _ in leaves
            for body_node in (getattr(node, "body", None) or [])
        )
        rows, total_rows, coordinate_cells = body_tree_rows(tree, len(leaves))
        rows = [row[:len(visible_leaves)] for row in rows]
        if has_nested_values or not rows:
            total_rows = max((len(getattr(node, "body", None) or []) for node, _ in leaves), default=0)
            rows = []
            coordinate_cells = 0
            for row_index in range(min(total_rows, 40)):
                row = []
                for node, _ in visible_leaves:
                    body = list(getattr(node, "body", None) or [])
                    body_node = body[row_index] if row_index < len(body) else None
                    if body_node is None:
                        row.append(None)
                        continue
                    if any(getattr(body_node, name, None) is not None for name in ("x1", "y1", "x2", "y2")):
                        coordinate_cells += 1
                    row.append(compact_table_value(getattr(body_node, "value", None)))
                rows.append(row)

        title = path[-1] if path else "HO-Tree table"
        index_values = [column_path[-1] for _, column_path in visible_leaves]
        meaningful_index_values = [
            value for value in index_values
            if str(value).strip().lower() not in {"", "none", "unnamed column"}
        ]
        layout = "table" if rows else "index_values" if meaningful_index_values else "group"
        sections.append({
            "id": f"table-{len(sections) + 1}",
            "title": title,
            "path": path or ["HO-Tree"],
            "columns": [
                {"label": column_path[-1], "path": column_path}
                for _, column_path in visible_leaves
            ],
            "rows": rows,
            "layout": layout,
            "indexValues": index_values if layout == "index_values" else [],
            "totalColumns": len(leaves),
            "totalRows": total_rows,
            "coordinateCells": coordinate_cells,
            "headerTree": index_tree_preview(tree),
            "bodyTree": body_tree_preview(tree),
            "truncated": len(leaves) > len(visible_leaves) or total_rows > len(rows),
        })

        for node, column_path in leaves:
            for body_node in getattr(node, "body", None) or []:
                value = getattr(body_node, "value", None)
                if isinstance(value, SafeFeatureTree):
                    collect(value, path + column_path, depth + 1)

    collect(feature_tree, [])
    counts = count_feature_tree_nodes(feature_tree)
    return {
        "kind": "table_tree",
        "title": "ST-Raptor HO-Tree",
        "sourceFormat": "FeatureTree pickle",
        "tree": {
            "sections": sections,
            "sectionCount": len(sections),
            "totalColumns": sum(section["totalColumns"] for section in sections),
            "totalRows": sum(section["totalRows"] for section in sections),
            "maxDepth": max_feature_depth,
            "nodeCounts": counts,
            "classes": ["FeatureTree", "IndexTree", "IndexNode", "BodyTree", "BodyNode"],
        },
    }


def loaded_pickle_summary(value) -> dict:
    if isinstance(value, SafeFeatureTree):
        return feature_tree_summary(value)
    try:
        import numpy as np
    except ImportError:
        return {"kind": "object", "object": safe_value_preview(value)}

    if isinstance(value, (list, tuple)) and value and isinstance(value[0], np.ndarray):
        embeddings = array_summary(value[0])
        labels = value[2] if len(value) > 2 and isinstance(value[2], (list, tuple)) else []
        row_ids = value[1] if len(value) > 1 and isinstance(value[1], (list, tuple)) else []
        cell_cache = bool(labels) and all(str(label) == "Cell" for label in labels)
        return {
            "kind": "embedding_cache",
            "title": "GraphOtter cell embedding cache" if cell_cache else "Embedding cache",
            "embeddings": embeddings,
            "rowIds": {"count": len(row_ids), "preview": list(row_ids)[:1024]},
            "labels": {"count": len(labels), "distribution": dict(Counter(str(label) for label in labels).most_common(12))},
            "components": len(value),
            "cellCache": cell_cache,
        }
    try:
        if isinstance(value, np.ndarray):
            return {"kind": "array", "title": "NumPy array", "embeddings": array_summary(value)}
    except AttributeError:
        pass
    return {"kind": "object", "title": type(value).__name__, "object": safe_value_preview(value)}


class CompactHtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.rows = []
        self.current_row = None
        self.current_cell = None
        self.current_attrs = {}
        self.current_caption = None
        self.caption = ""

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self.current_row = []
        elif tag == "caption":
            self.current_caption = []
        elif tag in {"td", "th"} and self.current_row is not None:
            self.current_cell = []
            self.current_attrs = dict(attrs)

    def handle_data(self, data):
        if self.current_cell is not None:
            self.current_cell.append(data)
        elif self.current_caption is not None:
            self.current_caption.append(data)

    def handle_endtag(self, tag):
        if tag in {"td", "th"} and self.current_row is not None and self.current_cell is not None:
            self.current_row.append({
                "text": " ".join("".join(self.current_cell).split()),
                "rowspan": max(1, int(self.current_attrs.get("rowspan", "1") or 1)),
                "colspan": max(1, int(self.current_attrs.get("colspan", "1") or 1)),
            })
            self.current_cell = None
            self.current_attrs = {}
        elif tag == "caption" and self.current_caption is not None:
            self.caption = " ".join("".join(self.current_caption).split())
            self.current_caption = None
        elif tag == "tr" and self.current_row is not None:
            self.rows.append(self.current_row)
            self.current_row = None

    def matrix(self) -> list[list[str]]:
        matrix = []
        pending = {}
        for source_row in self.rows:
            row = []
            column = 0

            def consume_pending():
                nonlocal column
                while column in pending:
                    text, remaining = pending[column]
                    row.append(text)
                    if remaining <= 1:
                        del pending[column]
                    else:
                        pending[column] = (text, remaining - 1)
                    column += 1

            consume_pending()
            for cell in source_row:
                consume_pending()
                for offset in range(cell["colspan"]):
                    row.append(cell["text"])
                    if cell["rowspan"] > 1:
                        pending[column + offset] = (cell["text"], cell["rowspan"] - 1)
                column += cell["colspan"]
            consume_pending()
            matrix.append(row)
        width = max((len(row) for row in matrix), default=0)
        return [row + [""] * (width - len(row)) for row in matrix]


def graphotter_table_context(path: Path, context_id: str = "", table_index: int | None = None) -> dict | None:
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, ValueError):
        return None

    title = ""
    header_rows = 1
    header_columns = 1
    matrix = None
    if isinstance(raw, dict) and isinstance(raw.get("texts"), list):
        matrix = copy.deepcopy(raw["texts"])
        title = str(raw.get("title") or path.stem)
        header_rows = max(0, int(raw.get("top_header_rows_num") or 1))
        header_columns = max(0, int(raw.get("left_header_columns_num") or 1))
        for region in raw.get("merged_regions") or []:
            try:
                r1, r2 = int(region["first_row"]), int(region["last_row"])
                c1, c2 = int(region["first_column"]), int(region["last_column"])
                merged_value = matrix[r1][c1]
                for row_index in range(r1, min(r2 + 1, len(matrix))):
                    for column_index in range(c1, min(c2 + 1, len(matrix[row_index]))):
                        matrix[row_index][column_index] = merged_value
            except (IndexError, KeyError, TypeError, ValueError):
                continue
    elif isinstance(raw, list) and context_id:
        record = next((item for item in raw if isinstance(item, dict) and str(item.get("uid")) == context_id), None)
        tables = record.get("tables") if isinstance(record, dict) else None
        if isinstance(tables, list) and table_index is not None and 0 <= table_index < len(tables):
            parser = CompactHtmlTableParser()
            parser.feed(str(tables[table_index]))
            matrix = parser.matrix()
            title = f"Selected source table {table_index}"

    if not matrix or not all(isinstance(row, list) for row in matrix):
        return None
    width = max((len(row) for row in matrix), default=0)
    if not width:
        return None
    normalized = [
        [compact_table_value(value, 160) for value in row]
        for row in matrix
    ]
    return {
        "title": title,
        "rows": len(normalized),
        "columns": width,
        "rowLengths": [len(row) for row in normalized],
        "headerRows": min(header_rows, len(normalized)),
        "headerColumns": min(header_columns, width),
        "values": normalized,
        "source": path.resolve().relative_to(DATA.resolve()).as_posix(),
    }


def attach_embedding_table(summary: dict, table: dict | None) -> None:
    if not table or summary.get("kind") != "embedding_cache":
        return
    embeddings = summary.get("embeddings") or {}
    norms = (embeddings.get("norms") or {}).get("preview") or []
    ids = (summary.get("rowIds") or {}).get("preview") or []
    expected = sum(table.get("rowLengths") or [])
    if expected != (summary.get("rowIds") or {}).get("count") or len(norms) < expected:
        return
    cells = []
    index = 0
    for row in table["values"]:
        for value in row:
            cells.append({
                "id": ids[index] if index < len(ids) else index,
                "value": value,
                "norm": norms[index],
            })
            index += 1
    table["cells"] = cells
    table.pop("values", None)
    summary["embeddingTable"] = table


@lru_cache(maxsize=32)
def pickle_summary_cached(path_text: str, modified_ns: int, size_bytes: int) -> dict:
    path = Path(path_text)
    if size_bytes > MAX_PICKLE_BYTES:
        return {"error": f"PKL is larger than the {MAX_PICKLE_BYTES // (1024 * 1024)} MB inspection limit."}
    data = path.read_bytes()
    static = static_pickle_summary(data)
    try:
        loader = RestrictedPickleLoader(io.BytesIO(data))
        loaded = loader.load()
        if loader.class_references:
            static["classReferences"] = loader.class_references
        result = loaded_pickle_summary(loaded)
        result["safeLoad"] = True
    except (pickle.UnpicklingError, ImportError, AttributeError, TypeError, ValueError, EOFError) as error:
        result = {
            "kind": "static_pickle",
            "title": "Serialized Python object",
            "safeLoad": False,
            "loadNote": str(error),
        }
        class_references = static.get("classReferences") or []
        if any(reference.startswith("table2tree.") for reference in class_references):
            result.update({
                "kind": "table_tree",
                "title": "Hierarchical table tree cache",
                "tree": {
                    "estimatedNodes": static.get("opcodeCounts", {}).get("NEWOBJ", 0),
                    "classCount": len(class_references),
                    "classes": [reference.rsplit(".", 1)[-1] for reference in class_references],
                    "contentPreview": static.get("contentPreview") or [],
                },
            })
    result.update({
        "fileSize": size_bytes,
        "modifiedNs": modified_ns,
        "static": static,
    })
    return result


def reveal_file(path: Path) -> None:
    if sys.platform == "darwin":
        command = ["open", str(path)] if path.is_dir() else ["open", "-R", str(path)]
    elif os.name == "nt":
        command = ["explorer.exe", f"/select,{path}"]
    else:
        opener = shutil.which("xdg-open")
        if not opener:
            raise OSError("No supported file manager opener was found")
        command = [opener, str(path if path.is_dir() else path.parent)]
    subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def value(record: dict, keys: tuple[str, ...]):
    lowered = {str(key).lower().replace("-", "_"): item for key, item in record.items()}
    return next((lowered[key] for key in keys if key in lowered), None)


def flatten(item):
    if isinstance(item, list):
        for child in item:
            yield from flatten(child)
    elif isinstance(item, dict):
        rows = next((item[key] for key in ("records", "results", "examples", "items", "data") if key in item), None)
        if isinstance(rows, list):
            yield from flatten(rows)
        else:
            yield item


def answers_match(gold, prediction) -> bool:
    left, right = str(gold).strip().casefold(), str(prediction).strip().casefold()
    if left == right:
        return True
    number = r"[-+]?\d+(?:\.\d+)?"
    left_numbers, right_numbers = re.findall(number, left), re.findall(number, right)
    return len(left_numbers) == len(right_numbers) == 1 and float(left_numbers[0]) == float(right_numbers[0])


def read_file(path: Path):
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return list(csv.DictReader(handle))
    if path.suffix.lower() == ".jsonl":
        return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    return json.loads(path.read_text(encoding="utf-8"))


@lru_cache(maxsize=16)
def load_records(pipeline_filter: str = "", dataset_filter: str = "") -> list[dict]:
    records = []
    search_root = DATA
    if pipeline_filter and dataset_filter:
        search_root = DATA / "Artifacts" / pipeline_filter / dataset_filter
    if not search_root.exists():
        return records
    for path in search_root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".json", ".jsonl", ".csv"}:
            continue
        try:
            for raw in flatten(read_file(path)):
                if not isinstance(raw, dict):
                    continue
                question = value(raw, QUESTION_KEYS)
                gold = value(raw, GOLD_KEYS)
                prediction = value(raw, PRED_KEYS)
                missing_output = path.name.lower() == "input.jsonl" and not (path.parent / "output.jsonl").exists()
                failed = raw.get("error") or str(raw.get("status", "")).lower() in {"error", "failed", "failure"}
                if question is None or (prediction is None and not missing_output and not failed):
                    continue

                parts = path.relative_to(DATA).parts
                pipeline = pipeline_filter or (parts[1] if len(parts) > 1 else "Unknown")
                dataset = dataset_filter or (parts[2] if len(parts) > 2 else "Unknown")
                explicit_correct = value(raw, ("correct", "is_correct", "exact_match", "em"))
                if missing_output or failed or prediction is None:
                    status = "error"
                elif explicit_correct is not None:
                    status = "correct" if explicit_correct in {True, 1, "1", "true", "True"} else "wrong"
                else:
                    status = "correct" if gold is not None and answers_match(gold, prediction) else "wrong"

                source = str(path.relative_to(DATA))
                records.append({
                    "id": f"{source}:{len(records)}",
                    "pipeline": pipeline,
                    "dataset": dataset,
                    "question": question,
                    "gold": gold,
                    "prediction": prediction,
                    "status": status,
                    "source": source,
                })
        except (csv.Error, OSError, ValueError, UnicodeDecodeError):
            continue
    return records


def load_catalog() -> dict[str, list[str]]:
    artifacts = DATA / "Artifacts"
    if not artifacts.exists():
        return {}
    return {
        pipeline.name: sorted(child.name for child in pipeline.iterdir() if child.is_dir())
        for pipeline in sorted(artifacts.iterdir())
        if pipeline.is_dir()
    }


def default_selection(catalog: dict[str, list[str]]) -> dict[str, str]:
    artifacts = DATA / "Artifacts"
    for pipeline, datasets in catalog.items():
        for dataset in datasets:
            dataset_path = artifacts / pipeline / dataset
            has_runs = (dataset_path / "runs").is_dir() or any(dataset_path.glob("*/runs"))
            if has_runs:
                return {"pipeline": pipeline, "dataset": dataset}
    first_pipeline = next(iter(catalog), "")
    first_dataset = catalog.get(first_pipeline, [""])[0] if first_pipeline else ""
    return {"pipeline": first_pipeline, "dataset": first_dataset}


@lru_cache(maxsize=4)
def load_common_questions(dataset: str) -> list[dict]:
    questions = []
    root = DATA / "Artifacts" / "ST-raptor" / dataset
    for path in root.rglob("input.jsonl"):
        try:
            for raw in flatten(read_file(path)):
                if not isinstance(raw, dict):
                    continue
                question = value(raw, QUESTION_KEYS)
                if question is None:
                    continue
                questions.append({
                    "key": str(raw.get("id") or path.parent.name),
                    "question": question,
                    "gold": value(raw, GOLD_KEYS),
                    "source": str(path.relative_to(DATA)),
                })
        except (OSError, ValueError, UnicodeDecodeError):
            continue
    return questions


def missing_prediction_records(pipeline: str, dataset: str) -> list[dict]:
    return [{
        "id": f"{pipeline}:{dataset}:{item['key']}",
        "pipeline": pipeline,
        "dataset": dataset,
        "question": item["question"],
        "gold": item["gold"],
        "prediction": None,
        "status": "error",
        "source": f"{item['source']} (no {pipeline} prediction artifact)",
    } for item in load_common_questions(dataset)]


def load_output_records(pipeline: str, dataset: str) -> list[dict]:
    output_dir = DATA / "Outputs" / pipeline / dataset
    reports = list(output_dir.glob("report_*.json"))
    if not reports:
        return []

    def report_number(path: Path) -> int:
        suffix = path.stem.removeprefix("report_")
        return int(suffix) if suffix.isdigit() else 0

    report_path = max(reports, key=report_number)
    payload = json.loads(report_path.read_text(encoding="utf-8"))
    records = []
    for index, raw in enumerate(payload.get("results", [])):
        error = raw.get("error")
        exact_match = raw.get("exact_match")
        if error:
            status = "error"
        else:
            try:
                status = "correct" if float(exact_match or 0) >= 1 else "wrong"
            except (TypeError, ValueError):
                status = "wrong"
        artifacts = artifact_paths(raw, pipeline, dataset, report_path)
        representation = None
        if pipeline == "GraphOtter":
            metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
            preprocess = metadata.get("preprocess") if isinstance(metadata.get("preprocess"), dict) else {}
            selected_sources = preprocess.get("selected_sources") if isinstance(preprocess.get("selected_sources"), list) else []
            selected_match = re.search(r"table\[(\d+)\]", str(selected_sources[0])) if selected_sources else None
            representation = {
                "contextId": str(preprocess.get("table_id") or raw.get("table_id") or raw.get("sample_id") or ""),
                "tableIndex": int(selected_match.group(1)) if selected_match else None,
                "rows": preprocess.get("row_count"),
                "columns": preprocess.get("column_count"),
            }
        records.append({
            "id": f"{pipeline}:{dataset}:{raw.get('sample_id') or index}",
            "pipeline": pipeline,
            "dataset": dataset,
            "question": raw.get("question") or "",
            "gold": raw.get("gold_answer"),
            "prediction": raw.get("predicted_answer"),
            "status": status,
            "source": str(report_path.relative_to(DATA)),
            "artifacts": artifacts,
            "representation": representation,
        })
    return records


@lru_cache(maxsize=8)
def artifact_catalog(pipeline: str, dataset: str) -> tuple[dict[str, list[Path]], dict[str, list[Path]]]:
    root = DATA / "Artifacts" / pipeline / dataset
    by_name: dict[str, list[Path]] = {}
    by_part: dict[str, list[Path]] = {}
    if not root.exists():
        return by_name, by_part
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        by_name.setdefault(path.name.casefold(), []).append(path)
        for part in path.relative_to(root).parts[:-1]:
            by_part.setdefault(part.casefold(), []).append(path)
    return by_name, by_part


def relative_file(path: Path | None) -> str | None:
    if path and path.is_file():
        return path.relative_to(DATA).as_posix()
    return None


def external_filename(raw_path) -> str:
    if not raw_path:
        return ""
    cleaned = re.split(r":(?:line\s+\d+|json\[.*)$", str(raw_path), flags=re.IGNORECASE)[0]
    cleaned = cleaned.split("!")[-1]
    return re.split(r"[\\/]", cleaned)[-1]


def artifact_paths(raw: dict, pipeline: str, dataset: str, report_path: Path) -> dict[str, list[str]]:
    sections: dict[str, list[str]] = {key: [] for key in ("input", "interpreted", "workflow", "output")}
    artifact_root = DATA / "Artifacts" / pipeline / dataset
    by_name, by_part = artifact_catalog(pipeline, dataset)
    metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}

    def add(section: str, path: Path | None):
        relative = relative_file(path)
        if relative and relative not in sections[section]:
            sections[section].append(relative)

    sample_name = external_filename(raw.get("sample_path"))
    if sample_name:
        dataset_matches = list((DATA / "Datasets" / dataset).rglob(sample_name))
        add("input", dataset_matches[0] if dataset_matches else None)
    table_id = str(raw.get("table_id") or raw.get("sample_id") or "")
    add("input", DATA / "Datasets" / f"{dataset}_xlsx" / f"{table_id}.xlsx")

    if pipeline == "SpreadsheetAgent":
        contexts = metadata.get("contexts") if isinstance(metadata.get("contexts"), list) else []
        for context in contexts:
            if not isinstance(context, dict):
                continue
            cache_name = external_filename(context.get("cache_dir"))
            cache_root = artifact_root / cache_name
            for structure in context.get("structure_paths") or []:
                add("interpreted", cache_root / external_filename(structure))
            add("workflow", cache_root / "manifest.json")

    elif pipeline == "GraphOtter":
        preprocess = metadata.get("preprocess") if isinstance(metadata.get("preprocess"), dict) else {}
        table_name = external_filename(preprocess.get("table_path"))
        matches = by_name.get(table_name.casefold(), []) if table_name else []
        add("interpreted", matches[0] if matches else None)

        official_table_id = str(preprocess.get("official_table_id") or "")
        cache_table_id = official_table_id or str(preprocess.get("table_id") or table_id)
        cache_root = artifact_root / "cache"
        cache_matches = sorted(cache_root.glob(f"*-{cache_table_id}.pkl")) if cache_table_id else []
        if not cache_matches and official_table_id:
            cache_key = official_table_id.rsplit("-", 1)[-1]
            cache_matches = sorted(cache_root.glob(f"*-{cache_key}.pkl"))
        for cache_path in cache_matches[:3]:
            add("interpreted", cache_path)

        logs = by_name.get("graphotter_official.log", [])
        add("workflow", logs[0] if logs else None)
        cli_logs = by_name.get("cli.log", [])
        add("workflow", cli_logs[0] if cli_logs else None)

    elif pipeline == "ST-raptor":
        run_name = external_filename(metadata.get("run_dir"))
        run_files = by_part.get(run_name.casefold(), []) if run_name else []
        add("interpreted", next((path for path in run_files if path.suffix.lower() == ".html"), None))
        add("workflow", next((path for path in run_files if path.name == "payload.json"), None))

        preprocess = metadata.get("preprocess") if isinstance(metadata.get("preprocess"), dict) else {}
        official_table_id = str(preprocess.get("official_table_id") or "")
        pkl_matches = by_name.get(f"{official_table_id}.pkl".casefold(), []) if official_table_id else []
        add("interpreted", pkl_matches[0] if pkl_matches else None)

        log_name = external_filename(metadata.get("log_dir"))
        log_files = by_part.get(log_name.casefold(), []) if log_name else []
        add("workflow", next((path for path in log_files if path.suffix.lower() == ".log"), None))
        add("output", next((path for path in run_files if path.name == "output.jsonl"), None))

    add("output", report_path)
    return sections


def siflex_path_detail(raw_path, run_root: Path) -> dict | None:
    if not raw_path:
        return None
    original = str(raw_path)
    normalized = original.replace("\\", "/")
    cleaned = re.split(r":(?:cases|json)\[", normalized, maxsplit=1)[0]
    candidates = []

    run_marker = f"TableAgent/outputs/{run_root.name}/"
    if run_marker in cleaned:
        candidates.append(run_root / cleaned.split(run_marker, 1)[1])
    prepared_marker = "outputs/v2/prepared/"
    if prepared_marker in cleaned:
        candidates.append(SIFLEX_ROOT / "v2" / "prepared" / cleaned.split(prepared_marker, 1)[1])
    if cleaned.startswith(str(DATA).replace("\\", "/")):
        candidates.append(Path(cleaned))

    for candidate in candidates:
        if candidate.exists():
            return {"path": candidate.relative_to(DATA).as_posix(), "available": True}
    if candidates:
        return {"path": candidates[0].relative_to(DATA).as_posix(), "available": False}
    return {"path": original, "available": False}


def build_siflex_index() -> dict:
    run_roots = sorted(path for path in SIFLEX_ROOT.glob("siflex-table_agent-*") if path.is_dir())
    if not run_roots:
        return {"version": 1, "run": {}, "records": []}

    run_root = run_roots[-1]
    reports = sorted((run_root / "evaluations").glob("report_*.json"))
    if not reports:
        return {"version": 1, "run": {"name": run_root.name}, "records": []}

    report_path = reports[-1]
    payload = json.loads(report_path.read_text(encoding="utf-8"))
    report_detail = {"path": report_path.relative_to(DATA).as_posix(), "available": True}
    golden_cases = DATA / "Datasets" / "SiFlex" / "golden_tests" / "compiled" / "golden_cases.json"
    golden_detail = {"path": golden_cases.relative_to(DATA).as_posix(), "available": golden_cases.is_file()}
    golden_case_values = []
    if golden_cases.is_file():
        try:
            golden_payload = json.loads(golden_cases.read_text(encoding="utf-8"))
            golden_case_values = golden_payload.get("cases", [])
        except (OSError, ValueError):
            golden_case_values = []
    siflex_data_root = DATA / "Datasets" / "SiFlex" / "golden_tests" / "data"
    records = []

    for index, raw in enumerate(payload.get("results", [])):
        metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
        qa = metadata.get("qa") if isinstance(metadata.get("qa"), dict) else {}
        qa_artifacts = qa.get("artifacts") if isinstance(qa.get("artifacts"), dict) else {}

        def details(*paths):
            items = []
            for raw_path in paths:
                detail = siflex_path_detail(raw_path, run_root)
                if detail and detail["path"] not in {item["path"] for item in items}:
                    items.append(detail)
            return items

        q_paths = [golden_detail]
        table_name = external_filename(raw.get("table_path") or metadata.get("workbook_path"))
        table_matches = list(siflex_data_root.rglob(table_name)) if table_name else []
        x_paths = [{"path": table_matches[0].relative_to(DATA).as_posix(), "available": True}] if table_matches else details(raw.get("table_path") or metadata.get("workbook_path"))
        z_paths = details(metadata.get("structure_path"))
        w_paths = details(
            qa_artifacts.get("plan_json"),
            qa_artifacts.get("events_jsonl"),
            qa_artifacts.get("notebook_ipynb"),
            qa_artifacts.get("generated_code_dir"),
            qa_artifacts.get("cells_dir"),
        )
        y_paths = details(qa_artifacts.get("result_json"), qa_artifacts.get("answer_py")) + [report_detail]

        if raw.get("error"):
            status = "error"
        else:
            status = "correct" if raw.get("pass") else "wrong"

        golden_answer = raw.get("gold_answer")
        sample_path = str(raw.get("sample_path") or "")
        case_match = re.search(r":cases\[(\d+)\]$", sample_path)
        if case_match:
            case_index = int(case_match.group(1))
            if 0 <= case_index < len(golden_case_values):
                reference_answer = golden_case_values[case_index].get("reference_answer")
                if isinstance(reference_answer, str):
                    golden_answer = [reference_answer]

        records.append({
            "id": f"TableAgent-SIFLEX:{raw.get('sample_id') or index}",
            "pipeline": "TableAgent-SIFLEX",
            "dataset": "SiFlex",
            "question": raw.get("question") or "",
            "gold": golden_answer,
            "prediction": raw.get("predicted_answer"),
            "status": status,
            "source": report_detail["path"],
            "components": {
                "q": {"paths": q_paths},
                "X": {"paths": x_paths},
                "Z": {"paths": z_paths},
                "W": {"paths": w_paths},
                "Y": {"paths": y_paths},
                "Y*": {"paths": [golden_detail]},
            },
            "metrics": {
                "overall_score": raw.get("overall_score"),
                "factual_correctness": raw.get("factual_correctness"),
                "coverage": raw.get("coverage"),
                "structure_fidelity": raw.get("structure_fidelity"),
                "grounding": raw.get("grounding"),
            },
        })

    return {
        "version": 1,
        "run": {
            "name": payload.get("run_name") or run_root.name,
            "dataset": payload.get("dataset") or "siflex",
            "pass": payload.get("pass", 0),
            "fail": payload.get("fail", 0),
            "pass_rate": payload.get("pass_rate"),
        },
        "records": records,
    }


def normalized_structure_text(value) -> str:
    text = str(value or "").replace("\\n", " ").replace("\xa0", " ")
    return re.sub(r"[\W_]+", " ", text, flags=re.UNICODE).strip().casefold()


def siflex_header_summary(raw: dict) -> dict:
    sub_headers = raw.get("sub_headers") if isinstance(raw.get("sub_headers"), list) else []
    return {
        "id": str(raw.get("id") or ""),
        "label": str(raw.get("label") or raw.get("name") or raw.get("id") or "Unnamed field"),
        "description": str(raw.get("description") or ""),
        "orientation": str(raw.get("orientation") or ""),
        "headerRange": str(raw.get("header_range") or ""),
        "dataRange": str(raw.get("data_range") or ""),
        "children": [siflex_header_summary(item) for item in sub_headers if isinstance(item, dict)],
    }


def siflex_structure_sections(payload) -> list[dict]:
    sections = []

    def visit(value, key=""):
        if not isinstance(value, dict):
            return
        if any(field in value for field in ("sheet", "headers", "header_range", "data_range")):
            headers = value.get("headers") if isinstance(value.get("headers"), list) else []
            sections.append({
                "id": str(value.get("id") or key or f"section_{len(sections) + 1}"),
                "name": str(value.get("name") or value.get("id") or key or "Unnamed section"),
                "description": str(value.get("description") or ""),
                "sheet": str(value.get("sheet") or ""),
                "headers": [siflex_header_summary(item) for item in headers if isinstance(item, dict)],
            })
            return
        for child_key, child in value.items():
            visit(child, str(child_key))

    visit(payload)
    return sections


def flatten_siflex_headers(headers: list[dict]):
    for header in headers:
        yield header
        yield from flatten_siflex_headers(header.get("children", []))


def siflex_validation(source_path: Path | None, sections: list[dict], yaml_path: Path) -> dict:
    method = "SiFlex structure.yaml claims checked against the source workbook"
    if source_path is None or not source_path.is_file():
        return {
            "status": "unverifiable",
            "method": method,
            "reason": "The source workbook is unavailable, so the YAML structure cannot be checked against X.",
            "artifacts": [yaml_path.relative_to(DATA).as_posix()],
            "metrics": [],
        }

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries

        workbook = load_workbook(source_path, data_only=True, read_only=False)
    except (OSError, ValueError, ImportError) as error:
        return {
            "status": "unverifiable",
            "method": method,
            "reason": f"The source workbook could not be inspected: {error}",
            "artifacts": [yaml_path.relative_to(DATA).as_posix(), source_path.relative_to(DATA).as_posix()],
            "metrics": [],
        }

    sheet_total = len(sections)
    sheet_matches = 0
    range_total = 0
    range_matches = 0
    data_total = 0
    populated_data = 0
    label_total = 0
    label_matches = 0
    issues = []

    try:
        for section in sections:
            sheet_name = section.get("sheet", "")
            worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
            if worksheet is None:
                issues.append(f'Sheet "{sheet_name or "(missing)"}" is not present in X.')
            else:
                sheet_matches += 1

            for header in flatten_siflex_headers(section.get("headers", [])):
                header_label = str(header.get("label") or header.get("id") or "Unnamed field")
                header_range = header.get("headerRange")
                data_range = header.get("dataRange")
                parsed_ranges = {}

                for range_kind, raw_range in (("header", header_range), ("data", data_range)):
                    range_total += 1
                    if range_kind == "data":
                        data_total += 1
                    try:
                        bounds = range_boundaries(str(raw_range)) if raw_range else None
                    except ValueError:
                        bounds = None
                    valid = bool(
                        worksheet and bounds
                        and bounds[0] >= 1 and bounds[1] >= 1
                        and bounds[2] <= worksheet.max_column and bounds[3] <= worksheet.max_row
                    )
                    if not valid:
                        issues.append(f'{header_label}: {range_kind}_range "{raw_range or "(missing)"}" is invalid for sheet "{sheet_name}".')
                        continue
                    range_matches += 1
                    parsed_ranges[range_kind] = bounds

                    if range_kind == "data":
                        min_column, min_row, max_column, max_row = bounds
                        populated = any(
                            worksheet.cell(row=row, column=column).value not in (None, "")
                            for row in range(min_row, max_row + 1)
                            for column in range(min_column, max_column + 1)
                        )
                        if populated:
                            populated_data += 1
                        else:
                            issues.append(f'{header_label}: data_range {raw_range} contains no values in X.')

                normalized_label = normalized_structure_text(header_label)
                if normalized_label:
                    label_total += 1
                    bounds = parsed_ranges.get("header")
                    cell_labels = []
                    if worksheet and bounds:
                        min_column, min_row, max_column, max_row = bounds
                        cell_labels = [
                            normalized_structure_text(worksheet.cell(row=row, column=column).value)
                            for row in range(min_row, max_row + 1)
                            for column in range(min_column, max_column + 1)
                        ]
                    aligned = any(
                        cell_label and (normalized_label in cell_label or cell_label in normalized_label)
                        for cell_label in cell_labels
                    )
                    if aligned:
                        label_matches += 1
                    else:
                        issues.append(f'{header_label}: label does not align with header_range {header_range or "(missing)"}.')
    finally:
        workbook.close()

    scores = {
        "sheets": sheet_matches / max(1, sheet_total),
        "ranges": range_matches / max(1, range_total),
        "populatedData": populated_data / max(1, data_total),
        "headerAlignment": label_matches / max(1, label_total),
    }
    valid = bool(
        sections and label_total
        and scores["sheets"] == 1
        and scores["ranges"] == 1
        and scores["populatedData"] >= 0.50
        and scores["headerAlignment"] >= 0.60
    )
    return {
        "status": "validated" if valid else "invalid",
        "method": method,
        "description": "Every referenced sheet and range exists; at least 50% of data ranges contain values; at least 60% of YAML labels align with their declared workbook headers.",
        **{key: round(value, 4) for key, value in scores.items()},
        "metrics": [
            {"key": "sheets", "label": "Sheets", "value": round(scores["sheets"], 4), "detail": f"{sheet_matches}/{sheet_total}"},
            {"key": "ranges", "label": "Ranges", "value": round(scores["ranges"], 4), "detail": f"{range_matches}/{range_total}"},
            {"key": "populatedData", "label": "Data populated", "value": round(scores["populatedData"], 4), "detail": f"{populated_data}/{data_total}"},
            {"key": "headerAlignment", "label": "Header alignment", "value": round(scores["headerAlignment"], 4), "detail": f"{label_matches}/{label_total}"},
        ],
        "issues": issues[:12],
        "artifacts": [yaml_path.relative_to(DATA).as_posix(), source_path.relative_to(DATA).as_posix()],
    }


@lru_cache(maxsize=64)
def yaml_structure_summary_cached(
    yaml_path_string: str,
    yaml_modified_ns: int,
    yaml_size: int,
    source_path_string: str,
    source_modified_ns: int,
    source_size: int,
) -> dict:
    import yaml

    yaml_path = Path(yaml_path_string)
    source_path = Path(source_path_string) if source_path_string else None
    try:
        payload = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
    except (yaml.YAMLError, UnicodeError) as error:
        raise ValueError(f"Invalid YAML: {error}") from error
    sections = siflex_structure_sections(payload)
    headers = [header for section in sections for header in flatten_siflex_headers(section.get("headers", []))]
    return {
        "title": yaml_path.name,
        "path": yaml_path.relative_to(DATA).as_posix(),
        "sectionCount": len(sections),
        "headerCount": len(headers),
        "sheets": list(dict.fromkeys(section.get("sheet", "") for section in sections if section.get("sheet"))),
        "sections": sections,
        "validation": siflex_validation(source_path, sections, yaml_path),
    }


def yaml_structure_summary(yaml_path: Path, source_path: Path | None = None) -> dict:
    yaml_stat = yaml_path.stat()
    source_stat = source_path.stat() if source_path and source_path.is_file() else None
    return copy.deepcopy(yaml_structure_summary_cached(
        str(yaml_path),
        yaml_stat.st_mtime_ns,
        yaml_stat.st_size,
        str(source_path) if source_stat else "",
        source_stat.st_mtime_ns if source_stat else 0,
        source_stat.st_size if source_stat else 0,
    ))


def tabular_local_detail(path: Path) -> dict:
    return {"path": path.relative_to(DATA).as_posix(), "available": path.exists()}


def tabular_recorded_detail(path: str) -> dict:
    return {"path": path.replace("\\", "/"), "available": False}


def build_tabular_models_index() -> dict:
    readme_path = TABULAR_MODELS_ROOT / "README.md"
    readme_detail = tabular_local_detail(readme_path)
    statuses = {
        "Correct": ("correct", "Correct"),
        "Incorrect": ("wrong", "Incorrect"),
        "Insufficient": ("error", "Insufficient"),
    }
    model_labels = {
        "Log_Deepseek_V4_Flash": "DeepSeek V4 Flash",
        "Log_Deepseek_V4_Pro": "DeepSeek V4 Pro",
        "Log_Qwen3_Coder_30B": "Qwen3 Coder 30B",
    }
    models = []
    records = []
    matched_source_paths = set()

    if not TABULAR_MODELS_ROOT.is_dir() or not DATALAKE_CHALLENGE_ROOT.is_dir():
        return {"version": 2, "root": {}, "models": [], "records": []}

    uploaded_files = {
        path.relative_to(DATALAKE_CHALLENGE_ROOT).as_posix(): path
        for path in DATALAKE_CHALLENGE_ROOT.rglob("*")
        if path.is_file()
    }

    for model_root in sorted(path for path in TABULAR_MODELS_ROOT.iterdir() if path.is_dir()):
        counts = {name: 0 for name in statuses}
        model_record_count = 0

        for category, (status, verdict_label) in statuses.items():
            category_root = model_root / category
            if not category_root.is_dir():
                continue

            for case_root in sorted(path for path in category_root.iterdir() if path.is_dir()):
                run_path = case_root / "run.json"
                result_path = case_root / "result.json"
                if not run_path.is_file() or not result_path.is_file():
                    continue
                try:
                    run = json.loads(run_path.read_text(encoding="utf-8"))
                    result = json.loads(result_path.read_text(encoding="utf-8"))
                except (OSError, ValueError, UnicodeDecodeError):
                    continue

                data_dir = str(result.get("data_dir") or run.get("data_dir") or "").rstrip("\\/")
                used_files = result.get("used_files") if isinstance(result.get("used_files"), list) else []
                evidence = result.get("evidence") if isinstance(result.get("evidence"), list) else []
                source_names = []
                for raw_name in used_files + [item.get("path") for item in evidence if isinstance(item, dict)]:
                    normalized_name = str(raw_name).replace("\\", "/").removeprefix("./") if raw_name else ""
                    if normalized_name and normalized_name not in source_names:
                        source_names.append(normalized_name)

                local_sources = [uploaded_files[name] for name in source_names if name in uploaded_files]
                if not local_sources:
                    continue

                counts[category] += 1
                model_record_count += 1
                matched_source_paths.update(path.relative_to(DATALAKE_CHALLENGE_ROOT).as_posix() for path in local_sources)

                x_paths = [tabular_local_detail(path) for path in local_sources]
                for source_name in source_names:
                    if source_name in uploaded_files:
                        continue
                    recorded = f"{data_dir}/{source_name}" if data_dir else source_name
                    x_paths.append(tabular_recorded_detail(recorded))

                artifact_snapshots = sorted((case_root / "scratch" / "artifacts").glob("**/manifest.json"))
                artifact_snapshots += sorted((case_root / "scratch" / "artifacts").glob("**/inspection.json"))
                z_paths = [tabular_local_detail(path) for path in artifact_snapshots]
                work_state_path = case_root / "solver_work_state.json"
                if work_state_path.is_file():
                    z_paths.insert(0, tabular_local_detail(work_state_path))

                workflow_candidates = [
                    case_root / "transcript.md",
                    case_root / "events.jsonl",
                    case_root / "messages.json",
                    case_root / "model_exchanges",
                    case_root / "commands",
                ]
                scratch_root = case_root / "scratch"
                if scratch_root.is_dir():
                    workflow_candidates.extend(
                        path for path in sorted(scratch_root.iterdir()) if path.name != "artifacts"
                    )
                w_paths = [tabular_local_detail(path) for path in workflow_candidates if path.exists()]

                usage = result.get("usage") if isinstance(result.get("usage"), dict) else {}
                records.append({
                    "id": f"Tabular-Models:{model_root.name}:{case_root.name}",
                    "pipeline": "Tabular-Models",
                    "dataset": model_labels.get(model_root.name, model_root.name),
                    "modelId": model_root.name,
                    "model": run.get("model") or model_labels.get(model_root.name, model_root.name),
                    "caseId": case_root.name,
                    "category": category,
                    "verdictLabel": verdict_label,
                    "question": result.get("query") or run.get("query") or "",
                    "gold": None,
                    "prediction": result.get("answer"),
                    "status": status,
                    "source": result_path.relative_to(DATA).as_posix(),
                    "referenceNote": (
                        "The exact benchmark reference answer is not stored in Log_Tabular_Models. "
                        f"This run is filed as {category}, which is the available evaluator verdict."
                    ),
                    "components": {
                        "q": {"paths": [tabular_local_detail(run_path)]},
                        "X": {"paths": x_paths},
                        "Z": {"paths": z_paths},
                        "W": {"paths": w_paths},
                        "Y": {"paths": [tabular_local_detail(result_path)]},
                        "Y*": {"paths": [readme_detail]},
                    },
                    "runMeta": {
                        "runId": result.get("run_id") or run.get("run_id") or case_root.name,
                        "turns": result.get("turns"),
                        "confidence": result.get("confidence"),
                        "answerStatus": result.get("answer_status"),
                        "terminationReason": result.get("termination_reason"),
                        "cost": usage.get("cost"),
                        "usedFiles": source_names,
                        "matchedFiles": [path.relative_to(DATALAKE_CHALLENGE_ROOT).as_posix() for path in local_sources],
                    },
                })

        if model_record_count:
            models.append({
                "id": model_root.name,
                "label": model_labels.get(model_root.name, model_root.name),
                "records": model_record_count,
                "counts": counts,
            })

    return {
        "version": 2,
        "root": {
            "name": TABULAR_MODELS_ROOT.name,
            "readme": readme_detail,
            "records": len(records),
            "dataLake": tabular_local_detail(DATALAKE_CHALLENGE_ROOT),
            "matchedFiles": len(matched_source_paths),
        },
        "models": models,
        "records": records,
    }


def normalized_validation_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        text = format(value, ".12g")
    else:
        text = str(value)
    return re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip().casefold()


@lru_cache(maxsize=128)
def workbook_validation_matrices(path_string: str, modified_ns: int, size: int, expand_merged: bool) -> tuple[tuple[tuple[str, ...], ...], ...]:
    from openpyxl import load_workbook

    workbook = load_workbook(path_string, data_only=True, read_only=False)
    matrices = []
    try:
        for sheet in workbook.worksheets:
            matrix = [
                [normalized_validation_cell(sheet.cell(row=row, column=column).value) for column in range(1, sheet.max_column + 1)]
                for row in range(1, sheet.max_row + 1)
            ]
            if expand_merged:
                for merged in sheet.merged_cells.ranges:
                    value = matrix[merged.min_row - 1][merged.min_col - 1]
                    for row_index in range(merged.min_row - 1, merged.max_row):
                        for column_index in range(merged.min_col - 1, merged.max_col):
                            matrix[row_index][column_index] = value
            matrices.append(tuple(tuple(row) for row in matrix))
    finally:
        workbook.close()
    return tuple(matrices)


def workbook_matrices(path: Path, expand_merged: bool = False) -> list[list[list[str]]]:
    stat = path.stat()
    return [
        [list(row) for row in matrix]
        for matrix in workbook_validation_matrices(str(path), stat.st_mtime_ns, stat.st_size, expand_merged)
    ]


def markdown_table_matrices(text: str) -> list[list[list[str]]]:
    tables = []
    current = []
    for raw_line in text.splitlines() + [""]:
        line = raw_line.strip()
        if line.startswith("|") and line.endswith("|"):
            cells = [normalized_validation_cell(cell) for cell in line[1:-1].split("|")]
            if cells and all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells):
                continue
            current.append(cells)
        elif current:
            tables.append(current)
            current = []
    return tables


def validation_profile(matrices: list[list[list[str]]]) -> tuple[Counter, Counter]:
    values = Counter()
    rows = Counter()
    for matrix in matrices:
        for row in matrix:
            normalized = tuple(value for value in (normalized_validation_cell(cell) for cell in row) if value)
            if not normalized:
                continue
            rows[normalized] += 1
            values.update(normalized)
    return values, rows


def validation_scores(source_matrices: list[list[list[str]]], representation_matrices: list[list[list[str]]]) -> dict:
    source_values, source_rows = validation_profile(source_matrices)
    representation_values, representation_rows = validation_profile(representation_matrices)
    shared_values = sum((source_values & representation_values).values())
    shared_rows = sum((source_rows & representation_rows).values())
    return {
        "coverage": shared_values / max(1, sum(source_values.values())),
        "precision": shared_values / max(1, sum(representation_values.values())),
        "rowCoverage": shared_rows / max(1, sum(source_rows.values())),
    }


def representation_is_valid(scores: dict) -> bool:
    return scores["coverage"] >= 0.95 and scores["precision"] >= 0.75 and scores["rowCoverage"] >= 0.80


def best_sheet_validation(source_sheets: list[list[list[str]]], representation: list[list[list[str]]]) -> tuple[dict, int]:
    candidates = [(validation_scores([sheet], representation), index) for index, sheet in enumerate(source_sheets)]
    return max(candidates, key=lambda item: (item[0]["rowCoverage"], item[0]["coverage"], item[0]["precision"]), default=({"coverage": 0, "precision": 0, "rowCoverage": 0}, -1))


def validate_graphotter_representation(record: dict, source_path: Path, z_paths: list[Path]) -> dict:
    table_path = next((path for path in z_paths if path.suffix.lower() == ".json"), None)
    if table_path is None:
        return {"status": "unverifiable", "reason": "The saved Z contains embeddings but no recoverable table content."}
    try:
        payload = json.loads(table_path.read_text(encoding="utf-8"))
        texts = payload.get("texts") if isinstance(payload.get("texts"), list) else []
        representation = []
        if payload.get("title"):
            representation.append([[normalized_validation_cell(payload["title"])]] )
        representation.append([[normalized_validation_cell(cell) for cell in row] for row in texts])
        scores, sheet_index = best_sheet_validation(workbook_matrices(source_path), representation)
    except (OSError, ValueError, UnicodeDecodeError):
        return {"status": "invalid", "reason": "The GraphOtter table representation could not be parsed."}
    return validation_result("GraphOtter JSON vs source workbook", scores, sheet_index, [table_path])


def validate_straptor_representation(record: dict, source_path: Path, z_paths: list[Path]) -> dict:
    html_path = next((path for path in z_paths if path.suffix.lower() == ".html"), None)
    if html_path is None:
        return {"status": "unverifiable", "reason": "No converted table HTML is available for content validation."}
    try:
        parser = CompactHtmlTableParser()
        parser.feed(html_path.read_text(encoding="utf-8", errors="replace"))
        representation = [parser.matrix()]
        source_sheets = workbook_matrices(source_path, expand_merged=True)
        if parser.caption:
            caption = normalized_validation_cell(parser.caption)
            for sheet in source_sheets:
                if sheet and caption in {cell for cell in sheet[0] if cell}:
                    sheet.pop(0)
        scores, sheet_index = best_sheet_validation(source_sheets, representation)
    except (OSError, ValueError):
        return {"status": "invalid", "reason": "The ST-Raptor HTML representation could not be parsed."}
    return validation_result("ST-Raptor HTML vs source workbook", scores, sheet_index, [html_path])


def validate_spreadsheet_agent_representation(record: dict, source_path: Path, z_paths: list[Path]) -> dict:
    representation = []
    parsed_paths = []
    seen_tables = set()
    for path in z_paths:
        if path.suffix.lower() != ".json":
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, ValueError, UnicodeDecodeError):
            continue
        tables = markdown_table_matrices(str(payload.get("structure") or ""))
        if tables:
            for table in tables:
                signature = tuple(tuple(row) for row in table)
                if signature not in seen_tables:
                    seen_tables.add(signature)
                    representation.append(table)
            parsed_paths.append(path)
    if not representation:
        return {"status": "unverifiable", "reason": "No table content could be parsed from the SpreadsheetAgent structures."}
    scores = validation_scores(workbook_matrices(source_path), representation)
    return validation_result("SpreadsheetAgent Markdown vs source workbook", scores, None, parsed_paths)


def validation_result(method: str, scores: dict, sheet_index: int | None, paths: list[Path]) -> dict:
    return {
        "status": "validated" if representation_is_valid(scores) else "invalid",
        "method": method,
        "coverage": round(scores["coverage"], 4),
        "precision": round(scores["precision"], 4),
        "rowCoverage": round(scores["rowCoverage"], 4),
        "matchedSheet": sheet_index,
        "artifacts": [path.relative_to(DATA).as_posix() for path in paths],
    }


def validate_record_representation(record: dict) -> dict:
    artifacts = record.get("artifacts") if isinstance(record.get("artifacts"), dict) else {}
    source_relative = next((path for path in artifacts.get("input", []) if str(path).lower().endswith(".xlsx")), "")
    source_path = DATA / source_relative
    z_paths = [DATA / path for path in artifacts.get("interpreted", []) if (DATA / path).is_file()]
    if not source_relative or not source_path.is_file() or not z_paths:
        return {"status": "unverifiable", "reason": "The source workbook or interpreted representation is missing."}
    validators = {
        "GraphOtter": validate_graphotter_representation,
        "ST-raptor": validate_straptor_representation,
        "SpreadsheetAgent": validate_spreadsheet_agent_representation,
    }
    validator = validators.get(record.get("pipeline"))
    return validator(record, source_path, z_paths) if validator else {"status": "unverifiable", "reason": "No validator is available for this pipeline."}


def validate_siflex_representation(record: dict) -> dict:
    components = record.get("components") if isinstance(record.get("components"), dict) else {}

    def available_path(stage: str, suffixes: tuple[str, ...]) -> Path | None:
        paths = components.get(stage, {}).get("paths", []) if isinstance(components.get(stage), dict) else []
        for detail in paths:
            if not isinstance(detail, dict) or not detail.get("available"):
                continue
            path = DATA / str(detail.get("path") or "")
            if path.is_file() and path.suffix.lower() in suffixes:
                return path
        return None

    source_path = available_path("X", (".xlsx", ".xlsm"))
    yaml_path = available_path("Z", (".yaml", ".yml"))
    if source_path is None or yaml_path is None:
        return {"status": "unverifiable", "reason": "The SiFlex source workbook or structure.yaml artifact is missing."}
    try:
        return yaml_structure_summary(yaml_path, source_path)["validation"]
    except (OSError, ValueError, ImportError) as error:
        return {"status": "unverifiable", "reason": f"The SiFlex structure could not be validated: {error}"}


def normalized_index_answer(value) -> str:
    if isinstance(value, list):
        text = ", ".join(str(item).strip() for item in value)
    else:
        text = "" if value is None else str(value).strip()
    return re.sub(r"\s+", " ", text).strip().casefold()


def build_mismatch_index(index_payload: dict | None = None) -> dict:
    payload = index_payload or load_index()
    records = []
    checked = Counter()
    for pipeline in ("GraphOtter", "SpreadsheetAgent", "ST-raptor"):
        for dataset, items in payload.get("records", {}).get(pipeline, {}).items():
            for record in items:
                artifacts = record.get("artifacts") if isinstance(record.get("artifacts"), dict) else {}
                candidate = (
                    record.get("status") == "wrong"
                    and record.get("prediction") is not None
                    and record.get("gold") is not None
                    and normalized_index_answer(record.get("prediction")) != normalized_index_answer(record.get("gold"))
                    and bool(artifacts.get("workflow"))
                    and bool(artifacts.get("output"))
                )
                if not candidate:
                    continue
                validation = validate_record_representation(record)
                checked[validation["status"]] += 1
                if validation["status"] == "validated":
                    records.append({"id": record["id"], "validation": validation})
    for record in build_siflex_index().get("records", []):
        components = record.get("components") if isinstance(record.get("components"), dict) else {}
        workflow_paths = components.get("W", {}).get("paths", []) if isinstance(components.get("W"), dict) else []
        output_paths = components.get("Y", {}).get("paths", []) if isinstance(components.get("Y"), dict) else []
        workflow_complete = any(isinstance(item, dict) and item.get("available") for item in workflow_paths)
        result_saved = any(
            isinstance(item, dict) and item.get("available") and str(item.get("path", "")).endswith("/result.json")
            for item in output_paths
        )
        candidate = (
            record.get("status") == "wrong"
            and record.get("prediction") is not None
            and record.get("gold") is not None
            and normalized_index_answer(record.get("prediction")) != normalized_index_answer(record.get("gold"))
            and workflow_complete
            and result_saved
        )
        if not candidate:
            continue
        validation = validate_siflex_representation(record)
        checked[validation["status"]] += 1
        if validation["status"] == "validated":
            records.append({"id": record["id"], "validation": validation})
    counts = Counter(item["id"].split(":", 1)[0] for item in records)
    return {
        "version": 2,
        "criteria": "Z content validated against X; W completed; normalized Y differs from Y*.",
        "counts": dict(counts),
        "checked": dict(checked),
        "records": records,
    }


def build_index() -> dict:
    indexed = {}
    for pipeline, datasets in load_catalog().items():
        indexed[pipeline] = {}
        for dataset in datasets:
            print(f"Indexing {pipeline} / {dataset}...", flush=True)
            indexed[pipeline][dataset] = load_output_records(pipeline, dataset)
    payload = {"version": INDEX_VERSION, "records": indexed}
    INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    temporary = INDEX_PATH.with_suffix(".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    temporary.replace(INDEX_PATH)
    WEB_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    WEB_INDEX_PATH.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    WEB_SIFLEX_INDEX_PATH.write_text(json.dumps(build_siflex_index(), ensure_ascii=False), encoding="utf-8")
    WEB_TABULAR_MODELS_INDEX_PATH.write_text(json.dumps(build_tabular_models_index(), ensure_ascii=False), encoding="utf-8")
    WEB_MISMATCH_INDEX_PATH.write_text(json.dumps(build_mismatch_index(payload), ensure_ascii=False), encoding="utf-8")
    return payload


def compact_workflow_text(value, limit: int = 220) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text if len(text) <= limit else f"{text[:limit - 3]}..."


def workflow_event(operator: str, status: str, result: str, source: str, line: int | None = None, **extra) -> dict:
    event = {
        "operator": operator,
        "status": status,
        "result": compact_workflow_text(result),
        "source": source,
        "evidenceType": extra.pop("evidence_type", "observed"),
    }
    if line:
        event["line"] = line
    for key, value in extra.items():
        if value is not None and value != "" and value != []:
            event[key] = compact_workflow_text(value) if isinstance(value, str) else value
    return event


@lru_cache(maxsize=256)
def parse_log_entries_cached(path_string: str, modified_ns: int, size: int) -> tuple[tuple[int, str], ...]:
    path = Path(path_string)
    entries = []
    patterns = (
        re.compile(r"^\d{4}-\d{2}-\d{2}[^|]*\|\s*(?:TRACE|DEBUG|INFO|WARNING|ERROR|CRITICAL)\s+\|.*? -\s?(.*)$"),
        re.compile(r"^\d{4}-\d{2}-\d{2}.*? - .*? - (?:DEBUG|INFO|WARNING|ERROR|CRITICAL) -\s?(.*)$"),
    )
    for line_number, raw_line in enumerate(path.read_text(encoding="utf-8", errors="replace").splitlines(), 1):
        match = next((pattern.match(raw_line) for pattern in patterns if pattern.match(raw_line)), None)
        if match:
            entries.append({"line": line_number, "message": match.group(1).strip()})
        elif entries and raw_line.strip():
            entries[-1]["message"] = f"{entries[-1]['message']}\n{raw_line.strip()}".strip()
    return tuple((entry["line"], entry["message"]) for entry in entries)


def parse_log_entries(path: Path) -> list[dict]:
    stat = path.stat()
    return [
        {"line": line, "message": message}
        for line, message in parse_log_entries_cached(str(path), stat.st_mtime_ns, stat.st_size)
    ]


def next_log_value(entries: list[dict], index: int) -> tuple[str, int | None]:
    for candidate in entries[index + 1:index + 5]:
        message = candidate["message"].strip()
        if message:
            return message, candidate["line"]
    return "", None


def find_report_record(record: dict) -> dict:
    source = DATA / str(record.get("source") or "")
    if not source.is_file():
        return {}
    try:
        stat = source.stat()
        payload = json_payload_cached(str(source), stat.st_mtime_ns, stat.st_size)
    except (OSError, ValueError, UnicodeDecodeError):
        return {}
    sample_id = str(record.get("id") or "").rsplit(":", 1)[-1]
    return next(
        (item for item in payload.get("results", []) if str(item.get("sample_id")) == sample_id),
        {},
    )


@lru_cache(maxsize=24)
def json_payload_cached(path_string: str, modified_ns: int, size: int) -> dict:
    return json.loads(Path(path_string).read_text(encoding="utf-8"))


def st_raptor_workflow(record: dict) -> dict:
    workflow_paths = record.get("artifacts", {}).get("workflow", [])
    log_relative = next((path for path in workflow_paths if str(path).lower().endswith(".log")), "")
    log_path = DATA / log_relative
    raw = find_report_record(record)
    if not log_relative or not log_path.is_file():
        payload_relative = next((path for path in workflow_paths if str(path).lower().endswith("payload.json")), "")
        events = []
        if payload_relative:
            payload = {}
            try:
                payload = json.loads((DATA / payload_relative).read_text(encoding="utf-8"))
            except (OSError, ValueError, UnicodeDecodeError):
                payload = {}
            settings = payload.get("settings") if isinstance(payload.get("settings"), dict) else {}
            if not settings:
                raw_metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
                settings = raw_metadata.get("settings") if isinstance(raw_metadata.get("settings"), dict) else {}
            events.append(workflow_event(
                "Configure the ST-Raptor run",
                "info",
                f"The saved payload enables embedding={settings.get('enable_embedding')} and query decomposition={settings.get('enable_query_decompose')}.",
                payload_relative,
                evidence_type="configured",
                evidence="The case has a run payload but no execution log, so operator-level actions cannot be reconstructed.",
            ))
        source = str(record.get("source") or "")
        if raw.get("error"):
            metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
            returncode = metadata.get("returncode")
            events.append(workflow_event(
                "Continue to question answering",
                "failed",
                f"ST-Raptor subprocess failed{f' with exit code {returncode}' if returncode is not None else ''}.",
                source,
                fallback="No QA fallback was recorded; the execution log is missing.",
                evidence=compact_workflow_text(record.get("prediction"), 180),
            ))
        else:
            events.append(workflow_event("Return the recorded answer", "success", f"Returned {record.get('prediction')}.", source))
        return {"events": events, "note": "This record has no case-specific execution log; only saved payload/report evidence is shown."}

    entries = parse_log_entries(log_path)
    events = []

    preprocess_entry = next((entry for entry in entries if "process_table_vlm() Start" in entry["message"]), None)
    dimensions_entry = next((entry for entry in entries if "表格行数:" in entry["message"]), None)
    split_entry = next((entry for entry in entries if "拆分出" in entry["message"] and "子表" in entry["message"]), None)
    schema_entry = next((entry for entry in entries if "模型输出:" in entry["message"]), None)
    if preprocess_entry:
        details = []
        if dimensions_entry:
            dimensions = re.search(r"表格行数:\s*(\d+)\s*表格列数:\s*(\d+)", dimensions_entry["message"])
            if dimensions:
                details.append(f"Detected a {dimensions.group(1)} x {dimensions.group(2)} table")
        if split_entry:
            split_count = re.search(r"拆分出\s*(\d+)\s*个子表", split_entry["message"])
            if split_count:
                details.append(f"split it into {split_count.group(1)} subtable(s)")
        if schema_entry:
            schema = schema_entry["message"].split("模型输出:", 1)[-1].strip()
            if schema:
                details.append(f"VLM schema: {schema}")
        events.append(workflow_event(
            "Build the hierarchical table representation",
            "success",
            "; ".join(details) + "." if details else "Started VLM-based table preprocessing.",
            log_relative,
            preprocess_entry["line"],
        ))

    load_entries = [
        entry for entry in entries
        if "Loading PKL File:" in entry["message"] or "Loading Embedding Cache File:" in entry["message"]
    ]
    if load_entries:
        loaded = [external_filename(entry["message"].split(":", 1)[-1]) for entry in load_entries]
        events.append(workflow_event(
            "Load hierarchical table state",
            "success",
            f"Loaded {', '.join(loaded)}.",
            log_relative,
            load_entries[0]["line"],
            evidence="The case log records both the serialized table and its embedding cache.",
        ))

    decompose_index = next((index for index, entry in enumerate(entries) if "Query Decompose" in entry["message"]), None)
    if decompose_index is not None:
        decomposition, line = next_log_value(entries, decompose_index)
        subquery_count = None
        try:
            parsed = literal_eval(decomposition)
            if isinstance(parsed, list):
                subquery_count = len(parsed)
                nonempty = sum(bool(str(item).strip()) for item in parsed)
                decomposition = f"Created {subquery_count} subqueries; {nonempty} contained query text."
        except (SyntaxError, ValueError):
            pass
        retrieve_index = next((index for index, entry in enumerate(entries[decompose_index:], decompose_index) if "Retreive Flag" in entry["message"]), None)
        retrieve_flags, _ = next_log_value(entries, retrieve_index) if retrieve_index is not None else ("", None)
        events.append(workflow_event(
            "Decompose the query",
            "success",
            decomposition or "The query was decomposed into subqueries.",
            log_relative,
            line,
            evidence=f"Retrieval flags: {retrieve_flags}" if retrieve_flags else None,
        ))

    primitive_failures = [entry for entry in entries if "Primitive Need Regenerate" in entry["message"]]
    primitive_tries = [entry for entry in entries if "Generated Primitive" in entry["message"]]
    if primitive_failures:
        reason = primitive_failures[-1]["message"].split("Primitive Execution Error:", 1)[-1]
        events.append(workflow_event(
            "Generate and execute a primitive",
            "failed",
            f"{len(primitive_failures)} attempt(s) failed: {reason}",
            log_relative,
            primitive_failures[0]["line"],
            fallback=f"Regenerated the primitive up to {len(primitive_tries) or len(primitive_failures)} times.",
        ))
    elif primitive_tries:
        events.append(workflow_event(
            "Generate and execute a primitive",
            "success",
            f"Executed a generated primitive after {len(primitive_tries)} attempt(s).",
            log_relative,
            primitive_tries[0]["line"],
        ))

    entity_entry = next((entry for entry in entries if "Extracted Entities:" in entry["message"]), None)
    matched_entry = next((entry for entry in entries if "Matched Table Content:" in entry["message"]), None)
    if entity_entry or matched_entry:
        events.append(workflow_event(
            "Match query entities to table content",
            "success",
            matched_entry["message"] if matched_entry else "Matched entities against the table.",
            log_relative,
            (entity_entry or matched_entry)["line"],
            evidence=entity_entry["message"] if entity_entry else None,
        ))

    scratch_index = next((index for index, entry in enumerate(entries) if "Try to Reason from Scratch" in entry["message"]), None)
    if scratch_index is not None:
        fallback_result, _ = next_log_value(entries, scratch_index)
        events.append(workflow_event(
            "Depth-first table reasoning",
            "fallback",
            "The DFS operator failed.",
            log_relative,
            entries[scratch_index]["line"],
            fallback="Reasoned again from the full table without the failed DFS path.",
            fallbackResult=fallback_result,
        ))

    answer_index = next((index for index, entry in enumerate(entries) if "Final Answer for Whole Table Reasoning" in entry["message"]), None)
    if answer_index is not None:
        answer, line = next_log_value(entries, answer_index)
        reliability_entry = next((entry for entry in entries[answer_index:] if "Reliabillity:" in entry["message"]), None)
        events.append(workflow_event(
            "Verify and return the final answer",
            "success",
            f"Returned {answer or record.get('prediction')}." + (f" Reliability: {reliability_entry['message'].split(':', 1)[-1].strip()}." if reliability_entry else ""),
            log_relative,
            line or entries[answer_index]["line"],
        ))

    if raw.get("error"):
        metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
        returncode = metadata.get("returncode")
        failure = f"ST-Raptor subprocess failed{f' with exit code {returncode}' if returncode is not None else ''}."
        events.append(workflow_event(
            "Continue to question answering",
            "failed",
            failure,
            str(record.get("source") or log_relative),
            fallback="No QA fallback was recorded; execution stopped before an answer operator completed.",
            evidence=compact_workflow_text(record.get("prediction"), 180),
        ))

    return {"events": events[:7], "note": "Observed from the case-specific ST-Raptor execution log."}


def graphotter_workflow(record: dict) -> dict:
    raw = find_report_record(record)
    metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
    settings = metadata.get("settings") if isinstance(metadata.get("settings"), dict) else {}
    artifacts = record.get("artifacts", {})
    interpreted = artifacts.get("interpreted", [])
    workflow_paths = artifacts.get("workflow", [])
    official_log = next((path for path in workflow_paths if str(path).endswith("graphotter_official.log")), "")
    result_source = str(record.get("source") or "")
    events = []

    table_path = next((path for path in interpreted if str(path).lower().endswith(".json")), "")
    if table_path:
        events.append(workflow_event(
            "Prepare the table graph input",
            "success",
            f"Prepared table {raw.get('table_id') or record.get('representation', {}).get('contextId')} as {table_path.rsplit('/', 1)[-1]}.",
            table_path,
            evidence_type="observed",
        ))

    if official_log and (DATA / official_log).is_file():
        entries = parse_log_entries(DATA / official_log)
        loaded = next((entry for entry in entries if "Dense retriever embedder loaded" in entry["message"]), None)
        indexed = next((entry for entry in entries if "Initialize the index" in entry["message"]), None)
        if loaded:
            events.append(workflow_event(
                "Initialize dense cell retrieval",
                "success",
                loaded["message"],
                official_log,
                loaded["line"],
                evidence="Run-level log; the logger does not attach this initialization line to an individual question.",
            ))
        cache_path = next((path for path in interpreted if str(path).lower().endswith(".pkl")), "")
        if cache_path:
            events.append(workflow_event(
                "Load the cell embedding index",
                "success",
                f"Used the saved embedding cache {cache_path.rsplit('/', 1)[-1]}.",
                cache_path,
                evidence=(f"The run log records index initialization at line {indexed['line']}." if indexed else "The per-table cache is bundled with this record."),
            ))

    max_depth = settings.get("max_iteration_depth")
    events.append(workflow_event(
        "Select cells and expand graph neighbors",
        "info",
        "GraphOtter combines LLM-selected and retriever-selected cells, then follows row and column neighbors.",
        "pipeline/GraphOtter/GraphRetriever/graph_retriver.py",
        evidence_type="documented",
        evidence="This operator sequence comes from the bundled GraphOtter implementation.",
    ))
    events.append(workflow_event(
        "Iterative graph reasoning",
        "info",
        f"Reasoned over the connected cell graph{f' for up to {max_depth} iterations' if max_depth else ''}.",
        "pipeline/GraphOtter/iterative_reasoning.py",
        evidence_type="configured",
        evidence="The report confirms the official pipeline and its iteration setting.",
    ))

    if raw.get("error"):
        events.append(workflow_event("Return the answer", "failed", raw.get("error"), result_source))
    else:
        latency = raw.get("latency")
        events.append(workflow_event(
            "Generate the final answer",
            "success",
            f"Returned {record.get('prediction')}." + (f" Completed in {float(latency):.2f}s." if isinstance(latency, (int, float)) else ""),
            result_source,
            evidence_type="observed",
        ))

    return {
        "events": events[:7],
        "note": "GraphOtter workflow stages combine observed run artifacts with the configured operator sequence from the bundled implementation.",
    }


def spreadsheet_agent_workflow(record: dict) -> dict:
    raw = find_report_record(record)
    artifacts = record.get("artifacts", {})
    workflow_paths = artifacts.get("workflow", [])
    interpreted = artifacts.get("interpreted", [])
    manifest_relative = next((path for path in workflow_paths if str(path).endswith("manifest.json")), "")
    structure_relative = next((path for path in interpreted if str(path).endswith("input_table_siflex.json")), "")
    events = []

    manifest = {}
    if manifest_relative and (DATA / manifest_relative).is_file():
        try:
            manifest = json.loads((DATA / manifest_relative).read_text(encoding="utf-8"))
        except (OSError, ValueError, UnicodeDecodeError):
            manifest = {}
    if manifest:
        sheets = manifest.get("sheet_names") if isinstance(manifest.get("sheet_names"), list) else []
        events.append(workflow_event(
            "Cache and inspect the workbook",
            "success",
            f"Opened the workbook and found {len(sheets)} sheet(s): {', '.join(map(str, sheets)) or 'unnamed'}.",
            manifest_relative,
            evidence=f"Extractor requested: {manifest.get('descriptor', {}).get('extractor') or 'unknown'}.",
        ))

    structure = {}
    if structure_relative and (DATA / structure_relative).is_file():
        try:
            structure = json.loads((DATA / structure_relative).read_text(encoding="utf-8"))
        except (OSError, ValueError, UnicodeDecodeError):
            structure = {}
    extraction_error = structure.get("error")
    if extraction_error:
        events.append(workflow_event(
            "Extract a structured spreadsheet representation",
            "fallback",
            extraction_error,
            structure_relative,
            fallback="Converted the workbook to a Markdown table and used that as Z.",
            fallbackResult=f"Fallback representation contains {len(str(structure.get('structure') or ''))} characters.",
        ))
    elif structure:
        events.append(workflow_event(
            "Extract a structured spreadsheet representation",
            "success",
            "Produced the saved SIFLEX structure used for question answering.",
            structure_relative,
        ))

    source = str(record.get("source") or "")
    if raw.get("error"):
        events.append(workflow_event("Answer from the spreadsheet context", "failed", raw.get("error"), source))
    else:
        latency = raw.get("latency")
        tokens = raw.get("tokens") if isinstance(raw.get("tokens"), dict) else {}
        details = []
        if isinstance(latency, (int, float)):
            details.append(f"{float(latency):.2f}s")
        if tokens.get("completion") is not None:
            details.append(f"{tokens['completion']} completion tokens")
        events.append(workflow_event(
            "Answer from the available table context",
            "success",
            f"Returned {record.get('prediction')}." + (f" Run evidence: {', '.join(details)}." if details else ""),
            source,
            evidence="The report stores the final output, but this bundle does not contain intermediate answer-generation operators.",
        ))

    return {
        "events": events[:6],
        "note": "SpreadsheetAgent records preprocessing and fallback evidence; intermediate answer-generation actions are not bundled.",
    }


def workflow_summary(record: dict) -> dict:
    pipeline = record.get("pipeline")
    if pipeline == "ST-raptor":
        return st_raptor_workflow(record)
    if pipeline == "GraphOtter":
        return graphotter_workflow(record)
    if pipeline == "SpreadsheetAgent":
        return spreadsheet_agent_workflow(record)
    return {"events": [], "note": "A structured workflow summary is not available for this pipeline."}


@lru_cache(maxsize=1)
def load_index() -> dict:
    if INDEX_PATH.exists():
        try:
            payload = json.loads(INDEX_PATH.read_text(encoding="utf-8"))
            if payload.get("version") == INDEX_VERSION:
                return payload
        except (OSError, ValueError):
            pass
    return build_index()


class Handler(BaseHTTPRequestHandler):
    @staticmethod
    def data_path(relative: str) -> Path | None:
        normalized = unquote(relative).replace("\\", "/")
        path = (DATA / normalized).resolve()
        try:
            path.relative_to(DATA.resolve())
        except ValueError:
            return None
        return path

    def send_json(self, status: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        request = urlparse(self.path)
        if request.path == "/api/health":
            payload = {"ok": True}
        elif request.path == "/api/catalog":
            catalog = load_catalog()
            payload = {
                "catalog": catalog,
                "default": default_selection(catalog),
            }
        elif request.path == "/api/records":
            params = parse_qs(request.query)
            pipeline = params.get("pipeline", [""])[0]
            dataset = params.get("dataset", [""])[0]
            records = load_index().get("records", {}).get(pipeline, {}).get(dataset, [])
            query = params.get("q", [""])[0].strip().casefold()
            try:
                offset = max(0, int(params.get("offset", ["0"])[0]))
                limit = min(100, max(1, int(params.get("limit", ["40"])[0])))
            except ValueError:
                self.send_error(400, "offset and limit must be integers")
                return
            matches = [
                record for record in records
                if record["pipeline"] == pipeline
                and record["dataset"] == dataset
                and (not query or query in str(record["question"]).casefold())
            ]
            totals = Counter(record["status"] for record in records)
            payload = {
                "records": matches[offset:offset + limit],
                "total": len(matches),
                "offset": offset,
                "totals": {status: totals.get(status, 0) for status in ("correct", "wrong", "error")},
            }
        elif request.path == "/api/workflow-summary":
            params = parse_qs(request.query)
            pipeline = params.get("pipeline", [""])[0]
            dataset = params.get("dataset", [""])[0]
            record_id = params.get("record", [""])[0]
            candidates = load_index().get("records", {}).get(pipeline, {}).get(dataset, [])
            record = next((item for item in candidates if item.get("id") == record_id), None)
            if record is None:
                self.send_json(404, {"error": "The requested workflow record does not exist."})
                return
            payload = workflow_summary(record)
        elif request.path == "/api/pkl-summary":
            params = parse_qs(request.query)
            relative = params.get("path", [""])[0]
            path = self.data_path(relative)
            if path is None:
                self.send_json(403, {"error": "The requested path is outside the data directory."})
                return
            if path.suffix.lower() != ".pkl" or not path.is_file():
                self.send_json(404, {"error": "The requested PKL artifact does not exist."})
                return
            try:
                stat = path.stat()
                payload = copy.deepcopy(pickle_summary_cached(str(path), stat.st_mtime_ns, stat.st_size))
                context_relative = params.get("context", [""])[0]
                context_path = self.data_path(context_relative) if context_relative else None
                context_id = params.get("context_id", [""])[0]
                table_index_text = params.get("table_index", [""])[0]
                try:
                    table_index = int(table_index_text) if table_index_text else None
                except ValueError:
                    table_index = None
                if context_path and context_path.is_file() and context_path.suffix.lower() == ".json":
                    attach_embedding_table(payload, graphotter_table_context(context_path, context_id, table_index))
                payload["path"] = path.relative_to(DATA).as_posix()
            except (OSError, MemoryError) as error:
                self.send_json(500, {"error": f"Could not inspect this PKL: {error}"})
                return
        elif request.path == "/api/yaml-summary":
            params = parse_qs(request.query)
            relative = params.get("path", [""])[0]
            path = self.data_path(relative)
            source_relative = params.get("source", [""])[0]
            source_path = self.data_path(source_relative) if source_relative else None
            if path is None or (source_relative and source_path is None):
                self.send_json(403, {"error": "The requested path is outside the data directory."})
                return
            if path.suffix.lower() not in {".yaml", ".yml"} or not path.is_file():
                self.send_json(404, {"error": "The requested YAML artifact does not exist."})
                return
            if source_path and (source_path.suffix.lower() not in {".xlsx", ".xlsm"} or not source_path.is_file()):
                self.send_json(404, {"error": "The requested source workbook does not exist."})
                return
            try:
                payload = yaml_structure_summary(path, source_path)
            except (OSError, ValueError, ImportError) as error:
                self.send_json(500, {"error": f"Could not inspect this YAML structure: {error}"})
                return
        elif request.path.startswith("/api/files/"):
            path = self.data_path(request.path.removeprefix("/api/files/"))
            if path is None:
                self.send_error(403)
                return
            if not path.is_file():
                self.send_error(404)
                return
            content_type = {
                ".jsonl": "application/x-ndjson; charset=utf-8",
                ".log": "text/plain; charset=utf-8",
            }.get(path.suffix.lower(), mimetypes.guess_type(path.name)[0] or "application/octet-stream")
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(path.stat().st_size))
            self.end_headers()
            with path.open("rb") as handle:
                shutil.copyfileobj(handle, self.wfile)
            return
        else:
            self.send_error(404)
            return
        body = json.dumps(payload, ensure_ascii=False).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        request = urlparse(self.path)
        if request.path != "/api/reveal":
            self.send_error(404)
            return
        relative = parse_qs(request.query).get("path", [""])[0]
        path = self.data_path(relative)
        if path is None:
            self.send_json(403, {"error": "The requested path is outside the data directory."})
            return
        if not path.exists():
            self.send_json(404, {"error": "The requested artifact does not exist."})
            return
        try:
            reveal_file(path)
        except OSError as error:
            self.send_json(500, {"error": f"Could not open the file manager: {error}"})
            return
        self.send_json(200, {"ok": True})

    def log_message(self, *_):
        return


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Run the artifact viewer API")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8766)
    parser.add_argument("--build-index", action="store_true")
    parser.add_argument("--build-siflex-index", action="store_true")
    parser.add_argument("--build-mismatch-index", action="store_true")
    parser.add_argument("--ready-file", type=Path)
    args = parser.parse_args()
    if args.build_index:
        INDEX_PATH.unlink(missing_ok=True)
        load_records.cache_clear()
        load_index.cache_clear()
        build_index()
        return
    if args.build_siflex_index:
        WEB_SIFLEX_INDEX_PATH.write_text(json.dumps(build_siflex_index(), ensure_ascii=False), encoding="utf-8")
        return
    if args.build_mismatch_index:
        WEB_MISMATCH_INDEX_PATH.write_text(json.dumps(build_mismatch_index(), ensure_ascii=False), encoding="utf-8")
        return
    print("Preparing QA index...", flush=True)
    load_index()
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    if args.ready_file:
        args.ready_file.write_text(str(server.server_port), encoding="ascii")
    print(f"Artifact API: http://{args.host}:{args.port}/api/catalog", flush=True)
    try:
        server.serve_forever()
    finally:
        if args.ready_file:
            args.ready_file.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
