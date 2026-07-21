#!/usr/bin/env python3
"""Classify failed table-QA cases and export auditable failure-mode reports."""

from __future__ import annotations

import argparse
import csv
import html
import io
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from api import server as artifact_server  # noqa: E402


DATA = ROOT / "data"
DEFAULT_OUTPUT = ROOT / "analysis" / "failure_modes"
PIPELINES = ("GraphOtter", "ST-raptor", "SpreadsheetAgent")
DATASETS = ("HiTab", "MultiHiertt")
DISPLAY_DATASET = {"HiTab": "Hitab", "MultiHiertt": "mulhi"}
CASE_COLUMNS = (
    "benchmark",
    "solution",
    "sample_id",
    "outcome",
    "classification",
    "misalignment",
    "misinterpretation",
    "failed_code",
    "question",
    "gold_answer_y_star",
    "predicted_answer_y",
    "w_status",
    "w_error",
    "z_created",
    "z_matches_x",
    "z_structural_match",
    "z_scope_match",
    "expected_x_tables",
    "selected_z_tables",
    "gold_table_evidence",
    "x_z_evidence",
    "x_preview",
    "z_preview",
    "llm_judge_reason",
    "x_paths",
    "z_paths",
    "w_paths",
    "y_paths",
    "source_report",
    "human_verdict",
    "human_notes",
)


class HtmlTableParser(artifact_server.CompactHtmlTableParser):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def report_number(path: Path) -> int:
    suffix = path.stem.removeprefix("report_")
    return int(suffix) if suffix.isdigit() else -1


def latest_report(pipeline: str, dataset: str) -> Path:
    reports = sorted((DATA / "Outputs_llm" / pipeline / dataset).glob("report_*.json"), key=report_number)
    if not reports:
        raise FileNotFoundError(f"No evaluated report for {pipeline}/{dataset}")
    return reports[-1]


def root_path(raw_path: str) -> Path:
    normalized = raw_path.replace("\\", "/")
    if normalized.startswith("data/"):
        return ROOT / normalized
    return DATA / normalized


def normalized_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip().casefold()


def trim_matrix(matrix: list[list[str]]) -> list[list[str]]:
    matrix = [list(row) for row in matrix]
    while matrix and not any(matrix[-1]):
        matrix.pop()
    width = max(
        (max((index for index, value in enumerate(row) if value), default=-1) + 1 for row in matrix),
        default=0,
    )
    return [row[:width] for row in matrix]


def workbook_sheets(path: Path, expand_merged: bool = False) -> tuple[list[str], list[list[list[str]]]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    names = list(workbook.sheetnames)
    matrices = []
    try:
        for sheet in workbook.worksheets:
            matrix = [
                [normalized_cell(sheet.cell(row=row, column=column).value) for column in range(1, sheet.max_column + 1)]
                for row in range(1, sheet.max_row + 1)
            ]
            if expand_merged:
                for merged in sheet.merged_cells.ranges:
                    value = matrix[merged.min_row - 1][merged.min_col - 1]
                    for row in range(merged.min_row - 1, merged.max_row):
                        for column in range(merged.min_col - 1, merged.max_col):
                            matrix[row][column] = value
            matrices.append(trim_matrix(matrix))
    finally:
        workbook.close()
    return names, matrices


def html_matrix(value: str) -> list[list[str]]:
    parser = HtmlTableParser()
    parser.feed(value)
    return trim_matrix([[normalized_cell(cell) for cell in row] for row in parser.matrix()])


def matrix_preview(matrix: list[list[str]], rows: int = 4) -> str:
    return json.dumps(matrix[:rows], ensure_ascii=False, separators=(",", ":"))


def tables_preview(matrices: dict[int | str, list[list[str]]]) -> str:
    payload = {str(key): value[:4] for key, value in matrices.items()}
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def parse_selected_tables(raw: dict) -> list[int]:
    metadata = raw.get("metadata") if isinstance(raw.get("metadata"), dict) else {}
    preprocess = metadata.get("preprocess") if isinstance(metadata.get("preprocess"), dict) else {}
    selected = []
    for source in preprocess.get("selected_sources") or []:
        match = re.search(r"table\[(\d+)\]", str(source))
        if match:
            selected.append(int(match.group(1)))
    return sorted(set(selected))


def expected_tables(mulhi_record: dict | None) -> tuple[list[int], list[str]]:
    qa = mulhi_record.get("qa") if isinstance(mulhi_record, dict) and isinstance(mulhi_record.get("qa"), dict) else {}
    evidence = [str(item) for item in qa.get("table_evidence") or []]
    expected = sorted({int(match.group(1)) for item in evidence if (match := re.match(r"^(\d+)-", item))})
    return expected, evidence


def pickle_summary(path: Path) -> dict:
    stat = path.stat()
    return artifact_server.pickle_summary_cached(str(path), stat.st_mtime_ns, stat.st_size)


def straptor_tree_audit(path: Path, source_matrix: list[list[str]]) -> dict:
    loader = artifact_server.RestrictedPickleLoader(io.BytesIO(path.read_bytes()))
    feature_tree = loader.load()
    all_values = set()
    root_header_values = set()
    coordinates = []
    seen = set()

    def walk_node(node, collect_header: bool = False):
        if node is None or id(node) in seen:
            return
        seen.add(id(node))
        if isinstance(node, artifact_server.SafeFeatureTree):
            walk_index(getattr(getattr(node, "index_tree", None), "root", None), collect_header)
            walk_node(getattr(getattr(node, "body_tree", None), "root", None))
            return
        value = getattr(node, "value", None)
        if isinstance(value, artifact_server.SafeFeatureTree):
            walk_node(value)
        elif normalized_cell(value):
            all_values.add(normalized_cell(value))
        if isinstance(node, artifact_server.SafeBodyNode):
            position = tuple(getattr(node, name, None) for name in ("x1", "y1", "x2", "y2"))
            if any(item is not None for item in position):
                coordinates.append((position, normalized_cell(value)))
        for attribute in ("children", "body"):
            for child in list(getattr(node, attribute, None) or []):
                walk_node(child)

    def walk_index(node, collect_header: bool):
        if node is None or id(node) in seen:
            return
        seen.add(id(node))
        value = getattr(node, "value", None)
        normalized = normalized_cell(value)
        if normalized:
            all_values.add(normalized)
            if collect_header:
                root_header_values.add(normalized)
        for child in list(getattr(node, "children", None) or []):
            walk_index(child, collect_header)
        for body_node in list(getattr(node, "body", None) or []):
            walk_node(body_node)

    walk_index(getattr(getattr(feature_tree, "index_tree", None), "root", None), True)
    walk_node(getattr(getattr(feature_tree, "body_tree", None), "root", None))

    source_values = {value for row in source_matrix for value in row if value}
    missing_values = sorted(source_values - all_values)
    value_coverage = (len(source_values) - len(missing_values)) / max(1, len(source_values))
    header_values = {value for value in (source_matrix[0] if source_matrix else []) if value}
    missing_headers = sorted(header_values - root_header_values)
    header_coverage = (len(header_values) - len(missing_headers)) / max(1, len(header_values))

    body_match = None
    if coordinates:
        max_row = max(int(position[0] or 0) for position, _ in coordinates)
        max_column = max(int(position[1] or 0) for position, _ in coordinates)
        coordinate_matrix = [["" for _ in range(max_column)] for _ in range(max_row)]
        for position, value in coordinates:
            row, column = int(position[0] or 0), int(position[1] or 0)
            if row >= 1 and column >= 1:
                coordinate_matrix[row - 1][column - 1] = value
        body_match = trim_matrix(coordinate_matrix) == trim_matrix(source_matrix[1:])

    if body_match is None:
        matches = value_coverage == 1
        method = f"nested_value_coverage={len(source_values) - len(missing_values)}/{len(source_values)}"
    else:
        matches = body_match and value_coverage == 1
        method = (
            f"coordinate_body_match={body_match}; "
            f"overall_value_coverage={len(source_values) - len(missing_values)}/{len(source_values)}"
        )
    return {
        "matches": matches,
        "method": method,
        "value_coverage": value_coverage,
        "header_coverage": header_coverage,
        "missing_values": missing_values[:12],
        "missing_headers": missing_headers[:12],
    }


def locate_artifacts(raw: dict, pipeline: str, dataset: str, source_report: Path) -> dict[str, list[str]]:
    return artifact_server.artifact_paths(raw, pipeline, dataset, source_report)


def existing_paths(paths: list[str], suffix: str | None = None) -> list[Path]:
    candidates = [DATA / path for path in paths]
    return [path for path in candidates if path.is_file() and (suffix is None or path.suffix.casefold() == suffix)]


def base_validation(expected: list[int | str], selected: list[int | str]) -> dict:
    return {
        "created": False,
        "structural_match": False,
        "scope_match": False,
        "matches": False,
        "expected": expected,
        "selected": selected,
        "evidence": "Z could not be validated.",
        "x_preview": "",
        "z_preview": "",
    }


def validate_graphotter_hitab(raw: dict, artifacts: dict) -> dict:
    table_id = str(raw.get("table_id") or "")
    result = base_validation([table_id], [table_id])
    x_path = DATA / "Datasets" / "HiTab_xlsx" / f"{table_id}.xlsx"
    z_paths = existing_paths(artifacts["interpreted"], ".json")
    pkl_paths = existing_paths(artifacts["interpreted"], ".pkl")
    if not x_path.is_file() or not z_paths or not pkl_paths:
        result["evidence"] = "Missing X workbook, GraphOtter table JSON, or embedding cache."
        return result

    _, x_sheets = workbook_sheets(x_path)
    payload = json.loads(z_paths[0].read_text(encoding="utf-8"))
    z_matrix = trim_matrix([[normalized_cell(value) for value in row] for row in payload.get("texts") or []])
    x_matrix = x_sheets[0]
    z_title = normalized_cell(payload.get("title"))
    x_title = next((value for value in x_matrix[0] if value), "") if x_matrix else ""
    if z_matrix == x_matrix and not z_title:
        title_match = True
        table_match = True
        comparison_mode = "titleless full matrix"
    else:
        title_match = z_title == x_title
        table_match = z_matrix == trim_matrix(x_matrix[1:])
        comparison_mode = "separate title plus table matrix"
    cache = pickle_summary(pkl_paths[0])
    cache_cells = (cache.get("rowIds") or {}).get("count")
    cache_match = cache_cells == sum(len(row) for row in z_matrix)
    result.update({
        "created": bool(z_matrix) and cache.get("safeLoad") is True,
        "structural_match": title_match and table_match and cache_match,
        "scope_match": True,
        "matches": title_match and table_match and cache_match,
        "evidence": (
            f"Single-table HiTab ({comparison_mode}): title_match={title_match}; table_matrix_match={table_match}; "
            f"embedding_cells={cache_cells}/{sum(len(row) for row in z_matrix)}."
        ),
        "x_preview": matrix_preview(x_matrix),
        "z_preview": matrix_preview(z_matrix),
    })
    return result


def validate_straptor_hitab(raw: dict, artifacts: dict) -> dict:
    table_id = str(raw.get("table_id") or "")
    result = base_validation([table_id], [table_id])
    x_path = DATA / "Datasets" / "HiTab_xlsx" / f"{table_id}.xlsx"
    html_paths = existing_paths(artifacts["interpreted"], ".html")
    pkl_paths = existing_paths(artifacts["interpreted"], ".pkl")
    if not x_path.is_file() or not html_paths or not pkl_paths:
        result["evidence"] = "Missing X workbook, ST-Raptor HTML, or HO-Tree pickle."
        return result

    _, x_sheets = workbook_sheets(x_path, expand_merged=True)
    x_matrix = x_sheets[0]
    z_matrix = html_matrix(html_paths[0].read_text(encoding="utf-8"))
    html_match = z_matrix == trim_matrix(x_matrix[1:])
    tree = pickle_summary(pkl_paths[0])
    tree_audit = straptor_tree_audit(pkl_paths[0], z_matrix)
    tree_match = tree_audit["matches"]
    structural = html_match and tree_match and tree.get("safeLoad") is True
    result.update({
        "created": bool(z_matrix) and tree.get("safeLoad") is True,
        "structural_match": structural,
        "scope_match": True,
        "matches": structural,
        "evidence": (
            f"Single-table HiTab: HTML_matrix_match={html_match}; HO-Tree_match={tree_match}; "
            f"{tree_audit['method']}; missing={tree_audit['missing_values'] or tree_audit['missing_headers']}."
        ),
        "x_preview": matrix_preview(x_matrix),
        "z_preview": matrix_preview(z_matrix),
    })
    return result


def validate_graphotter_mulhi(raw: dict, artifacts: dict, mulhi_record: dict) -> dict:
    expected, _ = expected_tables(mulhi_record)
    selected = parse_selected_tables(raw)
    result = base_validation(expected, selected)
    x_path = DATA / "Datasets" / "MultiHiertt_xlsx" / f"{raw.get('sample_id')}.xlsx"
    pkl_paths = existing_paths(artifacts["interpreted"], ".pkl")
    tables = mulhi_record.get("tables") if isinstance(mulhi_record.get("tables"), list) else []
    if not x_path.is_file() or len(selected) != 1 or selected[0] >= len(tables) or not pkl_paths:
        result["evidence"] = "Missing X workbook, selected source table, or GraphOtter embedding cache."
        return result

    _, x_sheets = workbook_sheets(x_path, expand_merged=True)
    selected_index = selected[0]
    z_matrix = html_matrix(str(tables[selected_index]))
    selected_matrix_match = selected_index < len(x_sheets) and z_matrix == x_sheets[selected_index]
    cache = pickle_summary(pkl_paths[0])
    cache_cells = (cache.get("rowIds") or {}).get("count")
    cache_match = cache_cells == sum(len(row) for row in z_matrix)
    structural = selected_matrix_match and cache_match and cache.get("safeLoad") is True
    scope = bool(expected) and selected == expected
    expected_matrices = {index: x_sheets[index] for index in expected if index < len(x_sheets)}
    result.update({
        "created": bool(z_matrix) and cache.get("safeLoad") is True,
        "structural_match": structural,
        "scope_match": scope,
        "matches": structural and scope,
        "evidence": (
            f"MulHi expected table(s)={expected}; selected Z table(s)={selected}; scope_match={scope}; "
            f"selected_matrix_match={selected_matrix_match}; embedding_cells={cache_cells}/{sum(len(row) for row in z_matrix)}."
        ),
        "x_preview": tables_preview(expected_matrices),
        "z_preview": tables_preview({selected_index: z_matrix}),
    })
    return result


def validate_straptor_mulhi(raw: dict, artifacts: dict, mulhi_record: dict) -> dict:
    expected, _ = expected_tables(mulhi_record)
    selected = parse_selected_tables(raw)
    result = base_validation(expected, selected)
    x_path = DATA / "Datasets" / "MultiHiertt_xlsx" / f"{raw.get('sample_id')}.xlsx"
    html_paths = existing_paths(artifacts["interpreted"], ".html")
    pkl_paths = existing_paths(artifacts["interpreted"], ".pkl")
    if not x_path.is_file() or len(selected) != 1 or not html_paths or not pkl_paths:
        result["evidence"] = "Missing X workbook, selected table metadata, ST-Raptor HTML, or HO-Tree pickle."
        return result

    _, x_sheets = workbook_sheets(x_path, expand_merged=True)
    selected_index = selected[0]
    z_matrix = html_matrix(html_paths[0].read_text(encoding="utf-8"))
    html_match = selected_index < len(x_sheets) and z_matrix == x_sheets[selected_index]
    tree = pickle_summary(pkl_paths[0])
    tree_audit = straptor_tree_audit(pkl_paths[0], z_matrix)
    tree_match = tree_audit["matches"]
    structural = html_match and tree_match and tree.get("safeLoad") is True
    scope = bool(expected) and selected == expected
    expected_matrices = {index: x_sheets[index] for index in expected if index < len(x_sheets)}
    result.update({
        "created": bool(z_matrix) and tree.get("safeLoad") is True,
        "structural_match": structural,
        "scope_match": scope,
        "matches": structural and scope,
        "evidence": (
            f"MulHi expected table(s)={expected}; selected Z table(s)={selected}; scope_match={scope}; "
            f"HTML_matrix_match={html_match}; HO-Tree_match={tree_match}; {tree_audit['method']}; "
            f"missing={tree_audit['missing_values'] or tree_audit['missing_headers']}."
        ),
        "x_preview": tables_preview(expected_matrices),
        "z_preview": tables_preview({selected_index: z_matrix}),
    })
    return result


def validate_spreadsheet(raw: dict, artifacts: dict, dataset: str, mulhi_record: dict | None) -> dict:
    sample_id = str(raw.get("sample_id") or "")
    table_id = str(raw.get("table_id") or sample_id)
    x_path = DATA / "Datasets" / f"{dataset}_xlsx" / f"{table_id}.xlsx"
    expected, _ = expected_tables(mulhi_record) if dataset == "MultiHiertt" else ([table_id], [])
    result = base_validation(expected, [])
    z_paths = existing_paths(artifacts["interpreted"], ".json")
    if not x_path.is_file() or not z_paths:
        result["evidence"] = "Missing X workbook or SpreadsheetAgent representation JSON."
        return result

    sheet_names, x_sheets = workbook_sheets(x_path)
    payload = json.loads(z_paths[0].read_text(encoding="utf-8"))
    structure = str(payload.get("structure") or "")
    represented_names = re.findall(r"^### Sheet:\s*(.+?)\s*$", structure, flags=re.MULTILINE)
    normalized_structure = normalized_cell(structure.replace("\\|", "|").replace("<br>", " "))
    populated_cells = [value for matrix in x_sheets for row in matrix for value in row if value]
    matched_cells = sum(value in normalized_structure for value in populated_cells)
    coverage = matched_cells / max(1, len(populated_cells))
    sheets_match = represented_names == sheet_names
    structural = bool(structure) and sheets_match and coverage == 1
    selected: list[int | str]
    if dataset == "MultiHiertt":
        selected = list(range(len(sheet_names)))
        scope = bool(expected) and set(expected).issubset(selected)
        expected_matrices = {index: x_sheets[index] for index in expected if index < len(x_sheets)}
    else:
        selected = [table_id]
        scope = True
        expected_matrices = {table_id: x_sheets[0]}
    fallback = bool(payload.get("error")) and "fallback" in structure.casefold()
    result.update({
        "created": bool(structure),
        "structural_match": structural,
        "scope_match": scope,
        "matches": structural and scope,
        "expected": expected,
        "selected": selected,
        "evidence": (
            f"Workbook fallback used={fallback}; represented_sheets={represented_names}; expected_sheets={sheet_names}; "
            f"nonempty_cell_coverage={matched_cells}/{len(populated_cells)}; scope_match={scope}."
        ),
        "x_preview": tables_preview(expected_matrices),
        "z_preview": structure[:1600],
    })
    return result


def validate_z(raw: dict, pipeline: str, dataset: str, artifacts: dict, mulhi_record: dict | None) -> dict:
    if pipeline == "GraphOtter" and dataset == "HiTab":
        return validate_graphotter_hitab(raw, artifacts)
    if pipeline == "ST-raptor" and dataset == "HiTab":
        return validate_straptor_hitab(raw, artifacts)
    if pipeline == "GraphOtter":
        return validate_graphotter_mulhi(raw, artifacts, mulhi_record or {})
    if pipeline == "ST-raptor":
        return validate_straptor_mulhi(raw, artifacts, mulhi_record or {})
    return validate_spreadsheet(raw, artifacts, dataset, mulhi_record)


def json_text(value) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def paths_text(paths: list[str]) -> str:
    return "\n".join(paths)


def build_cases() -> tuple[list[dict], list[dict]]:
    mulhi_payload = json.loads((DATA / "Datasets" / "MultiHiertt" / "dev.json").read_text(encoding="utf-8"))
    mulhi_by_id = {str(item.get("uid")): item for item in mulhi_payload}
    cases = []
    summaries = []

    for pipeline in PIPELINES:
        for dataset in DATASETS:
            llm_path = latest_report(pipeline, dataset)
            llm_payload = json.loads(llm_path.read_text(encoding="utf-8"))
            source_path = root_path(str(llm_payload.get("source_report") or ""))
            source_payload = json.loads(source_path.read_text(encoding="utf-8"))
            source_rows = source_payload.get("results") or []
            group_cases = []

            for llm_index, judged in enumerate(llm_payload.get("results") or []):
                source_index = judged.get("source_result_index", llm_index)
                if not isinstance(source_index, int) or not 0 <= source_index < len(source_rows):
                    raise IndexError(f"Invalid source_result_index in {llm_path}: {source_index}")
                raw = source_rows[source_index]
                source_error = judged.get("evaluation_method") == "source_error" or bool(raw.get("error"))
                wrong_answer = judged.get("resolved_correct") is False and not source_error
                if not source_error and not wrong_answer:
                    continue

                sample_id = str(raw.get("sample_id") or judged.get("sample_id") or source_index)
                mulhi_record = mulhi_by_id.get(sample_id) if dataset == "MultiHiertt" else None
                artifacts = locate_artifacts(raw, pipeline, dataset, source_path)
                validation = validate_z(raw, pipeline, dataset, artifacts, mulhi_record)
                _, table_evidence = expected_tables(mulhi_record)

                if source_error:
                    classification = "failed_code"
                elif validation["matches"]:
                    classification = "misalignment"
                else:
                    classification = "misinterpretation"

                llm_judge = judged.get("llm_judge") if isinstance(judged.get("llm_judge"), dict) else {}
                w_error = raw.get("predicted_answer") if source_error else ""
                row = {
                    "benchmark": DISPLAY_DATASET[dataset],
                    "solution": pipeline,
                    "sample_id": sample_id,
                    "outcome": "execution_failed" if source_error else "wrong_answer",
                    "classification": classification,
                    "misalignment": int(classification == "misalignment"),
                    "misinterpretation": int(classification == "misinterpretation"),
                    "failed_code": int(classification == "failed_code"),
                    "question": raw.get("question") or judged.get("question") or "",
                    "gold_answer_y_star": json_text(raw.get("gold_answer")),
                    "predicted_answer_y": "" if raw.get("predicted_answer") is None else str(raw.get("predicted_answer")),
                    "w_status": "failed/incomplete" if source_error else "completed",
                    "w_error": "" if not w_error else str(w_error),
                    "z_created": validation["created"],
                    "z_matches_x": validation["matches"],
                    "z_structural_match": validation["structural_match"],
                    "z_scope_match": validation["scope_match"],
                    "expected_x_tables": json_text(validation["expected"]),
                    "selected_z_tables": json_text(validation["selected"]),
                    "gold_table_evidence": json_text(table_evidence),
                    "x_z_evidence": validation["evidence"],
                    "x_preview": validation["x_preview"],
                    "z_preview": validation["z_preview"],
                    "llm_judge_reason": str(llm_judge.get("reason") or ""),
                    "x_paths": paths_text(artifacts["input"]),
                    "z_paths": paths_text(artifacts["interpreted"]),
                    "w_paths": paths_text(artifacts["workflow"]),
                    "y_paths": paths_text(artifacts["output"]),
                    "source_report": source_path.relative_to(ROOT).as_posix(),
                    "human_verdict": "",
                    "human_notes": "",
                }
                cases.append(row)
                group_cases.append(row)

            counts = Counter(row["classification"] for row in group_cases)
            wrong_count = counts["misalignment"] + counts["misinterpretation"]
            fail_count = len(group_cases)
            expected_wrong = int(llm_payload.get("wrong") or 0)
            expected_errors = int(llm_payload.get("sourceErrors") or 0)
            if wrong_count != expected_wrong or counts["failed_code"] != expected_errors:
                raise ValueError(
                    f"Count mismatch for {pipeline}/{dataset}: classified wrong={wrong_count}, errors={counts['failed_code']}; "
                    f"report wrong={expected_wrong}, sourceErrors={expected_errors}"
                )
            summaries.append({
                "benchmark": DISPLAY_DATASET[dataset],
                "solution": pipeline,
                "total_wrong_cases": wrong_count,
                "misalignment_cases": counts["misalignment"],
                "misalignment_ratio": counts["misalignment"] / wrong_count if wrong_count else 0,
                "misinterpretation_cases": counts["misinterpretation"],
                "misinterpretation_ratio": counts["misinterpretation"] / wrong_count if wrong_count else 0,
                "total_fail_cases": fail_count,
                "failed_code_cases": counts["failed_code"],
                "failed_code_ratio": counts["failed_code"] / fail_count if fail_count else 0,
                "source_report": source_path.relative_to(ROOT).as_posix(),
            })
    return cases, summaries


def write_csv(path: Path, rows: list[dict], columns: tuple[str, ...] | list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def style_sheet(sheet, widths: dict[str, int], freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="17324D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_xlsx(path: Path, summaries: list[dict], cases: list[dict]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_columns = list(summaries[0])
    summary_sheet.append(summary_columns)
    for row in summaries:
        summary_sheet.append([row[column] for column in summary_columns])
    for row in range(2, summary_sheet.max_row + 1):
        for column in (5, 7, 10):
            summary_sheet.cell(row, column).number_format = "0.00%"
    style_sheet(summary_sheet, {"A": 14, "B": 22, "C": 18, "D": 20, "E": 20, "F": 22, "G": 22, "H": 18, "I": 18, "J": 18, "K": 48})

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Failure-mode ratios"
    chart.y_axis.title = "Ratio"
    chart.y_axis.scaling.max = 1
    chart.x_axis.title = "Benchmark / solution"
    categories = Reference(summary_sheet, min_col=2, min_row=2, max_row=summary_sheet.max_row)
    for column in (5, 7, 10):
        data = Reference(summary_sheet, min_col=column, max_col=column, min_row=1, max_row=summary_sheet.max_row)
        chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 18
    summary_sheet.add_chart(chart, "M2")

    def add_cases_sheet(title: str, rows: list[dict]) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(list(CASE_COLUMNS))
        for item in rows:
            sheet.append([item[column] for column in CASE_COLUMNS])
        style_sheet(sheet, {
            "A": 12, "B": 20, "C": 34, "D": 18, "E": 20, "F": 14, "G": 17, "H": 13,
            "I": 64, "J": 24, "K": 30, "L": 18, "M": 48, "N": 12, "O": 12, "P": 18,
            "Q": 14, "R": 22, "S": 22, "T": 28, "U": 70, "V": 65, "W": 65, "X": 55,
            "Y": 55, "Z": 55, "AA": 55, "AB": 55, "AC": 48, "AD": 20, "AE": 40,
        })
        classification_column = CASE_COLUMNS.index("classification") + 1
        column_letter = get_column_letter(classification_column)
        end_row = max(2, sheet.max_row)
        colors = {"misalignment": "FFF2CC", "misinterpretation": "F4CCCC", "failed_code": "D9EAD3"}
        for value, color in colors.items():
            sheet.conditional_formatting.add(
                f"A2:{get_column_letter(len(CASE_COLUMNS))}{end_row}",
                FormulaRule(formula=[f'${column_letter}2="{value}"'], fill=PatternFill("solid", fgColor=color)),
            )

    add_cases_sheet("All failed cases", cases)
    abbreviations = {"GraphOtter": "GO", "ST-raptor": "STR", "SpreadsheetAgent": "SSA"}
    for summary in summaries:
        group = [
            row for row in cases
            if row["benchmark"] == summary["benchmark"] and row["solution"] == summary["solution"]
        ]
        add_cases_sheet(f"{abbreviations[summary['solution']]} {summary['benchmark']}", group)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def write_markdown(path: Path, summaries: list[dict], cases: list[dict]) -> None:
    lines = [
        "# Failed-case analysis: Hitab and mulhi",
        "",
        "Only failed outcomes are included. A `wrong case` is a completed execution whose answer remains wrong after the LLM judge. A `fail case` is either a wrong case or a source execution error.",
        "",
        "| Benchmark | Solution | Wrong | Misalignment | Misinterpretation | All fail cases | Failed code |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for item in summaries:
        lines.append(
            f"| {item['benchmark']} | {item['solution']} | {item['total_wrong_cases']} | "
            f"{item['misalignment_cases']} / {item['total_wrong_cases']} ({pct(item['misalignment_ratio'])}) | "
            f"{item['misinterpretation_cases']} / {item['total_wrong_cases']} ({pct(item['misinterpretation_ratio'])}) | "
            f"{item['total_fail_cases']} | {item['failed_code_cases']} / {item['total_fail_cases']} ({pct(item['failed_code_ratio'])}) |"
        )
    lines.extend([
        "",
        "## Audit files",
        "",
        "- `failure_modes.xlsx`: summary chart, filters, conditional formatting, and one detail sheet per benchmark/solution.",
        "- `all_failed_cases.csv`: every failed case with X/Z previews, evidence, W error, Y/Y*, and artifact paths.",
        "- The Artifact Viewer website provides the filterable visual report and case inspector.",
        "- `cases/*.csv`: one CSV per benchmark/solution.",
        "",
        "## Classification rules",
        "",
        "- Misalignment: W completed, Y differs from Y*, and Z both faithfully represents X and covers every gold-relevant table.",
        "- Misinterpretation: W completed, Y differs from Y*, and Z is missing, structurally inconsistent with X, selects the wrong table, or omits a gold-relevant MulHi table.",
        "- Failed code: W is incomplete or failed to execute. Its denominator is all failed outcomes (`wrong + sourceErrors`).",
        "- MulHi expected table indices come from the first component of each gold `table_evidence` cell ID. For GraphOtter/ST-Raptor, the selected table set must equal that expected set; selecting only one of two required tables is a mismatch.",
        "- SpreadsheetAgent used an explicit workbook-Markdown fallback because the official extraction services were unavailable. The fallback counts as Z because the saved workflow used it; every represented sheet and non-empty X cell is checked.",
        "",
        f"Total detailed failed cases: {len(cases)}.",
        "",
    ])
    path.write_text("\n".join(lines), encoding="utf-8")


def write_html(path: Path, summaries: list[dict], cases: list[dict]) -> None:
    summary_cards = "".join(
        f"""
        <article class="summary-card">
          <div><span>{html.escape(item['benchmark'])}</span><strong>{html.escape(item['solution'])}</strong></div>
          <dl>
            <div><dt>Misalignment</dt><dd>{item['misalignment_cases']}/{item['total_wrong_cases']} <b>{pct(item['misalignment_ratio'])}</b></dd></div>
            <div><dt>Misinterpretation</dt><dd>{item['misinterpretation_cases']}/{item['total_wrong_cases']} <b>{pct(item['misinterpretation_ratio'])}</b></dd></div>
            <div><dt>Failed code</dt><dd>{item['failed_code_cases']}/{item['total_fail_cases']} <b>{pct(item['failed_code_ratio'])}</b></dd></div>
          </dl>
        </article>
        """
        for item in summaries
    )
    data = json.dumps(cases, ensure_ascii=False).replace("</", "<\\/")
    document = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Failed-case analysis</title>
<style>
:root{{--ink:#132536;--muted:#5c6c78;--paper:#f5f1e8;--panel:#fffdf8;--line:#d8d0c2;--navy:#17324d;--gold:#d69b2d;--red:#b5483c;--green:#3f7556}}
*{{box-sizing:border-box}} body{{margin:0;background:radial-gradient(circle at 12% 0,#f9dba7 0,transparent 30rem),linear-gradient(145deg,#f8f3e9,#e9eef0);color:var(--ink);font:15px/1.45 Georgia,serif}}
header{{padding:48px clamp(20px,5vw,76px) 28px;border-bottom:1px solid var(--line)}} h1{{margin:0;font-size:clamp(34px,6vw,70px);letter-spacing:-.04em;line-height:.98}} header p{{max-width:850px;color:var(--muted);font-size:17px}}
main{{padding:28px clamp(16px,4vw,64px) 80px}} .cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px}}
.summary-card{{background:rgba(255,253,248,.9);border:1px solid var(--line);box-shadow:0 12px 30px #17324d12;padding:18px}} .summary-card>div{{display:flex;justify-content:space-between;border-bottom:2px solid var(--navy);padding-bottom:8px}} .summary-card span{{text-transform:uppercase;letter-spacing:.12em;color:var(--red);font:700 12px sans-serif}} .summary-card strong{{font-size:18px}}
dl{{margin:12px 0 0}} dl div{{display:grid;grid-template-columns:1fr auto;gap:12px;padding:7px 0;border-bottom:1px dotted var(--line)}} dt{{color:var(--muted)}} dd{{margin:0}} dd b{{display:inline-block;min-width:64px;text-align:right;color:var(--navy)}}
.controls{{position:sticky;top:0;z-index:3;display:grid;grid-template-columns:repeat(4,minmax(150px,1fr)) auto;gap:10px;margin:26px 0 14px;padding:12px;background:#17324df2;box-shadow:0 10px 30px #17324d33}} input,select,button{{width:100%;border:1px solid #ffffff55;background:#fff;color:var(--ink);padding:10px 12px;font:600 13px sans-serif}} button{{background:var(--gold);cursor:pointer;border-color:var(--gold)}}
.count{{margin:0 0 10px;color:var(--muted)}} .table-wrap{{overflow:auto;border:1px solid var(--line);background:var(--panel)}} table{{width:100%;border-collapse:collapse;min-width:1180px}} th{{position:sticky;top:62px;background:var(--navy);color:#fff;text-align:left;font:700 12px sans-serif;letter-spacing:.04em;padding:10px}} td{{vertical-align:top;padding:10px;border-bottom:1px solid var(--line)}} tr[data-category=misinterpretation] td:first-child{{border-left:5px solid var(--red)}} tr[data-category=misalignment] td:first-child{{border-left:5px solid var(--gold)}} tr[data-category=failed_code] td:first-child{{border-left:5px solid var(--green)}} code{{font:12px/1.4 ui-monospace,SFMono-Regular,monospace;white-space:pre-wrap;word-break:break-word}} details{{max-width:430px}} summary{{cursor:pointer;font-weight:700}} .tag{{display:inline-block;padding:3px 7px;background:#e9e1d3;font:700 11px sans-serif;text-transform:uppercase}} .error{{max-width:360px;white-space:pre-wrap;word-break:break-word;color:#7f3028}} .paths{{max-height:140px;overflow:auto}}
@media(max-width:800px){{header{{padding-top:30px}}.controls{{position:static;grid-template-columns:1fr 1fr}}.controls input{{grid-column:1/-1}}th{{top:0}}}}
</style>
</head>
<body>
<header><h1>Failure anatomy</h1><p>Hitab and mulhi · GraphOtter, ST-raptor, SpreadsheetAgent. Only failed outcomes are shown. Open any evidence cell to compare expected X, selected/created Z, W status, Y, and Y*.</p></header>
<main>
<section class="cards">{summary_cards}</section>
<section class="controls">
  <input id="search" placeholder="Search question, sample ID, answer…">
  <select id="benchmark"><option value="">All benchmarks</option><option>Hitab</option><option>mulhi</option></select>
  <select id="solution"><option value="">All solutions</option><option>GraphOtter</option><option>ST-raptor</option><option>SpreadsheetAgent</option></select>
  <select id="category"><option value="">All failure modes</option><option>misalignment</option><option>misinterpretation</option><option>failed_code</option></select>
  <button id="download">Download filtered CSV</button>
</section>
<p class="count" id="count"></p>
<div class="table-wrap"><table><thead><tr><th>Case</th><th>Question</th><th>Y vs Y*</th><th>X / Z verification</th><th>W / evidence paths</th></tr></thead><tbody id="rows"></tbody></table></div>
</main>
<script id="case-data" type="application/json">{data}</script>
<script>
const allCases=JSON.parse(document.getElementById('case-data').textContent); let visible=[];
const esc=value=>String(value??'').replace(/[&<>"']/g,ch=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[ch]));
const controls=['search','benchmark','solution','category'].map(id=>document.getElementById(id));
function filter(){{const q=controls[0].value.toLowerCase();visible=allCases.filter(item=>(!controls[1].value||item.benchmark===controls[1].value)&&(!controls[2].value||item.solution===controls[2].value)&&(!controls[3].value||item.classification===controls[3].value)&&(!q||JSON.stringify(item).toLowerCase().includes(q)));render()}}
function render(){{document.getElementById('count').textContent=`${{visible.length}} of ${{allCases.length}} failed cases`;document.getElementById('rows').innerHTML=visible.map(item=>`<tr data-category="${{esc(item.classification)}}"><td><span class="tag">${{esc(item.classification)}}</span><br><b>${{esc(item.benchmark)}} · ${{esc(item.solution)}}</b><br><code>${{esc(item.sample_id)}}</code></td><td>${{esc(item.question)}}</td><td><b>Y</b><br><code>${{esc(item.predicted_answer_y)}}</code><br><b>Y*</b><br><code>${{esc(item.gold_answer_y_star)}}</code><details><summary>Judge reason</summary>${{esc(item.llm_judge_reason)||'Source execution error'}}</details></td><td><b>Expected X</b> <code>${{esc(item.expected_x_tables)}}</code><br><b>Selected Z</b> <code>${{esc(item.selected_z_tables)}}</code><p>${{esc(item.x_z_evidence)}}</p><details><summary>X preview</summary><code>${{esc(item.x_preview)}}</code></details><details><summary>Z preview</summary><code>${{esc(item.z_preview)}}</code></details></td><td><b>${{esc(item.w_status)}}</b><div class="error">${{esc(item.w_error)}}</div><details><summary>Artifact paths</summary><code class="paths">X\n${{esc(item.x_paths)}}\n\nZ\n${{esc(item.z_paths)}}\n\nW\n${{esc(item.w_paths)}}\n\nY\n${{esc(item.y_paths)}}</code></details></td></tr>`).join('')}}
controls.forEach(control=>control.addEventListener('input',filter));
document.getElementById('download').addEventListener('click',()=>{{const columns={json.dumps(list(CASE_COLUMNS))};const quote=value=>'"'+String(value??'').replaceAll('"','""')+'"';const csv='\ufeff'+columns.map(quote).join(',')+'\r\n'+visible.map(row=>columns.map(key=>quote(row[key])).join(',')).join('\r\n');const link=document.createElement('a');link.href=URL.createObjectURL(new Blob([csv],{{type:'text/csv;charset=utf-8'}}));link.download='filtered_failed_cases.csv';link.click();URL.revokeObjectURL(link.href)}});
filter();
</script>
</body></html>"""
    path.write_text(document, encoding="utf-8")


def main() -> int:
    args = parse_args()
    output_dir = args.output_dir.resolve()
    cases, summaries = build_cases()
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_columns = list(summaries[0])
    write_csv(output_dir / "summary.csv", summaries, summary_columns)
    write_csv(output_dir / "all_failed_cases.csv", cases, CASE_COLUMNS)
    case_dir = output_dir / "cases"
    for summary in summaries:
        group = [
            row for row in cases
            if row["benchmark"] == summary["benchmark"] and row["solution"] == summary["solution"]
        ]
        filename = f"{summary['benchmark']}_{summary['solution']}.csv".replace("-", "_")
        write_csv(case_dir / filename, group, CASE_COLUMNS)
    write_xlsx(output_dir / "failure_modes.xlsx", summaries, cases)
    write_markdown(output_dir / "REPORT.md", summaries, cases)
    print(json.dumps({"output_dir": str(output_dir), "cases": len(cases), "summary": summaries}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
