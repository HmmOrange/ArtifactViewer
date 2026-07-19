import os
import json
import time
import shutil
import pickle
import types
import openpyxl
import pandas as pd
import re
import hashlib
from typing import Any, Dict, List
from datetime import datetime
import gradio as gr
from loguru import logger
from utils.constants import DELIMITER, LOG_DIR
from embedding import EmbeddingModel
from utils.api_utils import vlm_generate, llm_generate, embedding_generate
from query.primitive_pipeline import *
from table2tree.feature_tree import *
from table2tree.extract_excel import process_sheet_vlm, preprocess_sheet
from config import api_config
from utils.sheet_utils import html2workbook, extract_markdown_tables

# 全局思维链条数据存储
thinking_chain_data = {}

def _normalize_artifact_prefix(name: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9._-]+", "_", str(name or "").strip())
    text = re.sub(r"_+", "_", text).strip("._")
    return text or f"table_{int(time.time())}"


def save_named_runtime_artifacts(f_tree, cache_dir, artifact_prefix):
    """Save per-file runtime artifacts: pkl/txt/embedding."""
    prefix = _normalize_artifact_prefix(artifact_prefix)
    tree_str = f_tree.__str__([1])
    with open(os.path.join(cache_dir, f"{prefix}.pkl"), "wb") as f:
        pickle.dump(f_tree, f)
    with open(os.path.join(cache_dir, f"{prefix}.txt"), "w", encoding="utf-8") as f:
        f.write(tree_str)
    try:
        raw_values = f_tree.all_value_list()
        texts = [str(x) for x in raw_values] if raw_values else []
        if texts:
            embedding_dict = EmbeddingModel().get_embedding_dict(texts)
            EmbeddingModel().save_embedding_dict(
                embedding_dict, os.path.join(cache_dir, f"{prefix}.embedding.json")
            )
    except Exception as ee:
        logger.error(f"embedding generate failed for {prefix}: {ee}")
    logger.info(
        f"[artifact_save] prefix={prefix}, dir={cache_dir}, "
        f"files={[f'{prefix}.pkl', f'{prefix}.txt', f'{prefix}.embedding.json']}"
    )
    return prefix


def save_tree_artifacts(f_tree, cache_dir, source_prefix=None):
    """Save view artifacts and per-file runtime artifacts."""
    tree_column_json = f_tree.__json_column__()

    with open(os.path.join(cache_dir, "temp.column.json"), "w", encoding='utf-8') as f:
        json.dump(tree_column_json, f, indent=4, ensure_ascii=False)

    # Per-file artifacts for distinguishing multiple sources.
    if source_prefix:
        save_named_runtime_artifacts(f_tree, cache_dir, source_prefix)


def _safe_dom_token(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "root"
    text = re.sub(r"[^a-zA-Z0-9_-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "root"


def _build_compact_semantic_id(
    legacy_semantic_id: str,
    aliases: List[str],
    target_kind: str = "node",
) -> str:
    alias_list = [str(x or "").strip() for x in (aliases or []) if str(x or "").strip()]
    first_alias = alias_list[0] if alias_list else ""
    tokens: List[str] = []
    if first_alias:
        alias_text = first_alias
        if alias_text.startswith("ft:"):
            alias_text = alias_text[3:]
        for seg in alias_text.split("/"):
            seg = str(seg or "").strip()
            if not seg:
                continue
            m = re.match(r"^m_\d+_(.+)$", seg)
            if m:
                label = _safe_dom_token(m.group(1))
                if label and label not in {"ho_tree", "root"}:
                    tokens.append(label)
    if not tokens:
        tokens = ["root"]
    base = "_".join(tokens[:6])
    digest = hashlib.md5(str(legacy_semantic_id or "").encode("utf-8")).hexdigest()[:8]
    # Keep readable path-like prefix while ensuring uniqueness by hash suffix.
    return f"ct_tree_root_{base}_{digest}"


def build_and_save_tree_id_mappings(cache_dir: str, typed_root_name: str = "HO_TREE") -> Dict[str, Any]:
    """
    Build stable id mappings from temp1.json/temp.column.json and save to:
    - temp.id_mappings.json
    """
    try:
        row_path = os.path.join(cache_dir, "temp1.json")
        column_path = os.path.join(cache_dir, "temp.column.json")
        if not os.path.exists(row_path) or not os.path.exists(column_path):
            return {}

        with open(row_path, "r", encoding="utf-8") as rf:
            raw_row = json.load(rf)
        with open(column_path, "r", encoding="utf-8") as cf:
            raw_column = json.load(cf)

        if not isinstance(raw_row, dict) or not isinstance(raw_column, dict):
            return {}

        from utils.tree_semantic_utils import (
            build_flat_column_alias_target_map,
            build_flat_row_alias_target_map,
            build_nested_index_projection_map,
            build_semantic_projection_bundle,
            make_canonical_trace_id,
            make_tree_canonical_id,
            make_tree_group_canonical_id,
        )

        row_alias_map = build_flat_row_alias_target_map(raw_row, typed_root_name=typed_root_name)
        column_alias_map = build_flat_column_alias_target_map(raw_column, typed_root_name=typed_root_name)
        semantic_bundle = build_semantic_projection_bundle(raw_row, raw_column, typed_root_name=typed_root_name)
        nested_projection = build_nested_index_projection_map(
            raw_column,
            semantic_bundle=semantic_bundle,
            typed_root_name=typed_root_name,
        )

        row_canonical_ids = sorted(
            {
                str((item or {}).get("canonical_id", "") or "").strip()
                for item in (row_alias_map or {}).values()
                if str((item or {}).get("canonical_id", "") or "").strip().startswith("ct_tree_")
            }
        )
        column_canonical_ids = sorted(
            {
                str((item or {}).get("canonical_id", "") or "").strip()
                for item in (column_alias_map or {}).values()
                if str((item or {}).get("canonical_id", "") or "").strip().startswith("ct_tree_")
            }
        )

        row_dom_map = {cid: f"row_{_safe_dom_token(cid)}" for cid in row_canonical_ids}
        column_dom_map = {cid: f"column_{_safe_dom_token(cid)}" for cid in column_canonical_ids}

        group_to_views: Dict[str, Dict[str, List[str]]] = {}
        for alias_id, target in (row_alias_map or {}).items():
            target_info = target if isinstance(target, dict) else {}
            if str(target_info.get("target_kind", "") or "").strip() != "group":
                continue
            group_id = str(target_info.get("canonical_id", "") or "").strip()
            if not group_id.startswith("ct_tree_group_"):
                continue
            entry = group_to_views.setdefault(group_id, {"row": [], "column": [], "aliases": []})
            alias_text = str(alias_id or "").strip()
            if alias_text and alias_text not in entry["aliases"]:
                entry["aliases"].append(alias_text)
            col_target = (column_alias_map or {}).get(alias_text, {})
            if isinstance(col_target, dict):
                col_cid = str(col_target.get("canonical_id", "") or "").strip()
                if col_cid.startswith("ct_tree_") and col_cid not in entry["column"]:
                    entry["column"].append(col_cid)

        def walk_row_groups(value: Any, ct_parts: List[str]) -> None:
            if isinstance(value, dict):
                for idx, (k, v) in enumerate(value.items()):
                    k_str = str(k)
                    cur_ct_parts = ct_parts + [f"k_{k_str}", f"idx_{idx}"]
                    if isinstance(v, dict):
                        walk_row_groups(v, cur_ct_parts + ["body"])
                    elif isinstance(v, list):
                        columns: List[Any] = []
                        for row in v:
                            if isinstance(row, dict):
                                for ck in row.keys():
                                    if ck not in columns:
                                        columns.append(ck)
                        for row_idx, row in enumerate(v):
                            if not isinstance(row, dict):
                                continue
                            row_parts = cur_ct_parts + ["body", f"i_{row_idx}", "group"]
                            for ck, _cv in row.items():
                                col_idx = columns.index(ck)
                                ck_str = str(ck)
                                group_id = make_tree_group_canonical_id(
                                    cur_ct_parts + ["body", "header_group", f"k_{ck_str}", f"idx_{col_idx}"]
                                )
                                cell_value_id = make_tree_canonical_id(
                                    row_parts + [f"k_{ck_str}", f"idx_{col_idx}", "body", "v"]
                                )
                                entry = group_to_views.setdefault(group_id, {"row": [], "column": [], "aliases": []})
                                if group_id not in entry["row"]:
                                    entry["row"].append(group_id)
                                if cell_value_id not in entry["row"]:
                                    entry["row"].append(cell_value_id)

        walk_row_groups(raw_row, ["root", "flat_row", "root"])

        for item in group_to_views.values():
            item["row"] = sorted({str(x or "").strip() for x in item.get("row", []) if str(x or "").strip()})
            item["column"] = sorted({str(x or "").strip() for x in item.get("column", []) if str(x or "").strip()})
            item["aliases"] = sorted({str(x or "").strip() for x in item.get("aliases", []) if str(x or "").strip()})

        semantic_to_views = (semantic_bundle or {}).get("semantic_to_views", {}) or {}
        semantic_rows = (semantic_bundle or {}).get("row_canonical_to_semantic", {}) or {}
        semantic_columns = (semantic_bundle or {}).get("column_canonical_to_semantic", {}) or {}

        row_to_group_ids: Dict[str, List[str]] = {}
        column_to_group_ids: Dict[str, List[str]] = {}
        for group_id, entry in (group_to_views or {}).items():
            if not isinstance(entry, dict):
                continue
            for row_id in (entry.get("row", []) or []):
                row_text = str(row_id or "").strip()
                if not row_text:
                    continue
                row_to_group_ids.setdefault(row_text, [])
                if group_id not in row_to_group_ids[row_text]:
                    row_to_group_ids[row_text].append(group_id)
            for col_id in (entry.get("column", []) or []):
                col_text = str(col_id or "").strip()
                if not col_text:
                    continue
                column_to_group_ids.setdefault(col_text, [])
                if group_id not in column_to_group_ids[col_text]:
                    column_to_group_ids[col_text].append(group_id)

        semantic_projection: Dict[str, Dict[str, Any]] = {}
        semantic_compact_projection: Dict[str, Dict[str, Any]] = {}
        semantic_legacy_to_compact: Dict[str, str] = {}
        semantic_compact_to_legacy: Dict[str, str] = {}
        semantic_legacy_to_group_ids: Dict[str, List[str]] = {}
        semantic_compact_to_group_ids: Dict[str, List[str]] = {}
        alias_to_semantic_compact: Dict[str, str] = {}
        row_canonical_to_semantic_compact: Dict[str, str] = {}
        column_canonical_to_semantic_compact: Dict[str, str] = {}
        for semantic_id, payload in semantic_to_views.items():
            semantic_text = str(semantic_id or "").strip()
            if not semantic_text:
                continue
            semantic_tree_id = semantic_text.replace("ct_semantic_", "ct_tree_semantic_", 1)
            row_ids = sorted(
                {
                    str(v or "").strip()
                    for v in (payload.get("row", []) if isinstance(payload, dict) else [])
                    if str(v or "").strip()
                }
            )
            column_ids = sorted(
                {
                    str(v or "").strip()
                    for v in (payload.get("column", []) if isinstance(payload, dict) else [])
                    if str(v or "").strip()
                }
            )
            alias_ids = sorted(
                {
                    str(v or "").strip()
                    for v in (payload.get("aliases", []) if isinstance(payload, dict) else [])
                    if str(v or "").strip()
                }
            )
            target_kind = str((payload or {}).get("target_kind", "") or "node")

            group_ids = set()
            for rid in row_ids:
                if rid.startswith("ct_tree_group_"):
                    group_ids.add(rid)
                for gid in row_to_group_ids.get(rid, []) or []:
                    group_ids.add(str(gid or "").strip())
            for cid in column_ids:
                for gid in column_to_group_ids.get(cid, []) or []:
                    group_ids.add(str(gid or "").strip())
            group_ids_sorted = sorted({str(x or "").strip() for x in group_ids if str(x or "").strip()})

            compact_semantic_id = _build_compact_semantic_id(
                legacy_semantic_id=semantic_text,
                aliases=alias_ids,
                target_kind=target_kind,
            )

            semantic_projection[semantic_tree_id] = {
                "semantic_id": semantic_text,
                "target_kind": target_kind,
                "row": row_ids,
                "column": column_ids,
                "group_ids": group_ids_sorted,
                "aliases": alias_ids,
                "semantic_compact_id": compact_semantic_id,
            }
            semantic_compact_projection[compact_semantic_id] = {
                "semantic_id": semantic_text,
                "target_kind": target_kind,
                "row": row_ids,
                "column": column_ids,
                "group_ids": group_ids_sorted,
                "aliases": alias_ids,
            }
            semantic_legacy_to_compact[semantic_text] = compact_semantic_id
            semantic_compact_to_legacy[compact_semantic_id] = semantic_text
            semantic_legacy_to_group_ids[semantic_text] = group_ids_sorted
            semantic_compact_to_group_ids[compact_semantic_id] = group_ids_sorted
            for alias_id in alias_ids:
                alias_to_semantic_compact[alias_id] = compact_semantic_id
            for rid in row_ids:
                row_canonical_to_semantic_compact[rid] = compact_semantic_id
            for cid in column_ids:
                column_canonical_to_semantic_compact[cid] = compact_semantic_id

        group_to_semantic_nodes: Dict[str, List[Dict[str, Any]]] = {}
        semantic_to_group_ids_compact: Dict[str, List[str]] = {}
        semantic_to_group_ids_legacy: Dict[str, List[str]] = {}
        row_to_column_ids: Dict[str, List[str]] = {}
        for compact_id, info in semantic_compact_projection.items():
            if not isinstance(info, dict):
                continue
            legacy_id = str(info.get("semantic_id", "") or "").strip()
            row_ids = [str(x or "").strip() for x in (info.get("row", []) or []) if str(x or "").strip()]
            column_ids = [str(x or "").strip() for x in (info.get("column", []) or []) if str(x or "").strip()]
            group_ids = [str(x or "").strip() for x in (info.get("group_ids", []) or []) if str(x or "").strip()]

            semantic_to_group_ids_compact[compact_id] = sorted({g for g in group_ids})
            if legacy_id:
                semantic_to_group_ids_legacy[legacy_id] = sorted({g for g in group_ids})

            for rid in row_ids:
                row_to_column_ids.setdefault(rid, [])
                for cid in column_ids:
                    if cid not in row_to_column_ids[rid]:
                        row_to_column_ids[rid].append(cid)

            for gid in group_ids:
                group_to_semantic_nodes.setdefault(gid, [])
                group_to_semantic_nodes[gid].append({
                    "semantic_compact_id": compact_id,
                    "semantic_id": legacy_id,
                    "target_kind": str(info.get("target_kind", "") or "node"),
                    "row": row_ids,
                    "column": column_ids,
                })

        for rid, values in list(row_to_column_ids.items()):
            row_to_column_ids[rid] = sorted({str(x or "").strip() for x in values if str(x or "").strip()})
        for gid, items in list(group_to_semantic_nodes.items()):
            uniq: Dict[str, Dict[str, Any]] = {}
            for item in items:
                key = str(item.get("semantic_compact_id", "") or "").strip()
                if key and key not in uniq:
                    uniq[key] = item
            group_to_semantic_nodes[gid] = [uniq[k] for k in sorted(uniq.keys())]

        nested_id_projection: Dict[str, Dict[str, Any]] = {}
        for semantic_id, item in (nested_projection or {}).items():
            entry = item if isinstance(item, dict) else {}
            path = [str(p or "").strip() for p in (entry.get("path", []) or []) if str(p or "").strip()]
            nested_id = make_canonical_trace_id(["tree_nested"] + path)
            nested_id_projection[nested_id] = {
                "semantic_id": str(semantic_id or "").strip(),
                "path": path,
                "indexName": str(entry.get("indexName", "") or ""),
                "drillable": bool(entry.get("drillable", False)),
                "childFeaturePath": [
                    str(p or "").strip()
                    for p in (entry.get("childFeaturePath", []) or [])
                    if str(p or "").strip()
                ],
            }

        payload = {
            "version": "v1",
            "typed_root_name": typed_root_name,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "row": {
                "canonical_to_dom_id": row_dom_map,
                "alias_to_target": row_alias_map,
            },
            "column": {
                "canonical_to_dom_id": column_dom_map,
                "alias_to_target": column_alias_map,
            },
            "nested": {
                "ct_tree_nested_to_projection": nested_id_projection,
            },
            "group": {
                "ct_tree_group_to_row_column": group_to_views,
            },
            "semantic": {
                "semantic_tree_id_to_views": semantic_projection,
                "semantic_compact_id_to_views": semantic_compact_projection,
                "semantic_id_to_group_ids": semantic_legacy_to_group_ids,
                "semantic_compact_id_to_group_ids": semantic_compact_to_group_ids,
                "semantic_legacy_to_compact": semantic_legacy_to_compact,
                "semantic_compact_to_legacy": semantic_compact_to_legacy,
                "alias_to_semantic_compact": alias_to_semantic_compact,
                "row_canonical_to_semantic": semantic_rows,
                "column_canonical_to_semantic": semantic_columns,
                "row_canonical_to_semantic_compact": row_canonical_to_semantic_compact,
                "column_canonical_to_semantic_compact": column_canonical_to_semantic_compact,
                "semantic_to_group_ids_compact": semantic_to_group_ids_compact,
                "semantic_to_group_ids_legacy": semantic_to_group_ids_legacy,
                "group_to_semantic_nodes": group_to_semantic_nodes,
                "row_to_column_ids": row_to_column_ids,
            },
        }

        output_path = os.path.join(cache_dir, "temp.id_mappings.json")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        try:
            mapping_log_path = os.path.join(cache_dir, "temp.semantic.mapping.log.json")
            mapping_log_payload = {
                "generated_at": payload.get("generated_at"),
                "semantic_to_group_ids_compact": payload.get("semantic", {}).get("semantic_to_group_ids_compact", {}),
                "semantic_to_group_ids_legacy": payload.get("semantic", {}).get("semantic_to_group_ids_legacy", {}),
                "group_to_semantic_nodes": payload.get("semantic", {}).get("group_to_semantic_nodes", {}),
                "row_to_column_ids": payload.get("semantic", {}).get("row_to_column_ids", {}),
            }
            with open(mapping_log_path, "w", encoding="utf-8") as f:
                json.dump(mapping_log_payload, f, ensure_ascii=False, indent=2)
            logger.info(f"[id_mapping] semantic log saved path={mapping_log_path}")
        except Exception as ee:
            logger.warning(f"[id_mapping] semantic log save failed: {ee}")

        logger.info(
            f"[id_mapping] saved path={output_path}, "
            f"row={len(row_dom_map)}, column={len(column_dom_map)}, "
            f"group={len(group_to_views)}, semantic={len(semantic_projection)}, nested={len(nested_id_projection)}"
        )
        return payload
    except Exception as e:
        logger.warning(f"[id_mapping] build failed: {e}")
        return {}


def _save_artifact_manifest(cache_dir, mapping):
    path = os.path.join(cache_dir, "temp.artifacts.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(mapping or {}, f, ensure_ascii=False, indent=2)
    try:
        logger.info(
            f"[artifact_manifest] saved path={path}, keys={list((mapping or {}).keys())}, count={len(mapping or {})}"
        )
    except Exception:
        pass


def _load_artifact_manifest(cache_dir):
    path = os.path.join(cache_dir, "temp.artifacts.json")
    if not os.path.exists(path):
        logger.info(f"[artifact_manifest] not found: {path}")
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        loaded = data if isinstance(data, dict) else {}
        logger.info(f"[artifact_manifest] loaded path={path}, keys={list(loaded.keys())}, count={len(loaded)}")
        return loaded
    except Exception as e:
        logger.warning(f"[artifact_manifest] load failed path={path}: {e}")
        return {}


def _choose_artifact_keys_for_query(query, manifest):
    keys = [str(k) for k in (manifest or {}).keys()]
    if not keys:
        return []
    q = str(query or "").strip().lower()

    def _query_tokens(text):
        text = re.sub(r"\s+", " ", str(text or "").strip().lower())
        if not text:
            return []
        if " " in text:
            return [t for t in text.split(" ") if len(t) >= 2]
        # 中文/无空格场景：退化为整体 token
        return [text]

    q_tokens = _query_tokens(q)
    key_scores = {}
    for k in keys:
        item = manifest.get(k, {}) if isinstance(manifest, dict) else {}
        meta = item.get("meta", {}) if isinstance(item, dict) else {}
        if not isinstance(meta, dict):
            meta = {}
        route_aliases = meta.get("route_aliases", []) if isinstance(meta.get("route_aliases", []), list) else []
        header_preview = meta.get("header_preview", []) if isinstance(meta.get("header_preview", []), list) else []
        index_preview = meta.get("index_preview", []) if isinstance(meta.get("index_preview", []), list) else []
        display_name = str(meta.get("display_name", "") or "")
        summary = str(meta.get("summary", "") or "")
        candidate_texts = [k, display_name, summary] + [str(x) for x in route_aliases[:20]] + [str(x) for x in header_preview[:20]] + [str(x) for x in index_preview[:20]]

        score = 0
        for token in q_tokens:
            for c in candidate_texts:
                c_l = c.lower()
                if not c_l:
                    continue
                if token in c_l:
                    score += 3
                elif len(token) >= 3 and c_l in token:
                    score += 1
        # 兼容短问句：直接子串判断
        q_compact = q.replace(" ", "")
        for c in candidate_texts:
            c_compact = str(c).lower().replace(" ", "")
            if not c_compact:
                continue
            if q_compact and q_compact in c_compact:
                score += 4
            elif len(c_compact) >= 4 and c_compact in q_compact:
                score += 2
        key_scores[k] = score

    base_rank = sorted(keys, key=lambda kk: key_scores.get(kk, 0), reverse=True)
    logger.info(f"[artifact_route] lexical score={key_scores}, base_rank={base_rank}, query={query}")

    # LLM hint for top-1 routing; fallback to score ranking.
    llm_pick = ""
    try:
        candidate_lines = []
        for idx, k in enumerate(base_rank, start=1):
            item = manifest.get(k, {}) if isinstance(manifest, dict) else {}
            meta = item.get("meta", {}) if isinstance(item, dict) else {}
            if not isinstance(meta, dict):
                meta = {}
            summary = str(meta.get("summary", "")).strip()
            display_name = str(meta.get("display_name", "")).strip()
            header_preview = meta.get("header_preview", []) if isinstance(meta.get("header_preview", []), list) else []
            header_preview_str = ", ".join([str(x) for x in header_preview[:10]])
            value_count = meta.get("value_count", 0)
            index_count = meta.get("index_count", 0)
            line = (
                f"{idx}. key={k} | title={display_name or 'N/A'} | summary={summary or 'N/A'} | "
                f"header_preview=[{header_preview_str}] | index_count={index_count} | value_count={value_count}"
            )
            candidate_lines.append(line)

        candidates_block = "\n".join(candidate_lines)
        prompt = (
            "You are selecting the most relevant table artifact for QA routing.\n"
            "Do NOT rely on random key names; prioritize semantic meaning in title/summary/header_preview.\n"
            "Return ONLY one key exactly as listed; no explanation.\n"
            f"Question: {query}\n"
            f"Candidate Keys: {base_rank}\n"
            f"Candidates Detail:\n{candidates_block}\n"
        )
        logger.info(f"[artifact_route] llm prompt={prompt}")
        llm_pick = str(get_llm_generate(prompt, max_tokens=40, temperature=0.1) or "").strip()
        logger.info(f"[artifact_route] llm pick={llm_pick}")
    except Exception as e:
        logger.warning(f"[artifact_route] llm route hint failed: {e}")

    final_rank = list(base_rank)
    if llm_pick in final_rank:
        final_rank.remove(llm_pick)
        final_rank.insert(0, llm_pick)
    logger.info(f"[artifact_route] final rank={final_rank}")
    return final_rank


def _build_artifact_meta(top_key, single_tree, top_body=None):
    file_key = str(top_key)
    summary = {
        "file_key": file_key,
        "display_name": "",
        "summary": "",
        "route_aliases": [],
        "sheet_preview": [],
        "section_preview": [],
        "header_preview": [],
        "index_preview": [],
        "index_count": 0,
        "value_count": 0,
        "value_preview": [],
        "body_type": type(top_body).__name__,
    }
    random_key_like = bool(re.match(r"^tmp[a-z0-9]+$", file_key, flags=re.IGNORECASE))
    try:
        root_children = getattr(single_tree.index_tree.root, "children", []) or []
        index_names = [str(getattr(n, "value", "")) for n in root_children]
        summary["index_preview"] = index_names[:12]
        summary["index_count"] = len(index_names)
    except Exception:
        pass
    try:
        values = single_tree.all_value_list() or []
        summary["value_count"] = len(values)
        summary["value_preview"] = [str(v) for v in values[:12]]
    except Exception:
        pass

    # 若调用方未传 top_body，则从单树列视图反推当前文件 body。
    if top_body is None:
        try:
            tree_column = single_tree.__json_column__()
            if isinstance(tree_column, dict) and tree_column:
                first_k = next(iter(tree_column.keys()))
                guessed = tree_column.get(first_k)
                if isinstance(guessed, (dict, list, str, int, float, bool)) or guessed is None:
                    top_body = guessed
                    summary["body_type"] = type(top_body).__name__
        except Exception:
            pass

    # 从 canonical body 中提取更语义化的字段信息，避免只依赖 tmpxxxx key。
    header_set = []
    alias_set = []
    sheet_set = []
    section_set = []
    try:
        def _append_unique(bucket, text, limit=24):
            t = str(text or "").strip()
            if not t or t in bucket:
                return
            bucket.append(t)
            if len(bucket) > limit:
                del bucket[limit:]

        # 提取 "文件 -> sheet -> 直接标题(section)" 这两层信息
        if isinstance(top_body, dict):
            for sheet_name, sheet_body in top_body.items():
                _append_unique(sheet_set, sheet_name, limit=16)
                if isinstance(sheet_body, dict):
                    for section_name in sheet_body.keys():
                        _append_unique(section_set, section_name, limit=24)

        def _walk(node, depth=0):
            if depth > 3:
                return
            if isinstance(node, dict):
                for k, v in node.items():
                    k_text = str(k).strip()
                    _append_unique(header_set, k_text, limit=24)
                    _append_unique(alias_set, k_text, limit=24)
                    _walk(v, depth + 1)
            elif isinstance(node, list):
                for item in node[:30]:
                    _walk(item, depth + 1)
            else:
                _append_unique(alias_set, str(node), limit=24)

        _walk(top_body, depth=0)
    except Exception:
        pass

    summary["sheet_preview"] = sheet_set[:12]
    summary["section_preview"] = section_set[:12]
    summary["header_preview"] = header_set[:12]
    display_name = ""
    if summary["section_preview"]:
        display_name = " / ".join(summary["section_preview"][:2])
    elif summary["header_preview"]:
        display_name = " / ".join(summary["header_preview"][:2])
    elif summary["sheet_preview"]:
        display_name = " / ".join(summary["sheet_preview"][:2])
    elif summary["index_preview"]:
        display_name = " / ".join(summary["index_preview"][:2])
    if not display_name:
        display_name = file_key
    if random_key_like and summary["header_preview"]:
        # 对 random key 场景，尽量显示真实表头
        display_name = " / ".join(summary["header_preview"][:2])
    if random_key_like and summary["section_preview"]:
        display_name = " / ".join(summary["section_preview"][:2])
    summary["display_name"] = display_name

    route_aliases = []
    for x in [display_name] + summary["sheet_preview"][:12] + summary["section_preview"][:12] + summary["header_preview"][:12] + summary["index_preview"][:12]:
        t = str(x or "").strip()
        if t and t not in route_aliases:
            route_aliases.append(t)
    summary["route_aliases"] = route_aliases[:20]

    sheet_preview_text = ", ".join(summary["sheet_preview"][:4]) if summary["sheet_preview"] else ""
    section_preview_text = ", ".join(summary["section_preview"][:6]) if summary["section_preview"] else ""
    header_preview_text = ", ".join(summary["header_preview"][:6]) if summary["header_preview"] else ""
    index_preview_text = ", ".join(summary["index_preview"][:6]) if summary["index_preview"] else ""
    value_preview_text = ", ".join(summary["value_preview"][:4]) if summary["value_preview"] else ""
    summary["summary"] = (
        f"title={summary['display_name']}; "
        f"sheets={sheet_preview_text or 'N/A'}; "
        f"sections={section_preview_text or 'N/A'}; "
        f"headers={header_preview_text or 'N/A'}; "
        f"indexes={index_preview_text or 'N/A'}; "
        f"value_count={summary['value_count']}; "
        f"value_preview={value_preview_text or 'N/A'}"
    )
    return summary

def ensure_cache_directories(cache_dir, temp_dir=None):
    """Ensure cache directories exist"""
    os.makedirs(cache_dir, exist_ok=True)
    if temp_dir:
        os.makedirs(temp_dir, exist_ok=True)

def handle_processing_error(e, error_prefix="处理"):
    """Standard error handling for processing functions"""
    import traceback
    error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
    gr.Warning(f"❌ {error_prefix}失败: {error_msg}")
    return f"{error_prefix}失败"

def setup_cache_directory(conversation_id, default_cache_dir="cache", default_temp_dir="data/SSTQA/temp_tables"):
    """Setup cache directory based on conversation_id"""
    if conversation_id:
        cache_dir = os.path.join("history", conversation_id)
        temp_dir = os.path.join("history", conversation_id)
    else:
        cache_dir = default_cache_dir
        temp_dir = default_temp_dir
    
    os.makedirs(cache_dir, exist_ok=True)
    os.makedirs(temp_dir, exist_ok=True)
    
    return cache_dir, temp_dir


def save_placeholder_data(cache_dir, placeholder_data, embedding_texts=None):
    """Save placeholder data and embeddings for conversation history"""
    try:
        # Save placeholder pkl file
        with open(os.path.join(cache_dir, "temp.pkl"), "wb") as f:
            pickle.dump(placeholder_data, f)
        
        # Generate and save embeddings if provided
        if embedding_texts:
            embedding_dict = EmbeddingModel().get_embedding_dict(embedding_texts)
            EmbeddingModel().save_embedding_dict(
                embedding_dict, os.path.join(cache_dir, "temp.embedding.json")
            )
    except Exception as e:
        logger.error(f"Failed to save placeholder data: {e}")


def ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts=None):
    """Ensure conversation cache directory exists and save placeholder data"""
    if conversation_id:
        cache_dir = os.path.join("history", conversation_id)
        os.makedirs(cache_dir, exist_ok=True)
        save_placeholder_data(cache_dir, placeholder_data, embedding_texts)
        return cache_dir
    return None


def clear_directory_contents(dir_path):
    """Clear all contents from a directory, preserving the directory itself"""
    if os.path.exists(dir_path):
        for item in os.listdir(dir_path):
            item_path = os.path.join(dir_path, item)
            if os.path.isfile(item_path):
                os.remove(item_path)
            elif os.path.isdir(item_path):
                shutil.rmtree(item_path)  # 递归删除子目录

def generate_conversation_id():
    """生成唯一的对话ID，基于时间戳"""
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 包含毫秒以确保唯一性

def _contains_cjk(text):
    return bool(re.search(r"[\u4e00-\u9fff]", text or ""))


def _truncate_title(title, max_words=8, max_chars=40):
    title = (title or "").strip()
    if not title:
        return ""
    words = title.split()
    if len(words) > max_words:
        title = " ".join(words[:max_words])
    if len(title) > max_chars:
        title = title[:max_chars].rstrip() + "..."
    return title


def generate_history_title_from_questions(chat_history):
    """Generate a concise English history title from user questions."""
    if not chat_history or not isinstance(chat_history, list):
        return None
    
    # 提取所有用户问题
    user_questions = []
    for msg in chat_history:
        if isinstance(msg, dict) and msg.get("role") == "user":
            question = msg.get("content", "").strip()
            if question:
                user_questions.append(question)
    
    if not user_questions:
        return None
    
    # 如果只有一个问题，直接使用它；如果有多个，用LLM概括
    if len(user_questions) == 1:
        question_text = user_questions[0]
    else:
        # 合并所有问题
        questions_text = "\n".join([f"{i+1}. {q}" for i, q in enumerate(user_questions)])
        question_text = questions_text
    
    # 限制问题文本长度，避免超出LLM上下文
    if len(question_text) > 500:
        question_text = question_text[:500] + "..."
    
    # 使用LLM生成标题
    prompt = f"""Generate a concise English title (no more than 8 words) that summarizes the core topic of the conversation.

User questions:
{question_text}

Return the title only. Do not add explanations or quotes."""
    
    try:
        title = get_llm_generate(prompt, max_tokens=50, temperature=0.3)
        title = title.strip().strip('"').strip("'")
        title = _truncate_title(title, max_words=8, max_chars=40)
        if _contains_cjk(title):
            return "Conversation Summary"
        return title if title else None
    except Exception as e:
        logger.error(f"生成历史记录标题失败: {e}")
        return None

def create_conversation_record(conversation_id, file_list, upload_time, summary, chat_history=None):
    """创建对话记录文件，用于历史记录显示
    
    参数:
    conversation_id: 对话ID
    file_list: 文件列表
    upload_time: 上传时间
    summary: 摘要（如果提供chat_history且有用户问题，会用LLM生成标题覆盖summary）
    chat_history: 对话历史，格式为[{"role": "user", "content": "问题"}, ...]
    """
    history_dir = "history"
    os.makedirs(history_dir, exist_ok=True)
    
    record_file = os.path.join(history_dir, "history_records.json")
    
    # 读取现有记录
    records = []
    if os.path.exists(record_file):
        try:
            with open(record_file, 'r', encoding='utf-8') as f:
                records = json.load(f)
        except:
            records = []
    
    # 如果有对话历史且有用户问题，尝试用LLM生成标题
    final_summary = summary
    if chat_history:
        llm_title = generate_history_title_from_questions(chat_history)
        if llm_title:
            final_summary = llm_title
        # 如果没有生成标题，使用默认summary
    final_summary = _truncate_title(final_summary, max_words=8, max_chars=40) or "Conversation Summary"
    
    # 添加新记录
    new_record = {
        "conversation_id": conversation_id,
        "file_list": file_list,
        "upload_time": upload_time,
        "summary": final_summary
    }
    
    records.append(new_record)
    
    # 保存记录
    with open(record_file, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def get_conversation_records():
    """获取所有对话记录，用于历史记录显示"""
    history_dir = "history"
    os.makedirs(history_dir, exist_ok=True)
    
    record_file = os.path.join(history_dir, "history_records.json")
    
    # 读取记录
    records = []
    if os.path.exists(record_file):
        try:
            with open(record_file, 'r', encoding='utf-8') as f:
                records = json.load(f)
        except:
            records = []
    
    # 转换为表格格式数据
    table_data = []
    for record in records:
        # 将文件列表转换为字符串
        file_names_str = ", ".join(record.get("file_list", []))
        table_data.append([
            record.get("conversation_id", ""),
            file_names_str,
            record.get("upload_time", ""),
            record.get("summary", ""),
            "查看"  # 操作列
        ])
    
    return table_data

def get_llm_generate(prompt, max_tokens=8192, temperature=0.5):
    return llm_generate(
        prompt=prompt,
        key=api_config["llm_api_key"],
        url=api_config["llm_api_url"],
        model=api_config["llm_model"],
        max_tokens=max_tokens,
        temperature=temperature
    )

def reshape_question_with_context(current_question, chat_history, temperature=0.5):
    """
    使用上下文重塑用户问题，明确并代替可能指代不明确的代词
    
    参数:
    current_question: 当前用户问题
    chat_history: 对话历史，格式为[{"role": "user", "content": "问题"}, {"role": "assistant", "content": "回答"}, ...]
    temperature: LLM温度参数
    
    返回:
    重塑后的清晰问题
    """
    if not chat_history or len(chat_history) == 0:
        return current_question
    
    # 构建上下文提示
    context_prompt = "对话历史:\n"
    for message in chat_history:
        role = "用户" if message["role"] == "user" else "助手"
        context_prompt += f"{role}: {message['content']}\n"
    
    context_prompt += f"\n当前问题: {current_question}\n"
    context_prompt += "\n请根据对话历史，重塑当前问题，明确并代替可能指代不明确的代词，保持问题的核心意思不变。"
    context_prompt += "\n重塑后的问题:"
    
    try:
        reshaped_question = get_llm_generate(
            prompt=context_prompt,
            max_tokens=256,
            temperature=temperature
        )
        logger.info(f"原始问题: {current_question}")
        logger.info(f"重塑问题: {reshaped_question}")
        return reshaped_question.strip()
    except Exception as e:
        logger.error(f"问题重塑失败: {str(e)}")
        # 如果重塑失败，返回原始问题
        return current_question

def get_vlm_generate():
    # 返回一个已经配置好API参数的vlm_generate函数
    def configured_vlm_generate(prompt, image, temperature=0.5):
        return vlm_generate(
            prompt=prompt,
            image=image,
            key=api_config["vlm_api_key"],
            url=api_config["vlm_api_url"],
            model=api_config["vlm_model"],
            temperature=temperature
        )
    return configured_vlm_generate

def get_embedding_generate():
    # 返回一个已经配置好API参数的embedding_generate函数
    def configured_embedding_generate(input_texts, dimensions=1024):
        return embedding_generate(
            input_texts=input_texts,
            key=api_config["embedding_api_key"],
            url=api_config["embedding_api_url"],
            model=api_config["embedding_model"],
            dimensions=dimensions
        )
    return configured_embedding_generate


def convert_to_xlsx(src_path, dest_path):
    """将各种格式的文件转换为 xlsx 格式"""
    ext = os.path.splitext(src_path)[1].lower()
    try:
        if ext == ".xlsx":
            shutil.copy2(src_path, dest_path)
        elif ext == ".csv":
            df_src = pd.read_csv(src_path)
            df_src.to_excel(dest_path, index=False, engine="openpyxl")
        elif ext == ".html":
            html_content = open(src_path, "r", encoding="utf-8").read()
            html2workbook(html_content).save(dest_path)
        elif ext == ".md":
            md_content = open(src_path, "r", encoding="utf-8").read()
            table = extract_markdown_tables(md_content)
            if table and len(table) > 1:
                df_src = pd.DataFrame(table[1:], columns=table[0])
                df_src.to_excel(dest_path, index=False, engine="openpyxl")
            else:
                shutil.copy2(src_path, dest_path)
        else:
            shutil.copy2(src_path, dest_path)
    except Exception as e:
        logger.error(f"转换文件 {src_path} 到 xlsx 失败: {e}")
        shutil.copy2(src_path, dest_path)


def get_multiple_excel_feature_tree(files, log_dir=LOG_DIR, vlm_cache=False):
    """处理多个 Excel 文件，并构建成一棵总树，根节点为 'alldocument'"""
    all_docs_dict = {}
    temp_dir = "data/SSTQA/temp_tables"
    os.makedirs(temp_dir, exist_ok=True)
    
    for i, file_obj in enumerate(files):
        # file_obj 可能是 Gradio 的 File 对象或 SimpleNamespace
        src_path = file_obj.name if hasattr(file_obj, 'name') else str(file_obj)
        filename = os.path.basename(src_path)
        
        # 为每个文件创建一个唯一的临时 xlsx 名
        temp_file = os.path.join(temp_dir, f"temp_{i}.xlsx")
        try:
            convert_to_xlsx(src_path, temp_file)
            
            # 开启处理逻辑
            wb = openpyxl.load_workbook(temp_file, data_only=True)
            file_tree_dict = {}
            for sheet_name in wb.sheetnames:
                logger.info(f"正在处理文件 {filename} 的 Sheet: {sheet_name}")
                sheet = preprocess_sheet(wb[sheet_name])
                # 获取该 sheet 的结构字典 (tree_dict)
                sheet_tree_dict = process_sheet_vlm(sheet, get_json=False, cache=vlm_cache)
                file_tree_dict[sheet_name] = sheet_tree_dict
            
            # 将该文件的所有 sheet 挂在文件名节点下
            all_docs_dict[filename] = file_tree_dict
        except Exception as e:
            logger.error(f"处理文件 {filename} 失败: {e}")
            continue

    # 构建带 'alldocument' 根节点的字典
    combined_tree_dict = {"alldocument": all_docs_dict}
    
    # 建树并打标签
    total_tree = construct_feature_tree(combined_tree_dict)
    total_tree = tag_feature_tree(total_tree)
    
    return total_tree


def process_multiple_tables_for_tree(files, conversation_id=None):
    """专门处理多个表格，生成统一的 H-OTree 结构"""
    global thinking_chain_data
    thinking_chain_data = {"question_answering": {}, "retrieval_chains": []}
    
    if not files:
        return None
    
    try:
        # 如果提供了 conversation_id，则创建专用文件夹，否则使用 cache 目录
        if conversation_id:
            cache_dir = os.path.join("history", conversation_id)
            os.makedirs(cache_dir, exist_ok=True)
        else:
            cache_dir = "cache"
            os.makedirs(cache_dir, exist_ok=True)
        
        log_dir = LOG_DIR
        
        # 处理表格生成总树
        start_time = time.time()
        f_tree = get_multiple_excel_feature_tree(files, log_dir=log_dir, vlm_cache=False)
        tree_json = f_tree.__json__()
        end_time = time.time()
        
        # 保存汇总产物
        save_tree_artifacts(f_tree, cache_dir, source_prefix="alldocument")
        with open(os.path.join(cache_dir, "temp1.json"), "w", encoding="utf-8") as f:
            json.dump(tree_json, f, ensure_ascii=False, indent=4)

        # 多文件场景：按最外层文件名额外保存每文件一组运行产物
        artifact_manifest = {}
        try:
            column_payload = f_tree.__json_column__()
            if isinstance(column_payload, dict):
                logger.info(f"[multi_file] column top keys={list(column_payload.keys())}")
                for top_key, top_body in column_payload.items():
                    single_tree = tag_feature_tree(
                        construct_feature_tree_for_fixed_json({top_key: top_body})
                    )
                    prefix = save_named_runtime_artifacts(single_tree, cache_dir, top_key)
                    artifact_manifest[str(top_key)] = {
                        "prefix": prefix,
                        "pkl": f"{prefix}.pkl",
                        "txt": f"{prefix}.txt",
                        "embedding": f"{prefix}.embedding.json",
                        "meta": _build_artifact_meta(top_key, single_tree, top_body=top_body),
                    }
                    logger.info(
                        f"[multi_file] built artifact key={top_key}, prefix={prefix}, "
                        f"pkl={prefix}.pkl, emb={prefix}.embedding.json"
                    )
        except Exception as ee:
            logger.error(f"save per-file runtime artifacts failed: {ee}")
        _save_artifact_manifest(cache_dir, artifact_manifest)
        logger.info(
            f"[multi_file] completed conversation_id={conversation_id}, cache_dir={cache_dir}, "
            f"artifact_count={len(artifact_manifest)}"
        )
            
        return tree_json
    except Exception as e:
        import traceback
        logger.error(f"多文件处理失败: {traceback.format_exc()}")
        return None


def analyze_multiple_files_for_route(files):
    """分析多个文件以确定处理线路"""
    if not files:
        return "请选择文件"
    
    # 分析所有文件
    has_image = False
    has_xlsx = False
    has_text = False
    file_details = []
    
    for file in files:
        file_path = file.name if hasattr(file, 'name') else file
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
            has_image = True
        elif ext in [".xlsx", ".xls", ".docx", ".doc"]:
            has_xlsx = True
        else:
            has_text = True  # 包括 .txt, .md, .json, .csv 等
        
        file_size = os.path.getsize(file_path)
        file_details.append({
            "path": file_path,
            "size": file_size,
            "ext": ext
        })
    
    # 按照优先级判断处理线路
    # 1. 如果有任何图片文件，所有文件一起走VLM
    if has_image:
        return "vlm"
    # 2. 如果有xlsx文件和纯文本内容，走HOTree
    elif has_xlsx:
        return "hotree"
    # 3. 如果只有纯文本文件，走LLM
    elif has_text:
        return "llm"
    else:
        return "llm"  # 默认


def determine_processing_route(file_path, file_size, file_content=None):
    """使用AI判断文件处理线路"""
    if not file_path:
        return "请选择文件"
    
    # 获取文件扩展名
    ext = os.path.splitext(file_path)[1].lower()
    
    # 构建提示词
    prompt = f"文件路径: {file_path}\n"
    prompt += f"文件大小: {file_size} 字节\n"
    prompt += f"文件类型: {ext}\n"
    if file_content:
        prompt += f"文件内容摘要: {file_content[:500]}...\n"
    prompt += "请根据以上信息判断应该使用哪种处理线路：\n"
    prompt += "1. 'llm'：纯文本内容，适合使用LLM处理\n"
    prompt += "2. 'vlm'：包含图片或需要视觉理解的内容，适合使用VLM处理\n"
    prompt += "3. 'hotree'：结构化数据或表格内容，适合使用H-OTree处理\n"
    prompt += "请只返回'llm'、'vlm'或'hotree'中的一个，不要添加任何其他解释。"
    
    try:
        # 调用LLM生成判断结果
        result = get_llm_generate(prompt, max_tokens=10, temperature=0.1)
        result = result.strip().lower()
        
        # 验证结果有效性
        if result in ["llm", "vlm", "hotree"]:
            return result
        else:
            # 如果AI返回无效结果，使用默认规则
            if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
                return "vlm"
            elif ext in [".xlsx", ".xls", ".docx", ".doc"]:
                return "hotree"
            else:
                return "llm"
    except Exception as e:
        # 如果AI调用失败，使用默认规则
        logger.error(f"AI判断线路失败: {e}")
        if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
            return "vlm"
        elif ext in [".xlsx", ".xls", ".docx", ".doc"]:
            return "hotree"
        else:
            return "llm"

def answer_question(
    qa_pair: dict,                          # 一条问答对
    table_file: str,                        # 表格原文件路径
    cache_dir: str,                           # 存储 HO-Tree 中间结果的路径
    enable_query_decompose: bool = True,    # 是否启用 Query Decomposition 机制
    enable_emebdding: bool = True,          # 是否启用 Embedding 机制
    log_dir: str = LOG_DIR,                 # Log 日志目录
    temperature: float = 0.5,               # LLM/VLM temperature
    max_tokens: int = 2048,                 # LLM/VLM max_tokens
    pkl_file_override: str = None,
    embedding_cache_file_override: str = None,
    table_id: str = "temp",
):
    
    query = qa_pair["query"]

    ##### 创建日志文件 命名为 表格id_问题id.log
    log_file = os.path.join(log_dir, f'temp.log')
    log_file_handler = logger.add(
        log_file,
        enqueue=False,  # 不使用队列，立即写入，避免缓冲
        backtrace=False,
        diagnose=False
    )

    logger.info(f"{DELIMITER} 开始问答问题 {DELIMITER}")

    start_time = time.time()

    logger.info(f"Question ID: temp")
    logger.info(f"Table ID: {table_id}")

    logger.info(f"Question: {query}")
    logger.info(f"Temperature: {temperature}")
    logger.info(f"Max tokens: {max_tokens}")

    ##### 加载 ho_tree
    pkl_file = pkl_file_override or os.path.join(cache_dir, 'temp.pkl')
    embedding_cache_file = embedding_cache_file_override or os.path.join(cache_dir, 'temp.embedding.json')
    with open(pkl_file, 'rb') as file:
        ho_tree = pickle.load(file)

    logger.info(f"Loading PKL File: {pkl_file}")
    logger.info(f"Loading Embedding Cache File: {embedding_cache_file}")

    final_answer, _, reliability = qa_RWP(
        query=query,
        ho_tree=ho_tree,
        table_file=table_file,
        table_id=table_id,
        embedding_cache_file=embedding_cache_file,
        enable_emebdding=enable_emebdding,
        enable_query_decompose=enable_query_decompose,
        temperature=temperature,
        max_tokens=max_tokens
    )
    qa_pair["reliability"] = reliability
    qa_pair["model_output"] = final_answer

    end_time = time.time()

    logger.info(f"{DELIMITER} 回答问题成功！ {DELIMITER}")
    logger.info(f"Cost time: {end_time - start_time}")
    
    logger.remove(log_file_handler)
    
    return qa_pair

def get_excel_feature_tree_multisheet(file: str,                   # 输入表格文件路径
                                     log_dir: str = LOG_DIR,      # LOG 日志记录路径
                                     vlm_cache: bool = False      # 是否保存转图片的中间结果
                                     ):
    """处理 Excel 文件中的所有 sheet，并构建成一棵总树"""
    # 1. 打开文件获取所有 sheet
    wb = openpyxl.load_workbook(file, data_only=True)
    combined_tree_dict = {}
    
    # 2. 循环处理每一个 sheet
    for sheet_name in wb.sheetnames:
        logger.info(f"正在处理 Sheet: {sheet_name}")
        sheet = preprocess_sheet(wb[sheet_name])
        # 获取该 sheet 的结构字典 (tree_dict)
        sheet_tree_dict = process_sheet_vlm(sheet, get_json=False, cache=vlm_cache)
        
        # 将每个 sheet 挂在以 sheet_name 命名的节点下
        combined_tree_dict[sheet_name] = sheet_tree_dict

    # 3. 传入大字典，一键生成多层级的总树
    # construct_feature_tree 会递归处理字典
    total_tree = construct_feature_tree(combined_tree_dict)
    
    # 4. 递归打标签
    total_tree = tag_feature_tree(total_tree)
    
    return total_tree


def process_table_for_tree(file, conversation_id=None):
    """专门处理表格，生成H-OTree结构"""
    global thinking_chain_data
    
    # 重置思维链条数据
    thinking_chain_data = {
        "question_answering": {},
        "retrieval_chains": []
    }
    
    if file is None:
        return "请先选择表格文件", ""
    
    # 注意：这里不再自动调用 clear_all()，由外部调用者根据需要决定
    try:
        # 设置缓存目录
        cache_dir, temp_dir = setup_cache_directory(conversation_id, "cache", "data/SSTQA/temp_tables")
        source_filename = os.path.splitext(os.path.basename(file.name))[0]
        
        # 定义日志目录（保持在全局LOG_DIR，用于调试目的）
        log_dir = LOG_DIR
                
        # 创建临时文件
        temp_file = os.path.join(temp_dir, "temp.xlsx")

        # 兼容多种格式，统一转为 xlsx
        src_path = file.name
        ext = os.path.splitext(src_path)[1].lower()
        try:
            if ext == ".xlsx":
                shutil.copy2(src_path, temp_file)
            elif ext == ".csv":
                df_src = pd.read_csv(src_path)
                df_src.to_excel(temp_file, index=False, engine="openpyxl")
            elif ext == ".html":
                html_content = open(src_path, "r", encoding="utf-8").read()
                html2workbook(html_content).save(temp_file)
            elif ext == ".md":
                md_content = open(src_path, "r", encoding="utf-8").read()
                table = extract_markdown_tables(md_content)
                with pd.ExcelWriter(temp_file, engine="openpyxl") as writer:
                    sheet_name = "sheet"
                    df_src = pd.DataFrame(table[1:], columns=table[0])
                    df_src.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                shutil.copy2(src_path, temp_file)
        except Exception as e:
            logger.error(f"格式转换失败: {e}")
            return "文件格式不支持或转换失败", ""
                
        # 读取表格
        df = pd.read_excel(temp_file)
            
        # 处理表格生成H-OTree
        start_time = time.time()
        # 使用多 Sheet 版本处理
        f_tree = get_excel_feature_tree_multisheet(temp_file, log_dir=log_dir, vlm_cache=False)
        tree_json = f_tree.__json__()
        end_time = time.time()
                
        # 保存中间文件
        save_tree_artifacts(f_tree, cache_dir, source_prefix=source_filename)
        # 保存额外的副本
        with open(os.path.join(cache_dir, f"temp1.json"), "w", encoding='utf-8') as f:
            json.dump(tree_json, f, indent=4, ensure_ascii=False)
        # 写入文件产物清单，供问答路由选择对应PKL
        # 顺序固定：先写 temp.column.json，再回读该文件生成 meta，避免元信息与最终列视图不一致。
        top_body_for_meta = None
        try:
            column_path = os.path.join(cache_dir, "temp.column.json")
            if os.path.exists(column_path):
                with open(column_path, "r", encoding="utf-8") as cf:
                    column_payload_for_meta = json.load(cf)
                if isinstance(column_payload_for_meta, dict):
                    if source_filename in column_payload_for_meta:
                        # 兼容最外层为文件名的结构
                        top_body_for_meta = column_payload_for_meta.get(source_filename)
                    elif len(column_payload_for_meta) == 1:
                        # 单文件常见结构：最外层是 sheet 名
                        only_sheet = next(iter(column_payload_for_meta.keys()))
                        top_body_for_meta = {only_sheet: column_payload_for_meta.get(only_sheet)}
                    else:
                        top_body_for_meta = column_payload_for_meta
            logger.info(
                f"[single_file] meta source resolved source={source_filename}, "
                f"top_body_type={type(top_body_for_meta).__name__}"
            )
        except Exception as e:
            logger.warning(f"[single_file] read temp.column.json for meta failed: {e}")

        manifest = {
            str(source_filename): {
                "prefix": _normalize_artifact_prefix(source_filename),
                "pkl": f"{_normalize_artifact_prefix(source_filename)}.pkl",
                "txt": f"{_normalize_artifact_prefix(source_filename)}.txt",
                "embedding": f"{_normalize_artifact_prefix(source_filename)}.embedding.json",
                "meta": _build_artifact_meta(source_filename, f_tree, top_body=top_body_for_meta),
            }
        }
        _save_artifact_manifest(cache_dir, manifest)
        build_and_save_tree_id_mappings(cache_dir, typed_root_name="HO_TREE")
        logger.info(f"[multi_file][single] saved source={source_filename}, cache_dir={cache_dir}")
        
        # 这里移除 gr.Info，避免在循环处理多文件时产生大量弹窗
        return tree_json
         
    except Exception as e:
        return handle_processing_error(e, "生成树")
    

def pure_llm_generate_answer(question, context="", temperature=0.5, max_tokens=2048):
    """Generate an answer using pure LLM."""
    if not question.strip():
        gr.Warning("Please enter a question")
        return "Please enter a question"
    try:
        # 构建提示词
        prompt = f"Question: {question}\n"
        if context:
            prompt += f"Context: {context}\n"
        prompt += "Please answer the question based on the information above."
        
        # 调用LLM生成答案
        answer = get_llm_generate(prompt, max_tokens, temperature)
        
        gr.Info("✅ LLM answer generated successfully!")
        return f"Answer: {answer}"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ LLM answer generation failed: {error_msg}")
        return "Failed to generate answer"


def pure_vlm_generate_answer(question, image_path, temperature=0.5, max_tokens=2048):
    """Generate an answer using pure VLM."""
    if not question.strip():
        gr.Warning("Please enter a question")
        return "Please enter a question"
    if not image_path:
        gr.Warning("Please select an image file")
        return "Please select an image file"
    try:
        # 构建提示词
        prompt = f"Question: {question}\n"
        prompt += "Please answer the question based on the image."
        
        # 调用VLM生成答案
        vlm_generate_func = get_vlm_generate()
        answer = vlm_generate_func(prompt, image_path, temperature)
        
        gr.Info("✅ VLM answer generated successfully!")
        return f"Answer: {answer}"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ VLM answer generation failed: {error_msg}")
        return "Failed to generate answer"


def process_file_with_route(file, question, temperature=0.5, max_tokens=2048, conversation_id=None):
    """根据文件类型自动选择处理线路，支持单个文件或多个文件"""
    if not file:
        gr.Warning("Please select a file")
        return "Please select a file"
    
    # 检查是否是多个文件
    if isinstance(file, list):
        return process_multiple_files_with_route(file, question, temperature, max_tokens, conversation_id=conversation_id)
    else:
        # 单个文件处理
        try:
            # 获取文件信息
            file_path = file.name
            file_size = os.path.getsize(file_path)
            
            # 读取文件内容摘要
            file_content = None
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    file_content = f.read(1000)
            except:
                # 二进制文件无法读取内容
                pass
            
            # 确定处理线路
            route = determine_processing_route(file_path, file_size, file_content)
            
            # 根据线路处理文件
            if route == "llm":
                # 纯LLM处理
                if not file_content:
                    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                        file_content = f.read()
                
                # 为LLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "text",
                        "file_path": file_path,
                        "file_content_preview": file_content[:500],  # 限制长度
                        "processing_method": "llm"
                    }
                    
                    # 为文本内容生成嵌入向量
                    embedding_texts = [f"Text file: {os.path.basename(file_path)}, content: {file_content[:500]}"]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                return pure_llm_generate_answer(question, file_content, temperature, max_tokens)
            elif route == "vlm":
                # 纯VLM处理
                # 为VLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "image",
                        "file_path": file_path,
                        "processing_method": "vlm"
                    }
                    
                    # 为图像上下文生成简单的嵌入向量
                    embedding_texts = [f"Image file: {os.path.basename(file_path)}"]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                return pure_vlm_generate_answer(question, file_path, temperature, max_tokens)
            elif route == "hotree":
                # H-OTree处理
                wrapped_file = types.SimpleNamespace(name=file.name)
                # process_table_for_tree会将处理结果保存到临时文件，供后续问答使用
                data = process_table_for_tree(wrapped_file, conversation_id=conversation_id)
                if data:
                    # 使用H-OTree方法回答问题
                    result = process_question_only(question, temperature, max_tokens, conversation_id=conversation_id)
                    return result
                else:
                    return "H-OTree处理失败"
            else:
                return f"未知处理线路: {route}"
        except Exception as e:
            import traceback
            error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
            gr.Warning(f"❌ 文件处理失败: {error_msg}")
            return "文件处理失败"


def process_multiple_files_with_route(files, question, temperature=0.5, max_tokens=2048, conversation_id=None):
    """处理多个文件，根据文件类型自动选择处理线路"""
    if not files or len(files) == 0:
        gr.Warning("请选择文件")
        return "请选择文件"
    
    try:
        # 分析多个文件以确定处理线路
        route = analyze_multiple_files_for_route(files)
        
        # 根据线路处理文件
        if route == "llm":
            # 纯LLM处理 - 合并所有文本文件内容
            combined_content = ""
            for file in files:
                file_path = file.name
                try:
                    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read()
                        combined_content += f"\n--- 文件: {os.path.basename(file_path)} ---\n{content}\n"
                except:
                    # 非文本文件跳过或简单描述
                    combined_content += f"\n--- 文件: {os.path.basename(file_path)} (非文本文件) ---\n"
            
            # 为LLM对话创建必要的处理文件，以便后续历史记录加载
            if conversation_id:
                placeholder_data = {
                    "file_type": "text",
                    "combined_content": combined_content[:500],  # 限制长度
                    "processing_method": "llm"
                }
                
                # 为文本内容生成嵌入向量
                embedding_texts = [f"Text content: {combined_content[:500]}"]
                ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
            
            return pure_llm_generate_answer(question, combined_content, temperature, max_tokens)
        elif route == "vlm":
            # VLM处理 - 优先处理图片文件，但需要考虑其他文件
            # 为了更好地处理混合内容，我们先处理图片，然后将其他文件内容作为上下文
            image_files = []
            table_content = ""  # 存储表格转换后的JSON内容
            other_content = ""  # 存储其他文件内容
            
            for file in files:
                file_path = file.name
                ext = os.path.splitext(file_path)[1].lower()
                
                if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
                    image_files.append(file_path)
                elif ext in [".xlsx", ".xls", ".csv", ".docx", ".doc"]:  # 表格文件
                    try:
                        # 将表格文件转换为HOTree JSON格式
                        wrapped_file = types.SimpleNamespace(name=file.name)
                        tree_json = process_table_for_tree(wrapped_file, conversation_id=conversation_id)
                        if tree_json:
                            table_content += f"\n--- 表格文件 {os.path.basename(file_path)} 的HOTree JSON结构: {json.dumps(tree_json, ensure_ascii=False, indent=2)} ---\n"
                        else:
                            # 如果转换失败，尝试作为普通文本读取
                            try:
                                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                                    content = f.read(1000)  # 读取前1000个字符作为上下文
                                    other_content += f"\n--- 其他文件 {os.path.basename(file_path)} 内容: {content} ---\n"
                            except:
                                other_content += f"\n--- 其他文件 {os.path.basename(file_path)} (非文本文件) ---\n"
                    except Exception as e:
                        # 如果表格转换失败，尝试作为普通文本读取
                        try:
                            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                                content = f.read(1000)  # 读取前1000个字符作为上下文
                                other_content += f"\n--- 其他文件 {os.path.basename(file_path)} 内容: {content} ---\n"
                        except:
                            other_content += f"\n--- 其他文件 {os.path.basename(file_path)} (非文本文件) ---\n"
                else:  # 其他类型的文件
                    try:
                        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                            content = f.read(1000)  # 读取前1000个字符作为上下文
                            other_content += f"\n--- 其他文件 {os.path.basename(file_path)} 内容: {content} ---\n"
                    except:
                        other_content += f"\n--- 其他文件 {os.path.basename(file_path)} (非文本文件) ---\n"
            
            # 如果有图片文件，使用第一个图片文件并附加上下文
            if image_files:
                combined_context = ""
                if table_content:
                    combined_context += f"表格数据: {table_content}\n"
                if other_content:
                    combined_context += f"其他文件信息: {other_content}"
                
                # 为VLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "image",
                        "image_files": image_files,
                        "table_content": table_content,
                        "other_content": other_content,
                        "processing_method": "vlm"
                    }
                    
                    # 为图像上下文生成简单的嵌入向量
                    context_text = f"Image files: {', '.join([os.path.basename(img) for img in image_files])}"
                    if table_content:
                        context_text += f"; Table content: {table_content[:200]}"  # 限制长度
                    if other_content:
                        context_text += f"; Other content: {other_content[:200]}"  # 限制长度
                    
                    embedding_texts = [context_text]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                if combined_context:
                    enhanced_question = f"{question}\n\n{combined_context}"
                    return pure_vlm_generate_answer(enhanced_question, image_files[0], temperature, max_tokens)
                else:
                    return pure_vlm_generate_answer(question, image_files[0], temperature, max_tokens)
            else:
                # 如果没有找到图片文件但路线是VLM，使用第一个文件
                # 为VLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "image",
                        "file_path": files[0].name,
                        "processing_method": "vlm"
                    }
                    
                    # 为图像上下文生成简单的嵌入向量
                    embedding_texts = [f"Image file: {os.path.basename(files[0].name)}"]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                return pure_vlm_generate_answer(question, files[0].name, temperature, max_tokens)
        elif route == "hotree":
            # H-OTree处理 - 将所有表格文件合并为一棵树
            # 过滤出所有表格文件
            table_files = []
            for file in files:
                file_path = file.name if hasattr(file, 'name') else str(file)
                ext = os.path.splitext(file_path)[1].lower()
                if ext in [".xlsx", ".xls", ".csv", ".docx", ".doc"]:
                    table_files.append(file)
            
            if table_files:
                # --- 简化逻辑：只要有 pkl 就不重新解析 ---
                # 如果提供了 conversation_id，则使用历史记录文件夹，否则使用 cache 目录
                if conversation_id:
                    cache_dir = os.path.join("history", conversation_id)
                else:
                    cache_dir = "cache"
                cache_pkl = os.path.join(cache_dir, "temp.pkl")
                
                if os.path.exists(cache_pkl):
                    logger.info("检测到本地 H-OTree 缓存，跳过解析流程。")
                else:
                    logger.info("未检测到缓存，开始执行 H-OTree 完整解析...")
                    # 只有在没有缓存时才进行解析
                    data = process_multiple_tables_for_tree(table_files, conversation_id=conversation_id)
                    if not data:
                        return "多文件 H-OTree 解析失败"

                # 使用H-OTree方法回答问题
                result = process_question_only(question, temperature, max_tokens, conversation_id=conversation_id)
                return result
        else:
            return f"未知处理线路: {route}"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ 多文件处理失败: {error_msg}")
        return "多文件处理失败"


def process_question_only(question, temperature=0.5, max_tokens=2048, conversation_id=None):
    """专门处理问题，返回答案"""
    # 如果提供了 conversation_id，则使用历史记录文件夹中的文件，否则使用默认路径
    if conversation_id:
        cache_dir = os.path.join("history", conversation_id)
        
        # 检查缓存目录中的必要文件是否存在（支持每文件独立产物）
        required_files = [
            os.path.join(cache_dir, "temp.pkl"),
            os.path.join(cache_dir, "temp.embedding.json")
        ]
        manifest = _load_artifact_manifest(cache_dir)
        has_per_file_artifacts = False
        if manifest:
            for item in manifest.values():
                if not isinstance(item, dict):
                    continue
                pkl = os.path.join(cache_dir, str(item.get("pkl", "")))
                emb = os.path.join(cache_dir, str(item.get("embedding", "")))
                if os.path.exists(pkl) and os.path.exists(emb):
                    has_per_file_artifacts = True
                    break
        missing_files = [f for f in required_files if not os.path.exists(f)]
        if missing_files and not has_per_file_artifacts:
            gr.Warning("无法继续此对话：缺少必要的处理文件。请重新上传文件后再继续。")
            return "无法继续此对话：缺少必要的处理文件。请重新上传文件后再继续。"
        
        # 读取占位符数据以确定对话类型
        pkl_file = os.path.join(cache_dir, "temp.pkl")
        if os.path.exists(pkl_file):
            with open(pkl_file, 'rb') as f:
                placeholder_data = pickle.load(f)
        else:
            placeholder_data = {"processing_method": "hotree", "file_type": "table"}
        
        # 根据对话类型选择处理方法
        if isinstance(placeholder_data, dict):
            processing_method = placeholder_data.get("processing_method", "hotree")
            file_type = placeholder_data.get("file_type", "table")
        else:
            # 旧版本数据格式，假设为表格类型
            processing_method = "hotree"
            file_type = "table"
        
        # 根据处理方法决定如何处理问题
        if processing_method == "vlm":
            # 对于VLM对话，使用VLM方法继续处理
            image_files = placeholder_data.get("image_files", [])
            file_path = placeholder_data.get("file_path", "")
            table_content = placeholder_data.get("table_content", "")
            other_content = placeholder_data.get("other_content", "")
            
            # 构建上下文
            combined_context = ""
            if table_content:
                combined_context += f"表格数据: {table_content}\n"
            if other_content:
                combined_context += f"其他文件信息: {other_content}"
            
            # 选择要使用的图片文件
            image_to_use = image_files[0] if image_files else file_path
            if image_to_use:
                enhanced_question = question
                if combined_context:
                    enhanced_question = f"{question}\n\n{combined_context}"
                
                return pure_vlm_generate_answer(enhanced_question, image_to_use, temperature, max_tokens)
            else:
                # 如果没有图片文件，返回错误
                gr.Warning("VLM对话缺少图片文件")
                return "VLM对话缺少图片文件"
                
        elif processing_method == "llm":
            # 对于LLM对话，使用LLM方法继续处理
            combined_content = placeholder_data.get("combined_content", "")
            file_content_preview = placeholder_data.get("file_content_preview", "")
            
            # 合并所有可用内容
            full_content = combined_content or file_content_preview
            if full_content:
                return pure_llm_generate_answer(question, full_content, temperature, max_tokens)
            else:
                # 如果没有内容，可以使用通用LLM回答
                return pure_llm_generate_answer(question, "", temperature, max_tokens)
        
        else:
            # 对于HOTree对话，使用原来的处理方法
            table_file = os.path.join("history", conversation_id, "temp.xlsx")
            if not os.path.exists(table_file):
                # 保存后的编辑以 JSON/PKL 为准；若原始 xlsx 缺失，继续尝试问答。
                logger.warning(f"HOTree conversation table file missing, continue with PKL only: {table_file}")
    else:
        table_file = "data/SSTQA/temp_tables/temp.xlsx"
        cache_dir = "cache"
        
        if not os.path.exists(table_file):
            pkl_file = os.path.join(cache_dir, "temp.pkl")
            if not os.path.exists(pkl_file):
                gr.Warning("Please upload a table first")
                return "Please upload a table first"
            logger.warning(f"Default table file missing, continue with PKL only: {table_file}")
    
    if not question.strip():
        gr.Warning("Please enter a question")
        return "Please enter a question"
    
    # 对于HOTree对话（包括没有conversation_id的新对话），使用原来的处理方法
    try:
        # 记录参数变更日志（使用 loguru 格式：时间 | 级别 | 内容）
        param_log_file = os.path.join(LOG_DIR, "param_change.log")
        os.makedirs(LOG_DIR, exist_ok=True)
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        msg = f"{timestamp} | PARAM_CHANGE | temperature={temperature}, max_tokens={max_tokens}\n"
        with open(param_log_file, "a", encoding="utf-8") as f:
            f.write(msg)
        qa_pair = {
            "id": "temp",
            "table_id": "temp",
            "query": question.strip()
        }
        manifest = _load_artifact_manifest(cache_dir)
        routed_keys = _choose_artifact_keys_for_query(question, manifest)
        if not routed_keys:
            routed_keys = ["__default__"]
        elif "__default__" not in routed_keys:
            routed_keys.append("__default__")
        logger.info(
            f"[artifact_route] start question={question}, routed_keys={routed_keys}, "
            f"manifest_keys={list(manifest.keys()) if isinstance(manifest, dict) else []}"
        )

        result = None
        last_error = None
        route_attempts = []
        for artifact_key in routed_keys:
            if artifact_key == "__default__":
                pkl_override = None
                emb_override = None
                inferred_scope = ""
                try:
                    manifest_keys = [str(k or "").strip() for k in (manifest.keys() if isinstance(manifest, dict) else [])]
                    manifest_keys = [k for k in manifest_keys if k]
                    if len(manifest_keys) == 1:
                        inferred_scope = manifest_keys[0]
                except Exception:
                    inferred_scope = ""
                table_id = inferred_scope or "temp"
                logger.info(
                    f"[artifact_route] try default artifact (legacy temp.*), "
                    f"table_scope={table_id}"
                )
                route_attempts.append({"key": artifact_key, "table_id": table_id, "status": "trying_default"})
            else:
                item = manifest.get(artifact_key, {}) if isinstance(manifest, dict) else {}
                pkl_name = str(item.get("pkl", "") or "")
                emb_name = str(item.get("embedding", "") or "")
                pkl_override = os.path.join(cache_dir, pkl_name) if pkl_name else None
                emb_override = os.path.join(cache_dir, emb_name) if emb_name else None
                if not (pkl_override and emb_override and os.path.exists(pkl_override) and os.path.exists(emb_override)):
                    logger.warning(
                        f"[artifact_route] skip key={artifact_key}, missing files pkl={pkl_override}, emb={emb_override}"
                    )
                    route_attempts.append({"key": artifact_key, "table_id": artifact_key, "status": "skip_missing_files"})
                    continue
                table_id = artifact_key
                logger.info(
                    f"[artifact_route] try key={artifact_key}, pkl={pkl_override}, emb={emb_override}"
                )
                route_attempts.append({"key": artifact_key, "table_id": table_id, "status": "trying"})
            try:
                result = answer_question(
                    qa_pair=dict(qa_pair),
                    table_file=table_file,
                    cache_dir=cache_dir,
                    enable_emebdding=True,
                    enable_query_decompose=True,
                    log_dir=LOG_DIR,
                    temperature=temperature,
                    max_tokens=max_tokens,
                    pkl_file_override=pkl_override,
                    embedding_cache_file_override=emb_override,
                    table_id=table_id,
                )
                reliability = result.get("reliability", None) if isinstance(result, dict) else None
                if reliability in (False, "False", "false", 0, "0"):
                    logger.warning(
                        f"[artifact_route] fallback from key={artifact_key}, reason=low_reliability, reliability={reliability}"
                    )
                    route_attempts.append(
                        {"key": artifact_key, "table_id": table_id, "status": "fallback_low_reliability", "reliability": reliability}
                    )
                    continue
                logger.info(f"[artifact_route] selected key={artifact_key}, reliability={reliability}")
                route_attempts.append(
                    {"key": artifact_key, "table_id": table_id, "status": "selected", "reliability": reliability}
                )
                break
            except Exception as ee:
                last_error = ee
                logger.warning(f"[artifact_route] fallback from key={artifact_key}, reason=exception, err={ee}")
                route_attempts.append(
                    {"key": artifact_key, "table_id": table_id, "status": "fallback_exception", "error": str(ee)}
                )
                continue
        logger.info(f"[artifact_route] attempts={json.dumps(route_attempts, ensure_ascii=False)}")
        if result is None and last_error is not None:
            raise last_error
        if result :
            gr.Info("✅ Answer generated successfully!")
            return f"Answer: {result.get('model_output', 'No answer')}\n\nConfidence: {result.get('reliability', 'Unknown')}"
        else:
            gr.Warning("❌ Failed to generate answer")
            return "Failed to generate answer"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ Failed to generate answer: {error_msg}")
        return "Failed to generate answer"

def clear_all():
    """清除所有内容并删除相关文件"""
    import shutil
    import os    
    # 返回空输出以匹配界面清空
    empty_outputs = (None, "", "", {}, None, "")
    # 删除临时表格文件
    temp_dir = "data/SSTQA/temp_tables"
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
        os.makedirs(temp_dir, exist_ok=True)  # 重新创建空目录
    
    # 删除log目录下的所有文件
    log_dir = LOG_DIR
    clear_directory_contents(log_dir)

    # 删除cache目录下的所有文件
    cache_dir = "cache"
    clear_directory_contents(cache_dir)
    
    # 重置思维链条数据
    global thinking_chain_data
    thinking_chain_data = {
        "question_answering": {},
        "retrieval_chains": []
    }
    
    return None, "", "", {}, None, ""  # 清空所有界面组件（含图谱）

def get_thinking_chain():
    """获取思维链条数据"""
    global thinking_chain_data
    return thinking_chain_data

def read_all_logs(log_dir=LOG_DIR, max_lines=200):
    """合并读取所有日志文件，按时间顺序显示，并添加颜色美化"""
    all_lines = []
    
    # 优先读取的特定日志文件
    priority_log_files = [
        os.path.join(log_dir, "temp.xlsx.log"),
        os.path.join(log_dir, "param_change.log"),
        os.path.join(log_dir, "temp.log"),
    ]
    
    # 读取所有 .log 文件（包括app.log）
    all_log_files = []
    if os.path.exists(log_dir):
        for file in os.listdir(log_dir):
            if file.endswith('.log'):
                file_path = os.path.join(log_dir, file)
                if os.path.isfile(file_path):
                    all_log_files.append(file_path)
    
    # 合并优先文件和所有日志文件，去重
    log_files = list(set(priority_log_files + all_log_files))
    
    for log_path in log_files:
        if os.path.exists(log_path):
            try:
                with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
                    lines = f.readlines()
                    all_lines.extend(lines)
            except Exception as e:
                all_lines.append(f"[ERROR] 读取 {log_path} 失败: {e}\n")
    
    # 按时间戳排序（loguru 格式：时间 | 级别 | ...）
    try:
        all_lines.sort(key=lambda x: x.split("|")[0].strip() if "|" in x else "")
    except Exception:
        pass
    
    # 取最后 max_lines 行
    log_content = "".join(all_lines[-max_lines:]) if all_lines else "暂无日志"
    
    # 添加颜色美化 - 将日志转换为HTML格式
    # 支持 loguru 格式: 时间 | 级别 | 内容
    html_lines = []
    for line in log_content.split("\n"):
        if "|" in line and len(line.split("|")) >= 3:
            parts = line.split("|", 2)
            timestamp = parts[0].strip()
            level = parts[1].strip()
            content = parts[2].strip()
            # 为时间戳添加蓝色，为日志级别添加绿色
            html_line = f"<span style='color: blue'>{timestamp}</span> | <span style='color: green'>{level}</span> | {content}<br>"
        else:
            # 非标准格式行保持原样
            html_line = line + "<br>"
        html_lines.append(html_line)
    
    # 包装在<pre>标签中以保留格式，但使用HTML允许颜色显示
    return f"<pre style='font-family: monospace; white-space: pre-wrap; word-wrap: break-word;'>{' '.join(html_lines)}</pre>"


def tree_json_to_table_dict(tree_json, table_name="edited_table"):
    """
    将前端树 JSON（列表）递归转换为 construct_feature_tree 可用的 table_dict:
    {table_name: <嵌套结构/list/值>}#处理三种情况 dict of dict /dict with list /listofdict 无listoflist

    规则：
     他的规则是这样的：读json 读到每一个node
     1.如果有children而且有很多个children 构建一个dict dict的name是dict的index 然后dict的children是他的value value是一个list，返回这个dict
     2.如果只有一个children 那么他的value就是个string/float 总而言之就是一个值，就构建一个dict index=项目 value=children 
     3.如果没有children 就返回name（string类型）不建立dict
    - 
    """
    def convert_node(node: dict):   #处理dict with list of dict  返回一个dict
                                    
        if not isinstance(node, dict):
            return None
        name = str(node.get("name", ""))
        children = node.get("children", [])
        if children:
            if len(children)==1:
                child=children[0]
                return {name:convert_node(child)} #返回一个dict 
            # 孩子是一个list 直接传入convert children
            else:
                merged=convert_children(children)
                return {name: merged}#name+list
        # 无 children -> 叶子，返回 name 字符串
        return name

    def convert_children(children: list):#输入一个list 如果list里面每一个dict名字都是序号 返回一个list 里面是去掉序号包装的dict
                                         #如果不是这样 那么返回一个dict
        if not children:#
            return []
        else:#如果所有的孩子都是name：【n】，那么处理每个孩子的孩子为一个dict，加到list 返回list
            if all(isinstance(c, dict) and re.fullmatch(r"\[\d+\]", str(c.get("name", ""))) for c in children):
                lst = []
                for c in children:
                    v = c.get("children", [])
                    if isinstance(v, list):#一般来说序号的value就是list
                        merged = {}
                        for item in v:
                            item1=convert_node(item);
                            if isinstance(item1, dict) and len(item1) == 1:
                                k, val = next(iter(item1.items()))
                                merged[k] = val#组装成一个dict
                            else:#防御性 应该不会
                                merged[str(len(merged))] = item1
                        lst.append(merged)#加到list中
                return lst#返回不带序号的list

        # 保持顺序的列表，不再合并/排序
        #如果孩子的名字都是正常的那么这个时候传入的
        return [convert_node(c) for c in children]

    if not isinstance(tree_json, list) or not tree_json:
        return {table_name: []}

    # 使用最外层节点作为表节点，不再下钻 children[0]
    #tree_json是一个list of dict 暂时认为只有一个dict
    outer = tree_json[0] if isinstance(tree_json[0], dict) else {}
    logger.info(f"[debug] parsed root: {outer}")
    table_node = outer
    table_name = table_node.get("name", table_name) if isinstance(table_node, dict) else table_name
    table_children = table_node.get("children", []) if isinstance(table_node, dict) else []

    logger.info(f"[debug] table_name: {table_name}, child_count: {len(table_children)}")
    table_body = convert_children(table_children)#直接传入children_list
    return {table_name: table_body}#返回一个dict


def _is_item_wrapper_name(name: str) -> bool:
    return bool(re.fullmatch(r"item\[\d+\]", str(name or "").strip()))


def _flat_column_tree_to_table_dict(tree_json, default_table_name="edited_table"):
    """
    将 flat column view 前端树（name/children + nodeType）还原为精炼 dict/list：
    - 去掉 item[i] 包装节点
    - 不生成 [0]/[1] 这类中间索引键
    """
    if not isinstance(tree_json, list) or not tree_json or not isinstance(tree_json[0], dict):
        return {default_table_name: []}

    root = tree_json[0]
    root_name = str(root.get("name", "") or "").strip().lower()
    root_children = root.get("children", []) if isinstance(root.get("children", []), list) else []

    def parse_children(children):
        if not isinstance(children, list) or not children:
            return []
        parsed = [parse_node(ch) for ch in children if isinstance(ch, dict)]
        parsed = [p for p in parsed if p is not None]

        if not parsed:
            return []

        # dict 分支：[{k:v}, {k:v}] -> {k:v, k:v}
        if all(isinstance(x, dict) for x in parsed):
            merged = {}
            for item in parsed:
                for k, v in item.items():
                    merged[str(k)] = v
            return merged

        # list 分支：保留顺序
        return parsed

    def parse_node(node):
        if not isinstance(node, dict):
            return None
        name = str(node.get("name", "") or "")
        children = node.get("children", []) if isinstance(node.get("children", []), list) else []
        node_type = str(node.get("nodeType", "") or "")

        if not children:
            return name

        parsed_child = parse_children(children)

        # 去掉 list item 包装层
        if _is_item_wrapper_name(name):
            return parsed_child

        # M_NODE 表示 key-value 的 key
        if node_type == "M_NODE":
            return {name: parsed_child}

        # 其他带 children 的节点（如 FEATURE_TREE/B_NODE group）直接透传子结构
        return parsed_child

    body = parse_children(root_children)

    # flat column view 根节点仅是视图标签，不作为业务 key 落盘
    if root_name in {"flat column view", "flat row view", "flat view"}:
        if isinstance(body, dict):
            return body
        return {default_table_name: body}

    # 非 flat root，保守保留根名
    return {str(root.get("name", default_table_name)): body}


def _looks_like_frontend_tree_payload(payload):
    if isinstance(payload, dict):
        if "name" in payload:
            return True
        return False
    if isinstance(payload, list) and payload:
        return all(isinstance(x, dict) and "name" in x for x in payload)
    return False


def normalize_save_payload_to_table_dict(payload, table_name="edited_table"):
    """
    统一保存入口数据：
    - 前端树格式（name/children） -> 转换为 dict/list canonical
    - 已是 dict/list canonical -> 直接使用
    """
    if _looks_like_frontend_tree_payload(payload):
        as_list = payload if isinstance(payload, list) else [payload]
        root_name = ""
        if as_list and isinstance(as_list[0], dict):
            root_name = str(as_list[0].get("name", "") or "").strip().lower()
        logger.info(
            f"[save_payload] detected frontend tree payload, root={root_name}, converting to canonical dict/list"
        )
        if "flat column view" in root_name:
            converted = _flat_column_tree_to_table_dict(as_list, default_table_name=table_name)
        else:
            converted = tree_json_to_table_dict(as_list, table_name=table_name)
        logger.info(
            f"[save_payload] converted keys={list(converted.keys()) if isinstance(converted, dict) else type(converted)}"
        )
        return converted

    if isinstance(payload, dict):
        logger.info(f"[save_payload] detected canonical dict payload, keys={list(payload.keys())}")
        return payload
    if isinstance(payload, list):
        logger.info("[save_payload] detected canonical list payload, wrapped under table_name")
        return {table_name: payload}

    logger.warning(f"[save_payload] unexpected payload type={type(payload)}, fallback empty")
    return {table_name: []}


def construct_feature_tree_simple(obj, name="root"):#认为一定是dict+list of dict的结构，没有listof list
    
    index_tree = IndexTree()
    body_tree = BodyTree()

    def add_pair(idx_value, body_value):
        index_node = IndexNode(value=idx_value)
        body_node = BodyNode(body_value)
        index_node.body = [body_node]
        index_tree.add_index(index_node)
        body_tree.add_deep(body_node)

    if isinstance(obj, dict):
        for k, v in obj.items():
            logger.info(f"[debug] dict key: {k}")
            if isinstance(v, (dict, list)):#如果value是list
                child_tree = construct_feature_tree_simple(v, name=str(k))
                add_pair(str(k), child_tree)
            else:
                add_pair(str(k), v)#如果value是string/float 那么直接添加到body tree中
        return FeatureTree(index_tree=index_tree, body_tree=body_tree)

    if isinstance(obj, list):  # 行列表或通用列表
        for i, v in enumerate(obj):
            idx = f"[{i}]"
            if isinstance(v, (dict, list)):
                child_tree = construct_feature_tree_simple(v, name=idx)
                add_pair(idx, child_tree)
            else:
                add_pair(idx, v)
        return FeatureTree(index_tree=index_tree, body_tree=body_tree)

    # 原子值兜底
    add_pair(str(name) if name is not None else "root", obj)
    return FeatureTree(index_tree=index_tree, body_tree=body_tree)


def rebuild_feature_tree_from_json(tree_json, cache_dir="cache", temp_dir="data/SSTQA/temp_tables", log_dir=LOG_DIR):
    try:
        def strip_ids(obj):
            if isinstance(obj, list):
                return [strip_ids(o) for o in obj]
            if isinstance(obj, dict):
                return {
                    k: strip_ids(v)
                    for k, v in obj.items()
                    if k != "id"
                }
            return obj

        # 临时调试日志，观测前端传入结构（仅取前 1 个元素，防止日志过大）
        try:
            sample = tree_json[:1] if isinstance(tree_json, list) else tree_json
            logger.info(f"[debug] tree_json sample: {sample}")
        except Exception as e:
            logger.warning(f"[debug] cannot log tree_json sample: {e}")

        os.makedirs(cache_dir, exist_ok=True)
        os.makedirs(temp_dir, exist_ok=True)
        os.makedirs(log_dir, exist_ok=True)

        # 去掉 id，防止写回文件
        cleaned_tree = strip_ids(tree_json)

        # 不再写入 temp.json（已停用该中间文件）

        # JSON -> FeatureTree
        tree_dict = tree_json_to_table_dict(cleaned_tree)
        logger.info(f"[debug] parsed tree_dict: {tree_dict}")

        # 尝试用简化构造逻辑，以便保留 list 内的 dict 结构
        ho_tree = None
        try:
            if isinstance(tree_dict, dict) and len(tree_dict) == 1:
                t_name, t_body = next(iter(tree_dict.items()))
                ho_tree = construct_feature_tree_simple({t_name: t_body}, name="root")
                logger.info("[debug] ho_tree built by construct_feature_tree_simple")
            else:
                ho_tree = construct_feature_tree_simple(tree_dict, name="root")
        except Exception as ee:
            logger.warning(f"[debug] simple construct failed, fallback to original: {ee}")
            ho_tree = construct_feature_tree(tree_dict)

        try:
            idx_children = [n.value for n in ho_tree.index_tree.root.children]
            logger.info(f"[debug] ho_tree index root children: {idx_children}")
        except Exception as ee:
            logger.warning(f"[debug] cannot log index tree children: {ee}")

        ho_tree = tag_feature_tree(ho_tree)

        # 保存文本、对象、embedding
        with open(os.path.join(cache_dir, "temp.txt"), "w", encoding="utf-8") as f:
            f.write(ho_tree.__str__([1]))
        with open(os.path.join(cache_dir, "temp.pkl"), "wb") as f:
            pickle.dump(ho_tree, f)
        # 额外保存 HO-Tree 规范化 JSON，若原始仅单表，则包上一层表名，避免 __json__ 展开丢失表头
        ho_json = ho_tree.__json__()
        ho_column_json = ho_tree.__json_column__()
        with open(os.path.join(cache_dir, "temp.ho.json"), "w", encoding="utf-8") as f:
            json.dump(ho_json, f, ensure_ascii=False, indent=4)
        with open(os.path.join(cache_dir, "temp.column.json"), "w", encoding="utf-8") as f:
            json.dump(ho_column_json, f, ensure_ascii=False, indent=4)

        # 生成 embedding，失败不阻塞（确保输入为字符串列表）
        try:
            raw_values = ho_tree.all_value_list()
            texts = [str(x) for x in raw_values] if raw_values else []
            if texts:
                embedding_dict = EmbeddingModel().get_embedding_dict(texts)
                EmbeddingModel().save_embedding_dict(
                    embedding_dict,
                    os.path.join(cache_dir, "temp.embedding.json")
                )
            else:
                logger.warning("embedding skipped: empty value list")
        except Exception as ee:
            logger.error(f"embedding generate failed: {ee}")

        return True, "ok"
    except Exception as e:
        logger.error(f"rebuild_feature_tree_from_json failed: {e}")
        return False, str(e)


def _split_fixed_body_values(body):
    """Split scalar/list values into multiple body items by comma."""
    values = []
    if isinstance(body, list):
        for item in body:
            text = "" if item is None else str(item)
            values.extend([x.strip() for x in text.split(",") if x.strip() != ""])
    elif body is None:
        values = [""]
    else:
        text = str(body)
        values = [x.strip() for x in text.split(",") if x.strip() != ""]
    if not values:
        values = [""]
    return values


def construct_feature_tree_for_fixed_json(tree_dict):
    """
    Secondary rebuild-only constructor.
    Keep the initial-generation construct_feature_tree() untouched.
    """
    index_tree = IndexTree()
    body_tree = BodyTree()

    if not isinstance(tree_dict, dict):
        tree_dict = {}

    index_nodes = []
    has_subtree_body = False

    for index, body in tree_dict.items():
        if isinstance(body, dict):
            body_nodes = [BodyNode(construct_feature_tree_for_fixed_json(body))]
        else:
            parts = _split_fixed_body_values(body)
            body_nodes = [BodyNode(part) for part in parts]

        if any(isinstance(bn.value, FeatureTree) for bn in body_nodes):
            has_subtree_body = True

        index_node = IndexNode(value=index)
        index_node.body = body_nodes
        index_tree.add_index(index_node)
        index_nodes.append(index_node)

    # Row-aligned BodyTree construction for pure column-value payloads:
    # root -> row_0_head -> row_0_col_2 -> ... ; root -> row_1_head -> ...
    # This matches __json__ list-mode expectation (one path per row).
    if index_nodes and not has_subtree_body:
        row_count = max((len(n.body) for n in index_nodes), default=0)
        if row_count > 0:
            # Pad short columns so each row path keeps full schema width.
            for idx_node in index_nodes:
                while len(idx_node.body) < row_count:
                    idx_node.body.append(BodyNode(""))

            root = body_tree.root
            for row_idx in range(row_count):
                row_nodes = [idx_node.body[row_idx] for idx_node in index_nodes]
                if not row_nodes:
                    continue
                head = row_nodes[0]
                root.add_child(head)
                head.add_father(root)
                prev = head
                for curr in row_nodes[1:]:
                    prev.add_child(curr)
                    curr.add_father(prev)
                    prev = curr
    else:
        # Keep legacy fallback for subtree-heavy structures.
        for idx_node in index_nodes:
            for body_node in idx_node.body:
                body_tree.add_deep(body_node)

    return FeatureTree(index_tree=index_tree, body_tree=body_tree)


def rebuild_feature_views_from_json(
    tree_json,
    conversation_id=None,
    cache_dir="cache",
    log_dir=LOG_DIR,
):
    """
    Rebuild editable view artifacts from frontend tree json.
    New policy:
    - canonical source for edited data is temp.column.json
    - temp1.json is synchronized from the same rebuilt FeatureTree
    - do NOT regenerate temp.json here
    """
    try:
        def strip_ids(obj):
            if isinstance(obj, list):
                return [strip_ids(o) for o in obj]
            if isinstance(obj, dict):
                return {k: strip_ids(v) for k, v in obj.items() if k != "id"}
            return obj

        cleaned_tree = strip_ids(tree_json)
        tree_dict = normalize_save_payload_to_table_dict(cleaned_tree)
        logger.info(f"[save_views] parsed tree_dict keys: {list(tree_dict.keys()) if isinstance(tree_dict, dict) else type(tree_dict)}")

        # 1) canonical payload -> fixed column json
        # NOTE: do NOT pass through construct_feature_tree_simple() here.
        # It may inject intermediate structures that are not part of canonical payload.
        if isinstance(tree_dict, dict):
            fixed_column_json = tree_dict
        else:
            fixed_column_json = {"edited_table": tree_dict}
        logger.info(
            f"[save_views] fixed_column_json keys: {list(fixed_column_json.keys()) if isinstance(fixed_column_json, dict) else type(fixed_column_json)}"
        )

        # 2) temp.fixed.json -> FeatureTree
        # Multi-file loop is intentionally handled here (caller level), not inside
        # the initial-generation constructor path.
        per_file_trees = {}
        if isinstance(fixed_column_json, dict) and fixed_column_json:
            merged_index_tree = IndexTree()
            merged_body_tree = BodyTree()
            for top_key, top_body in fixed_column_json.items():
                single_tree = construct_feature_tree_for_fixed_json({top_key: top_body})
                single_tree = tag_feature_tree(single_tree)
                per_file_trees[str(top_key)] = single_tree
                for idx_node in getattr(single_tree.index_tree.root, "children", []) or []:
                    merged_index_tree.add_index(idx_node)
                    for b_node in getattr(idx_node, "body", []) or []:
                        merged_body_tree.add_deep(b_node)
            ho_tree = FeatureTree(index_tree=merged_index_tree, body_tree=merged_body_tree)
        else:
            ho_tree = construct_feature_tree_for_fixed_json({})
        ho_tree = tag_feature_tree(ho_tree)
        row_json = ho_tree.__json__()
        column_json = ho_tree.__json_column__()

        target_dirs = []
        if conversation_id:
            history_dir = os.path.join("history", str(conversation_id).strip())
            target_dirs.append(history_dir)
        target_dirs.append(cache_dir)

        for target_dir in target_dirs:
            os.makedirs(target_dir, exist_ok=True)
            with open(os.path.join(target_dir, "temp.fixed.json"), "w", encoding="utf-8") as f:
                json.dump(fixed_column_json, f, ensure_ascii=False, indent=4)

            # Clear then rewrite canonical outputs.
            with open(os.path.join(target_dir, "temp.column.json"), "w", encoding="utf-8") as f:
                json.dump({}, f, ensure_ascii=False, indent=2)

            with open(os.path.join(target_dir, "temp1.json"), "w", encoding="utf-8") as f:
                json.dump(row_json, f, ensure_ascii=False, indent=4)
            with open(os.path.join(target_dir, "temp.column.json"), "w", encoding="utf-8") as f:
                json.dump(column_json, f, ensure_ascii=False, indent=4)

            # Build per-file runtime artifacts from top-level fixed dict.
            artifact_manifest = {}
            if isinstance(fixed_column_json, dict):
                logger.info(f"[save_views] fixed top keys: {list(fixed_column_json.keys())}")
                for top_key, top_body in fixed_column_json.items():
                    single_tree = per_file_trees.get(str(top_key))
                    if single_tree is None:
                        single_tree = tag_feature_tree(
                            construct_feature_tree_for_fixed_json({top_key: top_body})
                        )
                    prefix = save_named_runtime_artifacts(single_tree, target_dir, top_key)
                    artifact_manifest[str(top_key)] = {
                        "prefix": prefix,
                        "pkl": f"{prefix}.pkl",
                        "txt": f"{prefix}.txt",
                        "embedding": f"{prefix}.embedding.json",
                        "meta": _build_artifact_meta(top_key, single_tree, top_body=top_body),
                    }
                    logger.info(
                        f"[save_views] artifact key={top_key}, prefix={prefix}, target_dir={target_dir}"
                    )
            with open(os.path.join(target_dir, "temp.artifacts.json"), "w", encoding="utf-8") as f:
                json.dump(artifact_manifest, f, ensure_ascii=False, indent=2)
            logger.info(
                f"[save_views] manifest written target_dir={target_dir}, count={len(artifact_manifest)}"
            )

        return True, "ok"
    except Exception as e:
        logger.error(f"rebuild_feature_views_from_json failed: {e}")
        return False, str(e)


def save_conversation_history(conversation_id, chat_history):
    """
    保存对话历史到文件
    
    参数:
    conversation_id: 对话ID
    chat_history: 对话历史，格式为[{"role": "user", "content": "问题"}, {"role": "assistant", "content": "回答"}, ...]
    """
    if not conversation_id:
        print(f"[DEBUG] save_conversation_history: conversation_id 为空")
        return False
    
    try:
        # 获取当前工作目录和绝对路径
        current_dir = os.getcwd()
        print(f"[DEBUG] save_conversation_history: 当前工作目录: {current_dir}")
        
        # 确保历史目录存在
        history_dir = os.path.join("history", conversation_id)
        history_dir_abs = os.path.abspath(history_dir)
        print(f"[DEBUG] save_conversation_history: 历史目录(相对): {history_dir}")
        print(f"[DEBUG] save_conversation_history: 历史目录(绝对): {history_dir_abs}")
        
        os.makedirs(history_dir, exist_ok=True)
        print(f"[DEBUG] save_conversation_history: 目录创建/存在: {os.path.exists(history_dir)}")
        
        # 保存对话历史到文件
        history_file = os.path.join(history_dir, "chat_history.json")
        history_file_abs = os.path.abspath(history_file)
        print(f"[DEBUG] save_conversation_history: 保存到文件(相对): {history_file}")
        print(f"[DEBUG] save_conversation_history: 保存到文件(绝对): {history_file_abs}")
        print(f"[DEBUG] save_conversation_history: 消息数量 {len(chat_history)}")
        print(f"[DEBUG] save_conversation_history: 消息内容 {chat_history}")
        
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(chat_history, f, ensure_ascii=False, indent=2)
        
        # 验证文件是否真的被写入
        file_exists = os.path.exists(history_file)
        file_size = os.path.getsize(history_file) if file_exists else 0
        print(f"[DEBUG] save_conversation_history: 文件写入后存在: {file_exists}, 大小: {file_size} 字节")
        
        if not file_exists:
            print(f"[DEBUG] save_conversation_history: 警告！文件写入后不存在！")
            return False
        
        print(f"[DEBUG] save_conversation_history: 保存成功")
        return True
    except Exception as e:
        print(f"[DEBUG] save_conversation_history: 保存失败 - {e}")
        import traceback
        traceback.print_exc()
        logger.error(f"保存对话历史失败: {e}")
        return False


def load_conversation_history(conversation_id):
    """
    从文件加载对话历史
    
    参数:
    conversation_id: 对话ID
    
    返回:
    对话历史，格式为[{"role": "user", "content": "问题"}, {"role": "assistant", "content": "回答"}, ...]
    """
    if not conversation_id:
        print(f"[DEBUG] load_conversation_history: conversation_id 为空")
        return []
    
    try:
        # 获取当前工作目录和绝对路径
        current_dir = os.getcwd()
        print(f"[DEBUG] load_conversation_history: 当前工作目录: {current_dir}")
        
        history_file = os.path.join("history", conversation_id, "chat_history.json")
        history_file_abs = os.path.abspath(history_file)
        print(f"[DEBUG] load_conversation_history: 加载文件(相对): {history_file}")
        print(f"[DEBUG] load_conversation_history: 加载文件(绝对): {history_file_abs}")
        
        file_exists = os.path.exists(history_file)
        print(f"[DEBUG] load_conversation_history: 文件存在(相对路径): {file_exists}")
        
        # 也尝试绝对路径
        file_exists_abs = os.path.exists(history_file_abs)
        print(f"[DEBUG] load_conversation_history: 文件存在(绝对路径): {file_exists_abs}")
        
        # 列出目录内容，看看实际有什么文件
        history_dir = os.path.join("history", conversation_id)
        if os.path.exists(history_dir):
            dir_contents = os.listdir(history_dir)
            print(f"[DEBUG] load_conversation_history: 目录内容: {dir_contents}")
        
        if file_exists:
            with open(history_file, 'r', encoding='utf-8') as f:
                chat_history = json.load(f)
            print(f"[DEBUG] load_conversation_history: 加载成功，消息数量: {len(chat_history) if isinstance(chat_history, list) else 0}")
            print(f"[DEBUG] load_conversation_history: 消息内容: {chat_history}")
            return chat_history if isinstance(chat_history, list) else []
        else:
            print(f"[DEBUG] load_conversation_history: 文件不存在")
            return []
    except Exception as e:
        print(f"[DEBUG] load_conversation_history: 加载失败 - {e}")
        import traceback
        traceback.print_exc()
        logger.error(f"加载对话历史失败: {e}")
        return []
